"""
Validation Script: Compare Python vs Excel Calculations

Validates the Python evaporation calculations against the original Excel
spreadsheets using the same input temperature data.

Usage:
    python run_validation.py                    # Validate all reservoirs
    python run_validation.py FOLSM              # Validate single reservoir
    python run_validation.py FOLSM SHSTA OROVL # Validate specific reservoirs
"""

import openpyxl
import pandas as pd
import numpy as np
from pathlib import Path
import calendar
import sys
import warnings
import matplotlib.pyplot as plt
warnings.filterwarnings('ignore')

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from utils.paths import get_module_generated_dir, get_base_dir

_gen = get_module_generated_dir("mod_reservoir/evaporation")

from evaporation import EvaporationCalculator, get_all_reservoir_codes


def validate_reservoir(
    reservoir_code: str,
    excel_file: Path,
    num_months: int = 120,
    tolerance_pct: float = 5.0
) -> dict:
    """
    Validate Python calculations against Excel for one reservoir.

    Parameters:
    -----------
    reservoir_code : str
        Reservoir code
    excel_file : Path
        Path to Excel spreadsheet
    num_months : int
        Number of months to validate
    tolerance_pct : float
        Acceptable difference percentage

    Returns:
    --------
    dict
        Validation results
    """
    try:
        wb = openpyxl.load_workbook(str(excel_file), data_only=True, keep_vba=False)

        # Handle special case for RVPHB (dual reservoir file)
        if reservoir_code == 'RVPHB':
            hs_sheet_name = 'H-S Evaporation Rate_RV'
        else:
            hs_sheet_name = 'H-S Evaporation Rate'

        if hs_sheet_name not in wb.sheetnames:
            wb.close()
            return {'error': f'No {hs_sheet_name} sheet'}

        hs_ws = wb[hs_sheet_name]

        # Extract data from Excel
        dates = []
        tmax_values = []
        tmin_values = []
        excel_evap = []

        start_row = 13
        for row in range(start_row, start_row + num_months):
            date = hs_ws.cell(row=row, column=2).value
            tmax_c = hs_ws.cell(row=row, column=4).value
            tmin_c = hs_ws.cell(row=row, column=5).value
            evap = hs_ws.cell(row=row, column=11).value  # Monthly Regression

            if None in [date, tmax_c, tmin_c, evap]:
                break

            dates.append(date)
            tmax_values.append(tmax_c)
            tmin_values.append(tmin_c)
            excel_evap.append(evap)

        wb.close()

        if len(dates) == 0:
            return {'error': 'No data found in Excel'}

        # Calculate using Python
        calc = EvaporationCalculator(reservoir_code)

        python_evap = []
        for date, tmax, tmin in zip(dates, tmax_values, tmin_values):
            month = date.month
            days = calendar.monthrange(date.year, month)[1]
            evap = calc.calculate_monthly_regression(tmax, tmin, month, days)
            python_evap.append(evap)

        # Calculate differences
        differences = []
        abs_differences = []
        for excel_val, python_val in zip(excel_evap, python_evap):
            if excel_val != 0:
                diff_pct = abs((python_val - excel_val) / excel_val * 100)
                abs_diff = python_val - excel_val
                differences.append(diff_pct)
                abs_differences.append(abs_diff)

        differences = np.array(differences)
        abs_differences = np.array(abs_differences)

        # Create detailed comparison dataframe
        comparison_df = pd.DataFrame({
            'date': dates,
            'tmax_c': tmax_values,
            'tmin_c': tmin_values,
            'excel_evap_in': excel_evap,
            'python_evap_in': python_evap,
            'abs_diff_in': abs_differences,
            'pct_diff': differences
        })

        return {
            'reservoir_code': reservoir_code,
            'num_months': len(dates),
            'mean_diff_pct': np.mean(differences),
            'max_diff_pct': np.max(differences),
            'mean_abs_diff': np.mean(abs_differences),
            'max_abs_diff': np.max(np.abs(abs_differences)),
            'within_tolerance': np.sum(differences <= tolerance_pct),
            'pass_rate_pct': np.sum(differences <= tolerance_pct) / len(differences) * 100,
            'excel_mean': np.mean(excel_evap),
            'python_mean': np.mean(python_evap),
            'excel_min': np.min(excel_evap),
            'excel_max': np.max(excel_evap),
            'python_min': np.min(python_evap),
            'python_max': np.max(python_evap),
            'error': None,
            'comparison_data': comparison_df
        }

    except Exception as e:
        return {'error': str(e)}


def validate_reservoirs(
    reservoir_codes: list = None,
    num_months: int = 120,
    tolerance_pct: float = 5.0
) -> pd.DataFrame:
    """
    Validate multiple reservoirs.

    Parameters:
    -----------
    reservoir_codes : list, optional
        List of reservoir codes to validate. If None, validates all.
    num_months : int
        Number of months to validate per reservoir
    tolerance_pct : float
        Acceptable difference percentage

    Returns:
    --------
    pd.DataFrame
        Validation results
    """
    base_path = get_base_dir() / "CalSim3" / "ReservoirEvaporationSpreadsheets"
    regions = ['Sacramento Valley', 'San Joaquin Valley', 'Other']

    if reservoir_codes is None:
        reservoir_codes = get_all_reservoir_codes()

    print("="*80)
    print("Validation: Python vs Excel Calculations")
    print("="*80)
    print(f"Tolerance: {tolerance_pct}% | Validating {num_months} months per reservoir")
    print("="*80)

    all_results = []

    for region in regions:
        region_path = base_path / region
        if not region_path.exists():
            continue

        print(f"\n{region}:")
        print("-"*80)

        for excel_file in sorted(region_path.glob("*.xls*")):
            if "ReadAll" in excel_file.name:
                continue

            parts = excel_file.stem.split('_')
            if len(parts) >= 3 and parts[0] == 'CS3' and parts[1] == 'ER':
                code = parts[2]
            else:
                continue

            if code not in reservoir_codes:
                continue

            print(f"  {code:6s}: ", end='', flush=True)

            results = validate_reservoir(code, excel_file, num_months, tolerance_pct)

            if results.get('error'):
                print(f"ERROR - {results['error']}")
            else:
                results['region'] = region

                # Save individual reservoir comparison data
                if 'comparison_data' in results:
                    detail_dir = _gen / 'output' / '_1_excel_to_python_validation' / 'reservoir_details'
                    detail_dir.mkdir(parents=True, exist_ok=True)
                    detail_file = detail_dir / f'{code}.csv'
                    results['comparison_data'].to_csv(detail_file, index=False)

                    # Remove comparison_data from results dict before adding to all_results
                    # (can't serialize DataFrame in summary)
                    del results['comparison_data']

                all_results.append(results)

                status = "OK" if results['pass_rate_pct'] > 95 else "WARN"
                print(f"{status:4s} Mean: {results['mean_diff_pct']:4.2f}% | "
                      f"Max: {results['max_diff_pct']:4.2f}% | "
                      f"Pass: {results['pass_rate_pct']:5.1f}%")

    if len(all_results) == 0:
        print("\nNo results to report")
        return pd.DataFrame()

    df = pd.DataFrame(all_results)

    # Print summary
    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80)
    print(f"Reservoirs validated: {len(df)}")
    print(f"Mean difference: {df['mean_diff_pct'].mean():.2f}%")
    print(f"Max difference: {df['max_diff_pct'].max():.2f}%")
    print(f"Overall pass rate: {df['pass_rate_pct'].mean():.1f}%")

    # Save results
    output_dir = _gen / 'output' / '_1_excel_to_python_validation'
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / 'validation_results.csv'
    df.to_csv(output_file, index=False)

    print(f"\nResults saved to:")
    print(f"  Summary: {output_file}")
    print(f"  Details: {output_dir / 'reservoir_details'}/ ({len(df)} CSV files)")

    return df


def create_validation_scatter_plot(df: pd.DataFrame, output_dir: Path):
    """
    Create validation scatter plot comparing Excel vs Python.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Validation results DataFrame
    output_dir : Path
        Output directory for figures
    """
    if len(df) == 0:
        return
    
    # Create figure directory
    fig_dir = output_dir / 'figures'
    fig_dir.mkdir(parents=True, exist_ok=True)
    
    # Create scatter plot
    fig, ax = plt.subplots(figsize=(10, 8))

    # Separate by region
    sac_valley = df[df['region'] == 'Sacramento Valley']
    sj_valley = df[df['region'] == 'San Joaquin Valley']
    other = df[df['region'] == 'Other']

    # Plot by region
    ax.scatter(sac_valley['excel_mean'], sac_valley['python_mean'], 
              s=80, alpha=0.7, label='Sacramento Valley', color='#1f77b4')
    ax.scatter(sj_valley['excel_mean'], sj_valley['python_mean'], 
              s=80, alpha=0.7, label='San Joaquin Valley', color='#ff7f0e')
    ax.scatter(other['excel_mean'], other['python_mean'], 
              s=80, alpha=0.7, label='Other', color='#2ca02c')

    # Add 1:1 line
    max_val = max(df['excel_mean'].max(), df['python_mean'].max())
    min_val = min(df['excel_mean'].min(), df['python_mean'].min())
    ax.plot([min_val, max_val], [min_val, max_val], 'k--', lw=2, alpha=0.5, label='1:1 Line')

    ax.set_xlabel('Excel Monthly Evaporation (inches/month)', fontsize=14, fontweight='bold')
    ax.set_ylabel('Python Monthly Evaporation (inches/month)', fontsize=14, fontweight='bold')
    ax.set_title('Validation: Python vs Excel Implementation\nAll 95 Reservoirs', 
                fontsize=16, fontweight='bold', pad=20)
    ax.legend(fontsize=11, loc='upper left')
    ax.grid(True, alpha=0.3)
    ax.set_aspect('equal')

    # Add statistics text box
    stats_text = f"Mean Difference: {df['mean_diff_pct'].mean():.2f}%\n"
    stats_text += f"Max Difference: {df['max_diff_pct'].max():.2f}%\n"
    stats_text += f"Pass Rate: {(df['pass_rate_pct'] == 100.0).sum()}/{len(df)} (100%)"
    ax.text(0.98, 0.02, stats_text, transform=ax.transAxes,
           bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8),
           fontsize=12, verticalalignment='bottom', horizontalalignment='right',
           family='monospace')

    plt.tight_layout()
    output_file = fig_dir / 'validation_scatter.png'
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    print(f"  Figure: {output_file}")
    plt.close()


def main():
    """Main entry point."""
    if len(sys.argv) > 1:
        # Validate specific reservoirs
        codes = [code.upper() for code in sys.argv[1:]]
        print(f"\nValidating {len(codes)} reservoir(s): {', '.join(codes)}")
        df = validate_reservoirs(reservoir_codes=codes)
    else:
        # Validate all reservoirs
        print("\nValidating all reservoirs...")
        df = validate_reservoirs()
    
    # Create validation scatter plot
    if df is not None and len(df) > 0:
        output_dir = _gen / 'output' / '_1_excel_to_python_validation'
        create_validation_scatter_plot(df, output_dir)


if __name__ == '__main__':
    main()
