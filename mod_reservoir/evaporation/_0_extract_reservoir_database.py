"""
Reservoir Parameter Database Extractor
======================================
Extracts evaporation calculation parameters (location, Ra values, calibration
factors) from the CalSim 3.0 reservoir Excel spreadsheets and writes them to
reference/reservoir_parameters.json, which is tracked in the repository and
used by all downstream evaporation scripts.

Run --extract whenever the Excel spreadsheets are updated. The default mode
(no flag) reads the existing reference JSON and prints a summary.

Inputs
------
- CalSim3/ReservoirEvaporationSpreadsheets/ (reservoir Excel workbooks)

Outputs
-------
- mod_reservoir/evaporation/reference/reservoir_parameters.json

Dependencies
------------
- utils/paths.py  (data-dir resolution)

Usage
-----
    # Re-extract from Excel spreadsheets -> overwrites reference/reservoir_parameters.json
    python _0_extract_reservoir_database.py --extract

    # Print database summary from reference/reservoir_parameters.json (default)
    python _0_extract_reservoir_database.py
"""

import argparse
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from utils.paths import get_module_generated_dir, get_base_dir

_gen             = get_module_generated_dir("mod_reservoir/evaporation")
_spreadsheets_dir = get_base_dir() / "CalSim3" / "ReservoirEvaporationSpreadsheets"


class ReservoirDatabase:
    """
    Database of reservoir parameters for evaporation calculations.

    Provides access to location, Ra values, and calibration factors for
    all CalSim 3.0 reservoirs.
    """

    def __init__(self, params_file: Optional[str] = None):
        """
        Initialize the reservoir database.

        Parameters:
        -----------
        params_file : str, optional
            Path to reservoir_parameters.json. If None, uses default location.
        """
        if params_file is None:
            params_file = Path(__file__).resolve().parent / 'reference' / 'reservoir_parameters.json'

        with open(params_file, 'r') as f:
            self._params = json.load(f)

    def get_reservoir_codes(self, region: Optional[str] = None) -> List[str]:
        """
        Get list of reservoir codes.

        Parameters:
        -----------
        region : str, optional
            Filter by region ('Sacramento Valley', 'San Joaquin Valley')

        Returns:
        --------
        List[str]
            List of reservoir codes
        """
        if region is None:
            return sorted(list(self._params.keys()))
        else:
            return sorted([code for code, params in self._params.items()
                          if params['region'] == region])

    def get_reservoir_info(self, reservoir_code: str) -> Dict:
        """
        Get all parameters for a reservoir.

        Parameters:
        -----------
        reservoir_code : str
            Reservoir code (e.g., 'ALDER', 'FOLSM')

        Returns:
        --------
        Dict
            Complete parameter dictionary
        """
        if reservoir_code not in self._params:
            raise ValueError(f"Reservoir '{reservoir_code}' not found in database. "
                           f"Available: {', '.join(sorted(self._params.keys())[:10])}...")
        return self._params[reservoir_code]

    def get_location(self, reservoir_code: str) -> Dict[str, float]:
        """
        Get reservoir location (latitude, longitude, elevation).

        Parameters:
        -----------
        reservoir_code : str
            Reservoir code

        Returns:
        --------
        Dict with keys: 'latitude', 'longitude', 'elevation_ft'
        """
        params = self.get_reservoir_info(reservoir_code)
        return {
            'latitude': params['latitude'],
            'longitude': params['longitude'],
            'elevation_ft': params['elevation_ft']
        }

    def get_ra_monthly(self, reservoir_code: str) -> Dict[int, float]:
        """
        Get monthly Ra values (extraterrestrial radiation in mm/day).

        Parameters:
        -----------
        reservoir_code : str
            Reservoir code

        Returns:
        --------
        Dict[int, float]
            Ra values by month number (1-12)
        """
        params = self.get_reservoir_info(reservoir_code)
        # Convert keys from strings to ints if needed
        ra = params['ra_monthly']
        return {int(k): v for k, v in ra.items()}

    def get_monthly_calibration(self, reservoir_code: str) -> Dict[int, Dict]:
        """
        Get monthly calibration factors.

        Parameters:
        -----------
        reservoir_code : str
            Reservoir code

        Returns:
        --------
        Dict[int, Dict]
            Calibration parameters by month. Each month contains:
            - 'factor': Multiplicative factor
            - 'slope': Regression slope
            - 'intercept': Regression intercept
        """
        params = self.get_reservoir_info(reservoir_code)
        cal = params['monthly_calibration']
        return {int(k): v for k, v in cal.items()}

    def get_annual_calibration(self, reservoir_code: str) -> Dict[str, float]:
        """
        Get annual calibration factors.

        Parameters:
        -----------
        reservoir_code : str
            Reservoir code

        Returns:
        --------
        Dict with keys: 'factor', 'slope', 'intercept'
        """
        params = self.get_reservoir_info(reservoir_code)
        return params['annual_calibration']

    def find_nearest_reservoir(self, latitude: float, longitude: float,
                                region: Optional[str] = None) -> str:
        """
        Find the nearest reservoir to a given location.

        Parameters:
        -----------
        latitude : float
            Latitude in decimal degrees
        longitude : float
            Longitude in decimal degrees
        region : str, optional
            Limit search to specific region

        Returns:
        --------
        str
            Reservoir code of nearest reservoir
        """
        min_dist = float('inf')
        nearest_code = None

        codes = self.get_reservoir_codes(region)

        for code in codes:
            loc = self.get_location(code)
            # Simple Euclidean distance (good enough for small areas)
            dist = np.sqrt((loc['latitude'] - latitude)**2 +
                          (loc['longitude'] - longitude)**2)

            if dist < min_dist:
                min_dist = dist
                nearest_code = code

        return nearest_code

    def get_region_summary(self) -> Dict[str, int]:
        """
        Get count of reservoirs by region.

        Returns:
        --------
        Dict[str, int]
            Count of reservoirs per region
        """
        summary = {}
        for params in self._params.values():
            region = params['region']
            summary[region] = summary.get(region, 0) + 1
        return summary

    def export_to_csv(self, output_file: str):
        """
        Export reservoir database to CSV format.

        Parameters:
        -----------
        output_file : str
            Output CSV file path
        """
        import csv

        with open(output_file, 'w', newline='') as f:
            writer = csv.writer(f)

            # Header
            writer.writerow([
                'ReservoirCode', 'Region', 'Latitude', 'Longitude',
                'Elevation_ft', 'AnnualFactor', 'AnnualSlope', 'AnnualIntercept'
            ])

            # Data
            for code in sorted(self._params.keys()):
                params = self._params[code]
                annual = params['annual_calibration']

                writer.writerow([
                    code,
                    params['region'],
                    params['latitude'],
                    params['longitude'],
                    params['elevation_ft'],
                    annual['factor'],
                    annual['slope'],
                    annual['intercept']
                ])


def find_nearest_weather_file(reservoir_code: str,
                               weather_dir: str = '../_00_Data/Historical_Climate_LTO/1_Historical') -> Optional[str]:
    """
    Find the nearest weather data file for a reservoir.

    Weather files are named: data_LAT_LON

    Parameters:
    -----------
    reservoir_code : str
        Reservoir code
    weather_dir : str
        Directory containing weather data files

    Returns:
    --------
    str or None
        Path to nearest weather file, or None if not found
    """
    db = ReservoirDatabase()
    loc = db.get_location(reservoir_code)

    res_lat = loc['latitude']
    res_lon = loc['longitude']

    # Find all weather files
    weather_path = Path(weather_dir)
    if not weather_path.exists():
        return None

    min_dist = float('inf')
    nearest_file = None

    for file in weather_path.glob('data_*'):
        # Extract lat/lon from filename: data_LAT_LON
        parts = file.name.split('_')
        if len(parts) != 3:
            continue

        try:
            file_lat = float(parts[1])
            file_lon = float(parts[2])

            # Calculate distance
            dist = np.sqrt((file_lat - res_lat)**2 + (file_lon - res_lon)**2)

            if dist < min_dist:
                min_dist = dist
                nearest_file = str(file)
        except ValueError:
            continue

    return nearest_file


def extract_reservoir_parameters(
    spreadsheets_dir=None,
    output_file=None
) -> Dict:
    """
    Extract parameters from all Excel spreadsheets and create JSON database.

    Parameters:
    -----------
    spreadsheets_dir : str
        Directory containing Excel spreadsheets organized by region
    output_file : str
        Output JSON file path

    Returns:
    --------
    Dict
        Complete parameter database
    """
    print("Extracting reservoir parameters from Excel spreadsheets...")
    print("="*80)

    base_path = Path(spreadsheets_dir) if spreadsheets_dir is not None else _spreadsheets_dir

    # Define regions and their directories
    regions = {
        'Sacramento Valley': base_path / 'Sacramento Valley',
        'San Joaquin Valley': base_path / 'San Joaquin Valley',
        'Other': base_path / 'Other'
    }

    all_params = {}

    for region_name, region_path in regions.items():
        print(f"\nProcessing {region_name}...")

        if not region_path.exists():
            print(f"  Warning: Directory not found - {region_path}")
            continue

        excel_files = list(region_path.glob('*.xls*'))
        print(f"  Found {len(excel_files)} Excel files")

        for excel_file in sorted(excel_files):
            try:
                # Skip the ReadAll file
                if 'ReadAll' in excel_file.name:
                    continue

                # Extract reservoir code from filename: CS3_ER_XXXXX_Rev2022F.xlsm
                reservoir_code = excel_file.stem.split('_')[2]

                print(f"  Extracting {reservoir_code}...", end=' ')

                wb = openpyxl.load_workbook(str(excel_file), data_only=True, keep_vba=False)

                # Handle special case for RVPHB (dual reservoir file with _RV and _PB sheets)
                if reservoir_code == 'RVPHB':
                    temp_sheet_name = 'Temperature_RV'
                    ra_sheet_name = 'Extraterrestrial Radiation_RV'
                    hs_sheet_name = 'H-S Evaporation Rate_RV'
                else:
                    temp_sheet_name = 'Temperature Data'
                    ra_sheet_name = 'Extraterrestrial Radiation'
                    hs_sheet_name = 'H-S Evaporation Rate'

                # Extract location from temperature sheet
                temp_ws = wb[temp_sheet_name]
                elevation_ft = float(temp_ws.cell(8, 3).value)   # C8
                latitude = float(temp_ws.cell(9, 3).value)       # C9
                longitude = float(temp_ws.cell(10, 3).value)     # C10

                # Extract Ra monthly values from Extraterrestrial Radiation sheet (mm/day)
                ra_ws = wb[ra_sheet_name]
                ra_monthly = {}
                # Rows 6-17 contain Oct-Sep, column E (5) has mm/day
                month_order = [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]  # Oct-Sep
                for i, month in enumerate(month_order):
                    ra_monthly[month] = float(ra_ws.cell(6 + i, 5).value)

                # Extract monthly calibration from H-S Evaporation Rate sheet
                # Row 12 is header, rows 13-24 contain months 10-9
                # Column T (20) = Month No., U (21) = Factor, V (22) = Slope, W (23) = Intercept
                # Row 25 = Annual calibration (T is None)
                hs_ws = wb[hs_sheet_name]
                monthly_calibration = {}

                # Extract all 12 months from rows 13-24
                for row in range(13, 25):
                    month = hs_ws.cell(row, 20).value
                    if month is not None:
                        monthly_calibration[int(month)] = {
                            'factor': float(hs_ws.cell(row, 21).value),
                            'slope': float(hs_ws.cell(row, 22).value),
                            'intercept': float(hs_ws.cell(row, 23).value)
                        }

                # Verify we got all 12 months
                if len(monthly_calibration) != 12:
                    raise ValueError(f"Expected 12 months, got {len(monthly_calibration)}")

                # Extract annual calibration from row 25 (where column T is None)
                annual_calibration = {
                    'factor': float(hs_ws.cell(25, 21).value),
                    'slope': float(hs_ws.cell(25, 22).value),
                    'intercept': float(hs_ws.cell(25, 23).value)
                }

                # Extract adjustment factor from M2 and P2
                m2_value = hs_ws.cell(2, 13).value  # M2
                if m2_value and "No data" in str(m2_value):
                    adjustment_factor = float(hs_ws.cell(2, 16).value)  # P2
                else:
                    adjustment_factor = 1.0

                wb.close()

                # Store parameters
                all_params[reservoir_code] = {
                    'region': region_name,
                    'latitude': latitude,
                    'longitude': longitude,
                    'elevation_ft': elevation_ft,
                    'ra_monthly': ra_monthly,
                    'monthly_calibration': monthly_calibration,
                    'annual_calibration': annual_calibration,
                    'adjustment_factor': adjustment_factor
                }

                print("OK")

            except Exception as e:
                print(f"FAILED - {str(e)}")
                continue

    # Save to JSON
    _reference_dir = Path(__file__).resolve().parent / 'reference'
    output_path = Path(output_file) if output_file is not None else _reference_dir / 'reservoir_parameters.json'
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, 'w') as f:
        json.dump(all_params, f, indent=2)

    print("\n" + "="*80)
    print(f"Extracted {len(all_params)} reservoirs")
    print(f"Saved to: {output_path}")

    # Summary by region
    summary = {}
    for params in all_params.values():
        region = params['region']
        summary[region] = summary.get(region, 0) + 1

    print("\nReservoirs by region:")
    for region, count in sorted(summary.items()):
        print(f"  {region:25s}: {count:3d}")

    return all_params


def main():
    """
    Example usage of the reservoir database.
    """
    db = ReservoirDatabase()

    print("Reservoir Database Summary")
    print("="*80)

    # Region summary
    summary = db.get_region_summary()
    print("\nReservoirs by Region:")
    for region, count in summary.items():
        print(f"  {region:25s}: {count:3d} reservoirs")

    print(f"\nTotal: {sum(summary.values())} reservoirs")

    # Example: Get info for a specific reservoir
    print("\n" + "="*80)
    print("Example: Folsom Reservoir (FOLSM)")
    print("="*80)

    loc = db.get_location('FOLSM')
    print("\nLocation:")
    print(f"  Latitude:  {loc['latitude']:.4f}deg")
    print(f"  Longitude: {loc['longitude']:.4f}deg")
    print(f"  Elevation: {loc['elevation_ft']:.0f} ft")

    ra = db.get_ra_monthly('FOLSM')
    print("\nRa values (mm/day) by month:")
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    for i, name in enumerate(month_names, 1):
        print(f"  {name}: {ra[i]:6.2f}", end='  ')
        if i % 4 == 0:
            print()

    annual = db.get_annual_calibration('FOLSM')
    print("\nAnnual calibration:")
    print(f"  Factor:    {annual['factor']:.6f}")
    print(f"  Slope:     {annual['slope']:.6f}")
    print(f"  Intercept: {annual['intercept']:.6f}")

    # Find nearest weather file
    print("\n" + "="*80)
    print("Nearest Weather Files")
    print("="*80)

    for code in ['FOLSM', 'SHSTA', 'OROVL']:
        weather_file = find_nearest_weather_file(code)
        if weather_file:
            filename = Path(weather_file).name
            print(f"\n{code:6s}: {filename}")
        else:
            print(f"\n{code:6s}: No weather file found")

    # Export to CSV
    print("\n" + "="*80)
    output_dir = Path(__file__).resolve().parent / 'reference'
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_file = output_dir / 'reservoir_database.csv'
    db.export_to_csv(str(csv_file))
    print(f"Database exported to: {csv_file}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Reservoir Parameter Database - Extract or query CalSim 3.0 evaporation parameters'
    )
    parser.add_argument(
        '--extract',
        action='store_true',
        help='Extract parameters from Excel spreadsheets and create JSON database'
    )

    args = parser.parse_args()

    if args.extract:
        # Extract parameters from Excel files
        extract_reservoir_parameters()
    else:
        # Show database summary
        main()
