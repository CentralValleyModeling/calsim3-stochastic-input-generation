"""
Feather River Minimum Instream Flow (MINFLOWFEATHER)
====================================================
Computes MINFLOWFEATHER (Feather River minimum instream flow, cfs) by
replicating the logic in MIF_FEATHER_Logic_Reconstruction.xlsx.

--------------------------------------------------------------------------
LOGIC SUMMARY
--------------------------------------------------------------------------
The schedule is assigned on an APRIL-to-MARCH cycle ("CY" = calendar year of
April). For each CY N the schedule covers Apr(N)-Sep(N) and Oct(N)-Mar(N+1).

Three schedule inputs are evaluated each CY:

  1. Apr-Jul runoff (CY N)  - sum of Apr+May+Jun+Jul UNIMP_OROV (TAF)
        Dry  if  Apr_Jul < APR_JUL_THRESH  (1068.1 TAF)
        Wet  otherwise

  2. Annual WY (N) unimpaired Oroville inflow  (TAF, Oct-Sep water year)
        Deficiency  if  Annual_WY < ANN_THRESH  (1245.99 TAF)
        Normal      otherwise

  3. 2-year rolling average of Annual WY (TAF)
        Deficiency  if  TwoYrAvg_WY < TWOY_THRESH  (3170.64 TAF)
        Normal      otherwise

  Schedule selection:
        driest  if  Annual Deficiency  OR  2Y Deficiency
        wet     elif  Apr-Jul Wet
        dry     else  (Apr-Jul Dry, no Deficiency)

  Monthly flow schedules (cfs):
        Month:    Apr  May  Jun  Jul  Aug  Sep  Oct  Nov  Dec  Jan  Feb  Mar
        wet:     1000 1000 1000 1000 1000 1000 1700 1700 1700 1700 1700 1700
        dry:     1000 1000 1000 1000 1000 1000 1200 1200 1200 1200 1200 1000
        driest:   750  750  750  750  750  750   900  900  900  900  900  750

  Schedule-to-date assignment:
        For a date (Year=Y, Month=M):  CY = Y  if  M >= 4  else  Y - 1

--------------------------------------------------------------------------
THRESHOLDS  (calibrated on 1921-2021 historical record, fixed constants)
--------------------------------------------------------------------------
  APR_JUL_THRESH = 0.55 x 1942           =  1068.10  TAF  (hardcoded in xlsx)
  ANN_THRESH     = 0.28492 x 4373.11     =  1245.99  TAF
  TWOY_THRESH    = 0.72503 x 4373.11     =  3170.64  TAF
  Mean annual UNIMP_OROV (1921-2021)     =  4373.11  TAF

--------------------------------------------------------------------------
INPUTS / OUTPUTS
--------------------------------------------------------------------------
  UNIMP_OROV inflow sources by run:

  Validation  <-  CalSim SV DSS  __calsim_sv_default__.dss
    UNIMP_OROV/FLOW-UNIMPAIRED CalSim baseline monthly unimpaired inflow,
    read via utils.dss_io.  The spreadsheet still supplies the calibrated
    thresholds and the Value_Spreadsheet reference column.

  Product A  <-  rim-inflow Product A CSV  _riminflow_productA_1972_2018.csv
    Quantile-mapped UNIMP_OROV (WY 1972-2018).  Leading months that predate
    the first classifiable cycle are filled from the baseline MINFLOWFEATHER.

  Product B  <-  rim-inflow Product B chunks  UNIMP_OROV_qmo_n01..n10.csv
    Quantile-mapped UNIMP_OROV; the WY-1921 rolling-average seed comes from
    the CalSim SV DSS.

  Output 1 - Validation (WY 1922-2021)
    Classifies with historical CS3 data; dates from Oct 1921 - Sep 2021.
    Compares Python output against:
      - CalSim3 reference CSV (Value_CS3)
      - Spreadsheet Main_WA cols G-S, rows 4-104 (CY 1921-2021) (Value_Spreadsheet)
    Wide format: Year, Month, Value_CS3 (CalSim3 reference),
                 Value_Spreadsheet (original xlsx computation),
                 Value_Python (this script).
    Written to:
      output/_1_min_flow_feather/_minflowfeather_validation_1922_2021.csv
    Primary comparison: Value_Spreadsheet (MIF_FEATHER_Logic_Reconstruction.xlsx
      Main_WA cols G-S, rows 4-104, CY 1921-2021)

  Output 2 - Product A (WY 1972-2018)
    Classifies with the quantile-mapped Product A UNIMP_OROV; dates from
    Oct 1971 - Sep 2018.  The leading months (Oct 1971 - Mar 1972), whose
    schedule cycle needs WY 1971, are filled from the CalSim baseline
    MINFLOWFEATHER.
    Written to:
      output/_product_a_validation/_minflowfeather_productA_1972_2018.csv

  Output 3 - Product B (10 chunks, WY 1922-2021 each)
    Classifies with each chunk's own qmap_postAdj data plus CS3 WY 1921 seed
    for 2-year rolling average priming.
    Written to:
      output/_product_b_final/_minflowfeather_productB_qmo_n01.csv
      output/_product_b_final/_minflowfeather_productB_qmo_n02.csv
      ... through n10.csv

Reference:
  Feather River Levee Repair Project, Draft EIR - DWR / TRLIA
  https://cms9files1.revize.com/trlia/Environmental%20Docs/5.4%20Fisheries.pdf

Usage
-----
    python mod_other/instream_flows/_1_min_flow_feather.py --product validation  # Output 1
    python mod_other/instream_flows/_1_min_flow_feather.py --product A           # Output 2
    python mod_other/instream_flows/_1_min_flow_feather.py --product B           # Output 3
"""

import os
import sys
from pathlib import Path
import pandas as pd
import numpy as np

# Add repo root to path for utils imports
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from utils.paths import get_base_dir, get_module_generated_dir
from utils import dss_io

# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------
_SCRIPT_DIR = Path(__file__).resolve().parent
_rim_gen    = get_module_generated_dir("mod_hydrology/rim_inflow")
_gen        = get_module_generated_dir("mod_other/instream_flows")

# CalSim SV DSS: source of the validation input UNIMP_OROV, the Product B
# WY-1921 rolling-average seed, and the baseline MINFLOWFEATHER used to fill
# Product A leading months and for the validation Value_CS3 column.
DSS_FILE         = get_base_dir() / "CalSim3" / "__calsim_sv_default__.dss"
DSS_READ_START   = "1915-01-31"
DSS_READ_END     = "2021-12-31"
UNIMP_OROV_KEY = ("UNIMP_OROV", "FLOW-UNIMPAIRED")        # baseline inflow input
MINFLOW_KEY    = ("MINFLOWFEATHER", "FLOW-MIN-REQUIRED")  # CS3 actual (Value_CS3)

# Product A quantile-mapped UNIMP_OROV (WY 1972-2018), from the rim-inflow
# Product A CSV (Part B == UNIMP_OROV).
PATH_PRODUCT_A_CSV = (_rim_gen / "output" / "_2_qmap_historical_validation"
                      / "_product_a_validation"
                      / "_riminflow_productA_1972_2018.csv")

# Product B directory (10 chunk CSVs)
PATH_UNIMP_OROV_PB_DIR = _rim_gen / "output" / "_3_qmap_product_b"

# Spreadsheet with original computed schedule (for 3-way validation comparison)
PATH_XLSX = _gen / "term_development" / "MINFLOWFEATHER" / "MIF_FEATHER_Logic_Reconstruction.xlsx"

# All outputs go to a single folder
OUT_DIR = _gen / "output" / "_1_min_flow_feather"

# Output 1: wide-format validation (WY 1922-2021, capped by available historical data)
OUT_VALIDATION = OUT_DIR / "_minflowfeather_validation_1922_2021.csv"

# Output 2: Product A (WY 1972-2018, quantile-mapped input)
OUT_PRODUCT_A_DIR = _gen / "output" / "_product_a_validation"
OUT_PRODUCT_A = OUT_PRODUCT_A_DIR / "_minflowfeather_productA_1972_2018.csv"

# Output 3: Product B - one file per chunk (n01 ... n10)
OUT_PRODUCT_B_DIR = _gen / "output" / "_product_b_final"
OUT_PRODUCT_B_PATTERN = str(OUT_PRODUCT_B_DIR / "_minflowfeather_productB_qmo_n{:02d}.csv")

# -----------------------------------------------------------------------------
# Thresholds  (fixed calibrated constants from spreadsheet)
# -----------------------------------------------------------------------------
# Apr-Jul Wet/Dry threshold: formula in xlsx cell = IF(D > 0.55*1942, "Wet", "Dry")
APR_JUL_THRESH = 0.55 * 1942          # 1068.10 TAF

# Annual and 2-year deficiency thresholds (fraction x mean annual)
_ANN_FACTOR   = 0.28492126771039866
_TWOY_FACTOR  = 0.725030663907629
_MEAN_ANNUAL  = 4373.108910891089      # TAF  (AVERAGE of Annual Unimp_Orov WY 1921-2021)

ANN_THRESH  = _ANN_FACTOR  * _MEAN_ANNUAL  # 1245.99 TAF
TWOY_THRESH = _TWOY_FACTOR * _MEAN_ANNUAL  # 3170.64 TAF

# -----------------------------------------------------------------------------
# Flow schedules  (cfs),  keyed by calendar month 1-12
# -----------------------------------------------------------------------------
SCHEDULES = {
    'wet': {
        4: 1000, 5: 1000, 6: 1000, 7: 1000, 8: 1000, 9: 1000,
        10: 1700, 11: 1700, 12: 1700, 1: 1700, 2: 1700, 3: 1700,
    },
    'dry': {
        4: 1000, 5: 1000, 6: 1000, 7: 1000, 8: 1000, 9: 1000,
        10: 1200, 11: 1200, 12: 1200, 1: 1200, 2: 1200, 3: 1000,
    },
    'driest': {
        4: 750, 5: 750, 6: 750, 7: 750, 8: 750, 9: 750,
        10: 900, 11: 900, 12: 900, 1: 900, 2: 900, 3: 750,
    },
}

# Canonical WY-ordered month sequence for output sorting
_WY_MONTH_ORDER = {m: i for i, m in enumerate([10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9])}


# -----------------------------------------------------------------------------
# Core functions
# -----------------------------------------------------------------------------

def _water_year(dates: pd.DatetimeIndex) -> pd.Index:
    """Return integer water year (Oct-Dec of Y -> WY Y+1; Jan-Sep of Y -> WY Y)."""
    return pd.Index(dates.year.where(dates.month < 10, dates.year + 1))


def _select_schedule(ann_def: bool, twoy_def: bool, apjul_dry: bool) -> str:
    """Return schedule name from three boolean condition inputs."""
    if ann_def or twoy_def:
        return 'driest'
    return 'wet' if not apjul_dry else 'dry'


def compute_min_flow_feather(
    unimp_orov: pd.Series,
    ann_thresh: float = None,
    twoy_thresh: float = None,
    apjul_thresh: float = None,
) -> pd.Series:
    """
    Compute MINFLOWFEATHER (cfs) for each month in the input series.

    Parameters
    ----------
    unimp_orov : pd.Series
        Monthly unimpaired Oroville inflow (TAF), end-of-month DatetimeIndex.
        NaN rows are dropped before processing.
    ann_thresh, twoy_thresh, apjul_thresh : float, optional
        Override the module-level threshold constants.  Pass values read
        directly from the spreadsheet when running the validation, so that
        the Python classification uses exactly the same thresholds.

    Returns
    -------
    pd.Series
        Monthly MINFLOWFEATHER (cfs), DatetimeIndex (end-of-month).
        Only months for which a valid CY schedule can be determined are included.
    """
    _ann   = ANN_THRESH   if ann_thresh   is None else ann_thresh
    _twoy  = TWOY_THRESH  if twoy_thresh  is None else twoy_thresh
    _apjul = APR_JUL_THRESH if apjul_thresh is None else apjul_thresh

    data = unimp_orov.dropna().copy()

    # -- Annual WY totals (Oct-Sep) ----------------------------------------
    wy_index = _water_year(data.index)
    annual_wy: pd.Series = data.groupby(wy_index).sum()  # indexed by integer WY

    # -- 2-year rolling average (WY-based) --------------------------------
    # twoy_avg[WY] = (Annual[WY] + Annual[WY-1]) / 2; NaN for first WY
    twoy_avg: pd.Series = (annual_wy + annual_wy.shift(1)) / 2

    # -- Apr-Jul sums by calendar year ------------------------------------
    apjul_mask = data.index.month.isin([4, 5, 6, 7])
    apjul_cy: pd.Series = (
        data[apjul_mask].groupby(data[apjul_mask].index.year).sum()
    )  # indexed by integer calendar year

    # -- Determine schedule for each CY -----------------------------------
    first_cy = int(annual_wy.index[0])
    last_cy  = int(annual_wy.index[-1])

    schedule_by_cy = {}  # dict[int, str]
    for cy in range(first_cy, last_cy + 1):
        ann   = annual_wy.get(cy,   np.nan)
        twoy  = twoy_avg.get(cy,    np.nan)
        apjul = apjul_cy.get(cy,    np.nan)

        # Cannot determine schedule without Annual or Apr-Jul data
        if np.isnan(ann) or np.isnan(apjul):
            continue

        ann_def   = ann   < _ann
        twoy_def  = (not np.isnan(twoy)) and (twoy < _twoy)
        apjul_dry = apjul < _apjul

        schedule_by_cy[cy] = _select_schedule(ann_def, twoy_def, apjul_dry)

    # -- Assign monthly flow values ----------------------------------------
    # For date (Y, M):  CY = Y if M >= 4 else Y - 1
    flows = {}  # dict[pd.Timestamp, int]
    for dt in data.index:
        y, m = dt.year, dt.month
        cy = y if m >= 4 else y - 1
        sched = schedule_by_cy.get(cy)
        if sched is not None:
            flows[dt] = SCHEDULES[sched][m]

    return pd.Series(flows, name='Value', dtype=int).rename_axis('date')


# -----------------------------------------------------------------------------
# Output helpers
# -----------------------------------------------------------------------------

def to_calsim_csv(series: pd.Series, part_b: str, part_c: str) -> pd.DataFrame:
    """
    Convert a monthly Series to CalSim CSV format
    (Part B, Part C, Year, Month, Value), sorted in water-year order.
    """
    df = series.reset_index()
    df.columns = ['date', 'Value']
    df['Part B']  = part_b
    df['Part C']  = part_c
    df['Year']    = df['date'].dt.year
    df['Month']   = df['date'].dt.month
    df['_wy']     = df['Year'].where(df['Month'] < 10, df['Year'] + 1)
    df['_mo_ord'] = df['Month'].map(_WY_MONTH_ORDER)
    df = (df
          .sort_values(['_wy', '_mo_ord'])
          .drop(columns=['date', '_wy', '_mo_ord'])
          [['Part B', 'Part C', 'Year', 'Month', 'Value']]
          .reset_index(drop=True))
    return df


def _print_schedule_summary(schedule_by_cy) -> None:
    """Print a summary of schedule classification by CY."""
    totals = pd.Series(schedule_by_cy).value_counts()
    print("  Schedule distribution (by CY):")
    for sched in ['wet', 'dry', 'driest']:
        n = totals.get(sched, 0)
        print(f"    {sched:7s}: {n:3d} years")


def _load_xlsx_thresholds() -> dict:
    """
    Read the ANN_THRESH and TWOY_THRESH values directly from Main_WA in the
    spreadsheet.  These are recomputed each time the spreadsheet's source data
    changes, so reading them here keeps the validation in exact agreement.

    Main_WA layout:
      Row 0 (Excel row 1): ('Annual Thresh', factor, VALUE, ...)
      Row 1 (Excel row 2): ('2Y Thresh',     factor, VALUE, ...)
    """
    import openpyxl
    wb = openpyxl.load_workbook(PATH_XLSX, read_only=True, data_only=True)
    rows = list(wb['Main_WA'].iter_rows(values_only=True))
    wb.close()
    return {
        'ann_thresh':   float(rows[0][2]),
        'twoy_thresh':  float(rows[1][2]),
        'apjul_thresh': APR_JUL_THRESH,   # unchanged - hardcoded in spreadsheet
    }


def _read_dss(*keys) -> dict:
    """Open the CalSim SV DSS once and return {(B, C): month-end Series}
    (NaN dropped) for the requested upper-case (B-part, C-part) pairs."""
    want = set(keys)
    with dss_io.open_dss(str(DSS_FILE), version=6, catalog_flag=True) as dss:
        smap = dss_io.read_monthly_series(dss, want, DSS_READ_START, DSS_READ_END)
    missing = want - set(smap)
    if missing:
        raise KeyError(f"Not found in {DSS_FILE}: {sorted(missing)}")
    return {k: v.dropna() for k, v in smap.items()}


def _load_spreadsheet_values() -> pd.Series:
    """
    Extract the monthly MINFLOWFEATHER values (cfs) as originally computed
    in the Main_WA sheet of MIF_FEATHER_Logic_Reconstruction.xlsx.

    Main_WA layout (0-indexed columns, data begins at row index 3):
      Col 0  : CY (calendar year N)
      Cols 7-12 : Apr(N), May(N), Jun(N), Jul(N), Aug(N), Sep(N)
      Cols 13-18: Oct(N), Nov(N), Dec(N), Jan(N+1), Feb(N+1), Mar(N+1)

    Returns end-of-month DatetimeIndex Series of integer cfs values.
    """
    import openpyxl
    wb = openpyxl.load_workbook(PATH_XLSX, read_only=True, data_only=True)
    ws = wb['Main_WA']
    rows = list(ws.iter_rows(values_only=True))
    # Data rows start at index 3 (row 4 in Excel)
    month_order = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]  # cols 7-18
    flows = {}
    for row in rows[3:]:
        cy = row[0]
        if not isinstance(cy, int):
            continue
        for i, m in enumerate(month_order):
            val = row[7 + i]
            if val is None or not isinstance(val, (int, float)):
                continue
            yr = cy if m >= 4 else cy + 1
            dt = pd.Timestamp(yr, m, 1) + pd.offsets.MonthEnd(0)
            flows[dt] = int(val)
    wb.close()
    s = pd.Series(flows, name='Value_Spreadsheet', dtype=int)
    s.index.name = 'date'
    return s.sort_index()


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def _schedule_summary_from_series(unimp_orov: pd.Series) -> dict:
    """Return {cy: schedule_name} dict for a given UNIMP_OROV series."""
    data = unimp_orov.dropna().copy()
    wy_idx    = _water_year(data.index)
    annual_wy = data.groupby(wy_idx).sum()
    twoy_avg  = (annual_wy + annual_wy.shift(1)) / 2
    mask_aj   = data.index.month.isin([4, 5, 6, 7])
    apjul_cy  = data[mask_aj].groupby(data[mask_aj].index.year).sum()
    sched = {}
    for cy in range(int(annual_wy.index[0]), int(annual_wy.index[-1]) + 1):
        ann   = annual_wy.get(cy, np.nan)
        twoy  = twoy_avg.get(cy,  np.nan)
        apjul = apjul_cy.get(cy,  np.nan)
        if np.isnan(ann) or np.isnan(apjul):
            continue
        sched[cy] = _select_schedule(
            ann   < ANN_THRESH,
            (not np.isnan(twoy)) and (twoy < TWOY_THRESH),
            apjul < APR_JUL_THRESH,
        )
    return sched


def _load_product_b_chunk(chunk_num: int) -> pd.Series:
    """
    Load a single Product B UNIMP_OROV chunk (qmap_postAdj column, TAF).
    Files now carry historical-epoch dates (Oct 1921 - Sep 2021, 1200 rows).
    """
    path = os.path.join(
        PATH_UNIMP_OROV_PB_DIR,
        f'UNIMP_OROV_qmo_n{chunk_num:02d}.csv'
    )
    df = pd.read_csv(path, usecols=['Year', 'Month', 'qmap_postAdj'])
    # Build end-of-month DatetimeIndex from the file's Year/Month columns
    dates = pd.to_datetime(
        {'year': df['Year'], 'month': df['Month'], 'day': 1}
    ) + pd.offsets.MonthEnd(0)
    return pd.Series(df['qmap_postAdj'].values, index=dates, name='UNIMP_OROV')


def run_validation():
    """Output 1: validation wide-format vs spreadsheet (WY 1922-2021)."""
    os.makedirs(OUT_DIR, exist_ok=True)
    print("\n" + "-" * 65)
    print("OUTPUT 1 - Validation  (WY 1922-2021, historical data from xlsx)")
    print("  Spreadsheet reference: Main_WA cols G-S, rows 4-104 (CY 1921-2021)")
    print("-" * 65)

    dss_series = _read_dss(UNIMP_OROV_KEY, MINFLOW_KEY)
    unimp_hist = dss_series[UNIMP_OROV_KEY]

    xlsx_thresh = _load_xlsx_thresholds()
    print("\n  Spreadsheet thresholds (from Main_WA):")
    print(f"    ANN_THRESH  = {xlsx_thresh['ann_thresh']:.4f} TAF")
    print(f"    TWOY_THRESH = {xlsx_thresh['twoy_thresh']:.4f} TAF")
    print(f"    (module defaults: ANN={ANN_THRESH:.4f}, TWOY={TWOY_THRESH:.4f})")

    mf_hist = compute_min_flow_feather(unimp_hist, **xlsx_thresh)
    mf_val  = mf_hist.loc['1921-10-31':'2021-09-30']
    df_val_wide = to_calsim_csv(mf_val, 'MINFLOWFEATHER', 'FLOW-MIN-REQUIRED') \
                      .rename(columns={'Value': 'Value_Python'})

    cs3_vals = dss_series[MINFLOW_KEY].round().astype(int)
    cs3_df = pd.DataFrame({
        'Year':  cs3_vals.index.year,
        'Month': cs3_vals.index.month,
        'Value_CS3': cs3_vals.values,
    })
    df_val_wide = df_val_wide.merge(cs3_df, on=['Year', 'Month'], how='left')

    if os.path.exists(PATH_XLSX):
        ss_vals = _load_spreadsheet_values()
        ss_df = pd.DataFrame({
            'Year':  ss_vals.index.year,
            'Month': ss_vals.index.month,
            'Value_Spreadsheet': ss_vals.values,
        })
        df_val_wide = df_val_wide.merge(ss_df, on=['Year', 'Month'], how='left')
    else:
        print("  (Spreadsheet not found - Value_Spreadsheet will be NaN)")
        df_val_wide['Value_Spreadsheet'] = pd.NA

    df_val_wide = df_val_wide[
        ['Part B', 'Part C', 'Year', 'Month',
         'Value_CS3', 'Value_Spreadsheet', 'Value_Python']
    ]
    df_val_wide.to_csv(OUT_VALIDATION, index=False)
    print(f"  Written : {OUT_VALIDATION}")
    print(f"  Rows    : {len(df_val_wide):,}")

    def _match_summary(col_a, col_b, label):
        mask  = df_val_wide[col_a].notna() & df_val_wide[col_b].notna()
        n     = mask.sum()
        if n == 0:
            return
        diffs = (df_val_wide.loc[mask, col_a] != df_val_wide.loc[mask, col_b]).sum()
        pct   = 100.0 * (n - diffs) / n
        note  = f"  ({diffs} diffs)" if diffs else "  OK exact match"
        print(f"  {label}: {n - diffs:,}/{n:,} ({pct:.1f}%){note}")

    _match_summary('Value_Python',      'Value_CS3',         'Python  vs CS3        ')
    _match_summary('Value_Spreadsheet', 'Value_CS3',         'Sheet   vs CS3        ')
    _match_summary('Value_Python',      'Value_Spreadsheet', 'Python  vs Sheet      ')


def _read_product_a_unimp_orov(path: Path = PATH_PRODUCT_A_CSV) -> pd.Series:
    """Load quantile-mapped Product A UNIMP_OROV (TAF, WY 1972-2018) from the
    rim-inflow Product A CSV (rows where Part B == UNIMP_OROV)."""
    df = pd.read_csv(path)
    df = df[df['Part B'].astype(str).str.upper() == UNIMP_OROV_KEY[0]]
    idx = (pd.to_datetime({'year': df['Year'], 'month': df['Month'], 'day': 1})
           + pd.offsets.MonthEnd(0))
    return pd.Series(df['Value'].astype(float).to_numpy(), index=idx).sort_index()


def run_product_a():
    """Output 2: Product A (WY 1972-2018, quantile-mapped UNIMP_OROV)."""
    os.makedirs(OUT_PRODUCT_A_DIR, exist_ok=True)
    print("\n" + "-" * 65)
    print("OUTPUT 2 - Product A  (WY 1972-2018, quantile-mapped UNIMP_OROV)")
    print("-" * 65)

    unimp_pa = _read_product_a_unimp_orov()
    mf_pa    = compute_min_flow_feather(unimp_pa)

    # The first months of the input (Oct 1971-Mar 1972) belong to a schedule
    # cycle that needs water year 1971, which the WY-1972-start input lacks, so
    # they cannot be classified.  Fill them from the CalSim baseline
    # MINFLOWFEATHER.
    baseline_mf = _read_dss(MINFLOW_KEY)[MINFLOW_KEY]
    missing = unimp_pa.index.difference(mf_pa.index)
    mf_pa_out = (pd.concat([baseline_mf.reindex(missing).dropna(), mf_pa])
                 .sort_index().round().astype(int))

    df_pa = to_calsim_csv(mf_pa_out, 'MINFLOWFEATHER', 'FLOW-MIN-REQUIRED')
    df_pa.to_csv(OUT_PRODUCT_A, index=False)
    print(f"  Written : {OUT_PRODUCT_A}")
    wy_first = int(df_pa.loc[df_pa['Month'] == 10, 'Year'].min()) + 1
    wy_last  = int(df_pa.loc[df_pa['Month'] == 9,  'Year'].max())
    print(f"  Rows    : {len(df_pa):,}  (WY {wy_first} - WY {wy_last})")
    print(f"  Leading months filled from baseline: {len(missing)}")
    print("  Product A schedule distribution:")
    _print_schedule_summary(_schedule_summary_from_series(unimp_pa))


def run_product_b():
    """Output 3: Product B - 10 chunks, WY 1922-2021 each."""
    os.makedirs(OUT_PRODUCT_B_DIR, exist_ok=True)
    print("\n" + "-" * 65)
    print("OUTPUT 3 - Product B  (10 chunks, WY 1922-2021)")
    print("-" * 65)

    # CS3 WY 1921 seed (Oct 1920 - Sep 1921) for 2-year rolling avg priming.
    unimp_hist = _read_dss(UNIMP_OROV_KEY)[UNIMP_OROV_KEY]
    _cs3_seed  = unimp_hist.loc['1920-10-31':'1921-09-30'].copy()

    for n in range(1, 11):
        pb_path = os.path.join(
            PATH_UNIMP_OROV_PB_DIR,
            f'UNIMP_OROV_qmo_n{n:02d}.csv'
        )
        if not os.path.exists(pb_path):
            print(f"  [n{n:02d}] Skipped - file not found: {pb_path}")
            continue

        unimp_pb       = _load_product_b_chunk(n)
        unimp_combined = pd.concat([_cs3_seed, unimp_pb]).sort_index()
        mf_combined    = compute_min_flow_feather(unimp_combined)
        mf_pb          = mf_combined.loc[unimp_pb.index[0]:unimp_pb.index[-1]].copy()
        df_pb          = to_calsim_csv(mf_pb, 'MINFLOWFEATHER', 'FLOW-MIN-REQUIRED')

        out_path = OUT_PRODUCT_B_PATTERN.format(n)
        df_pb.to_csv(out_path, index=False)

        wy_first_pb = int(df_pb.loc[df_pb['Month'] == 10, 'Year'].min()) + 1
        wy_last_pb  = int(df_pb.loc[df_pb['Month'] == 9,  'Year'].max())
        print(f"  [n{n:02d}] {len(df_pb):,} rows  WY {wy_first_pb}-{wy_last_pb}  -> {os.path.basename(out_path)}")


def main(product: str):
    """Run one output.  product in {'validation', 'A', 'B'}."""
    print("=" * 65)
    print("MINFLOWFEATHER - Feather River Minimum Instream Flow")
    print("=" * 65)
    if product == 'validation':
        run_validation()
    elif product == 'A':
        run_product_a()
    elif product == 'B':
        run_product_b()
    print("\nOK  Done.\n")


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(
        description='Compute MINFLOWFEATHER (Feather River minimum instream flow).'
    )
    parser.add_argument(
        '--product', choices=['validation', 'A', 'B'], required=True,
        help='Which output to run: validation (3-way comparison wide CSV), '
             'A (Product A historical 1922-2018), or B (Product B 10 stochastic chunks).'
    )
    args = parser.parse_args()
    main(product=args.product)
