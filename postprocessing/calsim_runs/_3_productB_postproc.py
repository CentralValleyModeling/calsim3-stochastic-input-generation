r"""
Product B stochastic post-processing.

This module is tailored to the 10 x 100-year Product B block structure:
- one benchmark / baseline scenario
- ten stochastic block scenarios n01 ... n10
- each block spans the same 100 water years

Outputs:
1) annual_block_summary.xlsx
   - benchmark_summary
   - block_summary
   - annual_values_long

2) rolling_minima.xlsx
   - benchmark_rolling_minima
   - ensemble_rolling_minima

3) figures/annual_cdf/<metric>.png
   - annual water-year CDFs, with one line for the benchmark and one line per Product B block

Additional statistic requested:
- minimum 2, 3, ... 10-year rolling average for every metric
- reported with the Product B block (n1-n10) and the water-year range where the minimum occurs

Assumptions:
- values.pkl metric columns are already in TAF
- Product B block scenarios are named either n01..n10, or contain those tokens in the scenario name
- Non-storage metrics are monthly volumes that should be summed to water years
- Storage metrics are represented by end-of-September carryover values

Typical usage:

    python 3_productB_postproc.py ^
        --pickle-dir "output\_2_pickle_builder\cache\Benchmark_vs_ProductB" ^
        --benchmark-name Benchmark ^
        --out-dir "output\_3_productB_postproc"
"""

from __future__ import annotations

import argparse
from pathlib import Path
import pickle
import re
from typing import Dict, Iterable, List, Sequence, Tuple

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


FIXED_COLS = {"Date", "Scenario", "OctSeptYear", "MarFebYear", "Year", "Month", "JanDecYear"}
_BLOCK_RE = re.compile(r"(?<![A-Za-z0-9])n0*([1-9]|10)(?![A-Za-z0-9])", flags=re.IGNORECASE)


def load_pickles(pickle_dir: str | Path) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, str], Dict[str, str]]:
    pickle_path = Path(pickle_dir)
    with open(pickle_path / "values.pkl", "rb") as f:
        df_values = pickle.load(f)
    with open(pickle_path / "diffs.pkl", "rb") as f:
        df_diffs = pickle.load(f)
    with open(pickle_path / "units.pkl", "rb") as f:
        units = pickle.load(f)
    with open(pickle_path / "fields.pkl", "rb") as f:
        fields = pickle.load(f)
    return df_values, df_diffs, units, fields


def metric_groups_from_fields(fields: Dict[str, str]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for key, label in fields.items():
        if isinstance(label, str) and ":" in label:
            out[key] = label.split(":", 1)[0].strip()
        else:
            out[key] = ""
    return out


def metric_label_from_fields(metric_key: str, fields: Dict[str, str]) -> str:
    raw = fields.get(metric_key, metric_key)
    return raw.split(":", 1)[1].strip() if ":" in raw else raw


def extract_block_index(scenario_name: str) -> int | None:
    match = _BLOCK_RE.search(str(scenario_name))
    if not match:
        return None
    return int(match.group(1))


def block_label_from_index(block_index: int | float | None) -> str:
    if block_index is None or pd.isna(block_index):
        return ""
    return f"n{int(block_index)}"


def water_year_aggregate(df: pd.DataFrame, metric_key: str, group: str) -> pd.DataFrame:
    """
    Returns a DataFrame with columns:
      Scenario, OctSeptYear, WY_value

    Rule for mixed metric types:
      - Storage group: end-of-September value (end of water year)
      - Else: sum monthly values to annual volume
    """

    needed = {"Scenario", "OctSeptYear", metric_key, "Date"}
    missing = needed - set(df.columns)
    if missing:
        raise ValueError(f"values.pkl missing columns: {sorted(missing)}")

    if group.strip().lower() == "storage":
        tmp = df[["Scenario", "Date", "OctSeptYear", metric_key]].copy()
        tmp = tmp.dropna(subset=[metric_key])
        tmp = tmp[tmp["Date"].dt.month == 9]
        wy = tmp.groupby(["Scenario", "OctSeptYear"], as_index=False)[metric_key].last()
    else:
        tmp = df[["Scenario", "OctSeptYear", metric_key]].copy()
        tmp = tmp.dropna(subset=[metric_key])
        wy = tmp.groupby(["Scenario", "OctSeptYear"], as_index=False)[metric_key].sum()

    return wy.rename(columns={metric_key: "WY_Value"})


def annualize_all_metrics(
    df_values: pd.DataFrame,
    metric_keys: Sequence[str],
    metric_groups: Dict[str, str],
    fields: Dict[str, str],
) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []

    for metric_key in metric_keys:
        group = metric_groups.get(metric_key, "")
        label = metric_label_from_fields(metric_key, fields)

        wy = water_year_aggregate(df=df_values, metric_key=metric_key, group=group)
        wy = wy.rename(columns={"OctSeptYear": "WY"})
        wy.insert(0, "Metric_Label", label)
        wy.insert(0, "Metric", metric_key)
        wy.insert(0, "Group", group)
        wy["Block_Index"] = wy["Scenario"].map(extract_block_index)
        wy["Block"] = wy["Block_Index"].map(block_label_from_index)
        frames.append(wy)

    annual_long = pd.concat(frames, ignore_index=True)
    annual_long = annual_long.sort_values(["Metric", "Scenario", "WY"]).reset_index(drop=True)
    return annual_long


def benchmark_summary_table(annual_long: pd.DataFrame, benchmark_name: str) -> pd.DataFrame:
    benchmark = annual_long[annual_long["Scenario"] == benchmark_name].copy()
    table = (
        benchmark.groupby(["Group", "Metric", "Metric_Label"], as_index=False)["WY_Value"]
        .agg(
            Benchmark_Years="count",
            Benchmark_Mean_WY_TAF="mean",
            Benchmark_Median_WY_TAF="median",
            Benchmark_Min_WY_TAF="min",
            Benchmark_Max_WY_TAF="max",
            Benchmark_Std_WY_TAF="std",
        )
    )
    return table.sort_values(["Group", "Metric"]).reset_index(drop=True)


def block_summary_table(annual_long: pd.DataFrame, benchmark_name: str) -> pd.DataFrame:
    benchmark_table = benchmark_summary_table(annual_long=annual_long, benchmark_name=benchmark_name)

    blocks = annual_long[annual_long["Block_Index"].notna()].copy()
    table = (
        blocks.groupby(
            ["Group", "Metric", "Metric_Label", "Scenario", "Block", "Block_Index"],
            as_index=False,
        )["WY_Value"]
        .agg(
            Block_Years="count",
            Mean_WY_TAF="mean",
            Median_WY_TAF="median",
            Min_WY_TAF="min",
            Max_WY_TAF="max",
            Std_WY_TAF="std",
        )
    )

    table = table.merge(
        benchmark_table,
        on=["Group", "Metric", "Metric_Label"],
        how="left",
        validate="many_to_one",
    )
    table["Mean_Diff_vs_Benchmark_TAF"] = table["Mean_WY_TAF"] - table["Benchmark_Mean_WY_TAF"]
    table["Mean_Diff_vs_Benchmark_pct"] = np.where(
        table["Benchmark_Mean_WY_TAF"].ne(0),
        table["Mean_Diff_vs_Benchmark_TAF"] / table["Benchmark_Mean_WY_TAF"] * 100.0,
        np.nan,
    )
    table = table.sort_values(["Group", "Metric", "Block_Index"]).reset_index(drop=True)
    return table


def _rolling_candidates(metric_block_df: pd.DataFrame, window_years: int) -> pd.DataFrame:
    metric_block_df = metric_block_df.sort_values("WY").reset_index(drop=True)
    rolling = metric_block_df["WY_Value"].rolling(window_years).mean()

    rows: List[dict] = []
    for idx in range(window_years - 1, len(metric_block_df)):
        value = rolling.iloc[idx]
        if pd.isna(value):
            continue
        wy_start = int(metric_block_df.loc[idx - window_years + 1, "WY"])
        wy_end = int(metric_block_df.loc[idx, "WY"])
        rows.append(
            {
                "Window_Years": int(window_years),
                "RollingAvg_TAF": float(value),
                "WY_Start": wy_start,
                "WY_End": wy_end,
            }
        )

    return pd.DataFrame(rows)


def benchmark_rolling_minima_table(
    annual_long: pd.DataFrame,
    benchmark_name: str,
    min_window_years: int = 2,
    max_window_years: int = 10,
) -> pd.DataFrame:
    benchmark = annual_long[annual_long["Scenario"] == benchmark_name].copy()
    rows: List[dict] = []

    for (group, metric, label), metric_df in benchmark.groupby(["Group", "Metric", "Metric_Label"], sort=False):
        for window_years in range(min_window_years, max_window_years + 1):
            candidates = _rolling_candidates(metric_df, window_years)
            if candidates.empty:
                continue

            min_value = candidates["RollingAvg_TAF"].min()
            tie_count = int(np.isclose(candidates["RollingAvg_TAF"], min_value).sum())

            best = (
                candidates.sort_values(["RollingAvg_TAF", "WY_Start", "WY_End"])
                .iloc[0]
                .to_dict()
            )

            rows.append(
                {
                    "Group": group,
                    "Metric": metric,
                    "Metric_Label": label,
                    "Window_Years": int(best["Window_Years"]),
                    "Benchmark_Min_RollingAvg_TAF": float(best["RollingAvg_TAF"]),
                    "Benchmark_WY_Start": int(best["WY_Start"]),
                    "Benchmark_WY_End": int(best["WY_End"]),
                    "Benchmark_Tie_Count": tie_count,
                    "Benchmark_Scenario": benchmark_name,
                }
            )

    table = pd.DataFrame(rows)
    if table.empty:
        return table

    return table.sort_values(["Group", "Metric", "Window_Years"]).reset_index(drop=True)


def ensemble_rolling_minima_table(
    annual_long: pd.DataFrame,
    benchmark_rolling: pd.DataFrame,
    min_window_years: int = 2,
    max_window_years: int = 10,
) -> pd.DataFrame:
    blocks = annual_long[annual_long["Block_Index"].notna()].copy()
    rows: List[dict] = []

    for (group, metric, label), metric_df in blocks.groupby(["Group", "Metric", "Metric_Label"], sort=False):
        for window_years in range(min_window_years, max_window_years + 1):
            candidate_frames: List[pd.DataFrame] = []

            for (scenario, block, block_index), block_df in metric_df.groupby(
                ["Scenario", "Block", "Block_Index"],
                sort=True,
            ):
                candidates = _rolling_candidates(block_df, window_years)
                if candidates.empty:
                    continue
                candidates["Scenario"] = scenario
                candidates["Block"] = block
                candidates["Block_Index"] = int(block_index)
                candidate_frames.append(candidates)

            if not candidate_frames:
                continue

            candidates_all = pd.concat(candidate_frames, ignore_index=True)
            min_value = candidates_all["RollingAvg_TAF"].min()
            tie_count = int(np.isclose(candidates_all["RollingAvg_TAF"], min_value).sum())

            best = (
                candidates_all.sort_values(["RollingAvg_TAF", "Block_Index", "WY_Start", "WY_End"])
                .iloc[0]
                .to_dict()
            )

            rows.append(
                {
                    "Group": group,
                    "Metric": metric,
                    "Metric_Label": label,
                    "Window_Years": int(best["Window_Years"]),
                    "Min_RollingAvg_TAF": float(best["RollingAvg_TAF"]),
                    "Scenario": best["Scenario"],
                    "Block": best["Block"],
                    "Block_Index": int(best["Block_Index"]),
                    "WY_Start": int(best["WY_Start"]),
                    "WY_End": int(best["WY_End"]),
                    "Tie_Count": tie_count,
                }
            )

    table = pd.DataFrame(rows)
    if table.empty:
        return table

    table = table.merge(
        benchmark_rolling,
        on=["Group", "Metric", "Metric_Label", "Window_Years"],
        how="left",
        validate="one_to_one",
    )
    table["Min_RollingAvg_Diff_vs_Benchmark_TAF"] = (
        table["Min_RollingAvg_TAF"] - table["Benchmark_Min_RollingAvg_TAF"]
    )
    table["Min_RollingAvg_Diff_vs_Benchmark_pct"] = np.where(
        table["Benchmark_Min_RollingAvg_TAF"].ne(0),
        table["Min_RollingAvg_Diff_vs_Benchmark_TAF"] / table["Benchmark_Min_RollingAvg_TAF"] * 100.0,
        np.nan,
    )

    return table.sort_values(["Group", "Metric", "Window_Years"]).reset_index(drop=True)


def empirical_cdf(x: Iterable[float]) -> Tuple[np.ndarray, np.ndarray]:
    values = np.asarray(list(x), dtype=float)
    values = values[np.isfinite(values)]
    if values.size == 0:
        return np.array([]), np.array([])
    xs = np.sort(values)
    p = np.arange(1, xs.size + 1) / xs.size * 100.0
    return p, xs


def plot_annual_cdf(
    annual_long: pd.DataFrame,
    metric_key: str,
    metric_label: str,
    benchmark_name: str,
    out_png: str | Path,
    unit: str = "TAF",
) -> None:
    df_metric = annual_long[annual_long["Metric"] == metric_key].copy()

    fig, ax = plt.subplots(figsize=(6.5, 4.0))

    benchmark = df_metric[df_metric["Scenario"] == benchmark_name]
    p, xs = empirical_cdf(benchmark["WY_Value"].to_numpy(dtype=float))
    ax.plot(p, xs, linewidth=2.0, label=benchmark_name, color="black")

    block_order = (
        df_metric[df_metric["Block_Index"].notna()][["Scenario", "Block", "Block_Index"]]
        .drop_duplicates()
        .sort_values("Block_Index")
    )

    for _, row in block_order.iterrows():
        scenario = row["Scenario"]
        block_label = row["Block"]
        sub = df_metric[df_metric["Scenario"] == scenario]
        p, xs = empirical_cdf(sub["WY_Value"].to_numpy(dtype=float))
        ax.plot(p, xs, linewidth=1.1, alpha=0.9, label=block_label)

    ax.set_xlabel("Non-Exceedance Probability (%)")
    ax.set_ylabel(unit)
    ax.set_xlim(0, 100)
    ax.set_title(f"{metric_label} — Annual Water-Year CDF")
    ax.grid(True, linewidth=0.35, alpha=0.4)
    ax.legend(ncol=3, fontsize=7, frameon=False, loc="best")
    fig.tight_layout()

    out_path = Path(out_png)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=300, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def format_excel_workbook(path: str | Path) -> None:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    centered = Alignment(horizontal="center", vertical="center")
    left_aligned = Alignment(horizontal="left", vertical="center")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered

        column_widths: Dict[str, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                text = "" if cell.value is None else str(cell.value)
                width = min(max(len(text) + 2, 10), 42)
                column_widths[cell.column_letter] = max(column_widths.get(cell.column_letter, 0), width)

        for letter, width in column_widths.items():
            ws.column_dimensions[letter].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = str(ws.cell(row=1, column=cell.column).value or "").lower()
                if header in {"group", "metric", "metric_label", "scenario", "block"}:
                    cell.alignment = left_aligned
                else:
                    cell.alignment = centered

                if "pct" in header:
                    cell.number_format = '0.0'
                elif header in {"wy", "wy_start", "wy_end", "block_index", "window_years", "tie_count", "benchmark_tie_count", "benchmark_years", "block_years"}:
                    cell.number_format = '0'
                elif any(token in header for token in ["taf", "mean", "median", "min", "max", "std", "rollingavg"]):
                    cell.number_format = '#,##0.0'

    wb.save(workbook_path)


def run_post_processing_package(
    pickle_dir: str | Path,
    benchmark_name: str,
    out_dir: str | Path,
    min_window_years: int = 2,
    max_window_years: int = 10,
) -> Dict[str, str]:
    df_values, df_diffs, units, fields = load_pickles(pickle_dir)
    df_values["Date"] = pd.to_datetime(df_values["Date"])

    metric_groups = metric_groups_from_fields(fields)
    metric_keys = [col for col in df_values.columns if col not in FIXED_COLS]

    annual_long = annualize_all_metrics(
        df_values=df_values,
        metric_keys=metric_keys,
        metric_groups=metric_groups,
        fields=fields,
    )

    benchmark_summary = benchmark_summary_table(annual_long=annual_long, benchmark_name=benchmark_name)
    block_summary = block_summary_table(annual_long=annual_long, benchmark_name=benchmark_name)

    benchmark_rolling = benchmark_rolling_minima_table(
        annual_long=annual_long,
        benchmark_name=benchmark_name,
        min_window_years=min_window_years,
        max_window_years=max_window_years,
    )
    ensemble_rolling = ensemble_rolling_minima_table(
        annual_long=annual_long,
        benchmark_rolling=benchmark_rolling,
        min_window_years=min_window_years,
        max_window_years=max_window_years,
    )

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    annual_summary_xlsx = out_path / "annual_block_summary.xlsx"
    with pd.ExcelWriter(annual_summary_xlsx, engine="openpyxl") as writer:
        benchmark_summary.to_excel(writer, sheet_name="benchmark_summary", index=False)
        block_summary.to_excel(writer, sheet_name="block_summary", index=False)
        annual_long.to_excel(writer, sheet_name="annual_values_long", index=False)
    format_excel_workbook(annual_summary_xlsx)

    rolling_minima_xlsx = out_path / "rolling_minima.xlsx"
    with pd.ExcelWriter(rolling_minima_xlsx, engine="openpyxl") as writer:
        benchmark_rolling.to_excel(writer, sheet_name="benchmark_rolling_minima", index=False)
        ensemble_rolling.to_excel(writer, sheet_name="ensemble_rolling_minima", index=False)
    format_excel_workbook(rolling_minima_xlsx)

    fig_dir = out_path / "figures" / "annual_cdf"
    for metric_key in metric_keys:
        plot_annual_cdf(
            annual_long=annual_long,
            metric_key=metric_key,
            metric_label=metric_label_from_fields(metric_key, fields),
            benchmark_name=benchmark_name,
            out_png=fig_dir / f"{metric_key}.png",
            unit=units.get(metric_key, "TAF"),
        )

    return {
        "annual_summary_xlsx": str(annual_summary_xlsx),
        "rolling_minima_xlsx": str(rolling_minima_xlsx),
        "annual_cdf_dir": str(fig_dir),
        "n_metrics": str(len(metric_keys)),
        "n_rows_annual_values": str(len(annual_long)),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Run Product B stochastic post-processing.")
    parser.add_argument("--pickle-dir", required=True, help="Directory containing values.pkl / diffs.pkl / units.pkl / fields.pkl")
    parser.add_argument("--benchmark-name", default="Benchmark", help="Benchmark scenario name in values.pkl")
    parser.add_argument(
        "--out-dir",
        default=str(Path(__file__).resolve().parent / "output" / "_3_productB_postproc"),
        help="Output directory for Product B post-processing",
    )
    parser.add_argument("--min-window-years", type=int, default=2, help="Minimum rolling window in years")
    parser.add_argument("--max-window-years", type=int, default=10, help="Maximum rolling window in years")
    args = parser.parse_args()

    outputs = run_post_processing_package(
        pickle_dir=args.pickle_dir,
        benchmark_name=args.benchmark_name,
        out_dir=args.out_dir,
        min_window_years=args.min_window_years,
        max_window_years=args.max_window_years,
    )

    print("Created:")
    for key, value in outputs.items():
        print(f"  {key}: {value}")


if __name__ == "__main__":
    main()
