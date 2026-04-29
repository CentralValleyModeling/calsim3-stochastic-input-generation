r"""
Product B stochastic post-processing.

This module is tailored to the 10 x 100-year Product B block structure:
- one baseline scenario
- ten stochastic block scenarios n01 ... n10
- each block spans the same 100 water years

Outputs:
1) heatmap_block_summary.xlsx
   - % difference vs benchmark for mean annual, P5 annual,
     worst 5-year rolling average, and worst 10-year rolling average

2) annual_block_summary.xlsx
   - historical_summary
   - block_summary
   - annual_values_long

3) rolling_minima.xlsx
   - historical_rolling_minima
   - stochastic_rolling_minima

4) figures/annual_cdf/<metric>.png
   - annual water-year CDFs, with one line for the benchmark and one line per Product B block

5) figures/monthly_cdf/<metric>.png
   - monthly CDFs, with all monthly values on one CDF for the benchmark and Product B blocks

6) figures/block_boxplots/<metric>.png
   - benchmark distribution plus one Product B box per block; historical mean shown as
     the single horizontal reference line

7) figures/worst_window_sequences/<window>yr/<metric>_worst_<window>yr.png
   - overlays of the historical worst sequence and each block's own worst sequence for
     2-year and 5-year rolling windows. The 2-year plot uses a 6-year frame;
     the 5-year plot uses the default fixed 15-year frame.

8) compact_summary.xlsx
   - one-row-per-metric summary: historical annual avg, N1-10 block range, rolling minima

9) figures/timeseries_1000yr/<metric>.png
   - historical trace followed by stitched 1000-year stochastic sequence; optional
     Product A 10-yr rolling overlay in the historical region

10) rolling_minima_vs_historical_counts.xlsx
    - counts_by_block: per-block count of metrics whose worst rolling average is
      strictly worse than the historical worst rolling average
    - details: per-metric breakdown

11) figures/rolling_minima_vs_historical_counts/block_rolling_minima_below_historical_counts*.png
    - grouped bar charts of below-historical rolling-minimum counts per block

12) figures/annual_block_range/range_100yr_block_means_vs_historical.png / .svg
    - lollipop/range chart of 100-year block means vs the historical mean
      for key Delta, Delivery, and Storage metrics


Dependencies:
- Python 3.10+ (uses PEP 604 union syntax, e.g. ``int | None``)
- pandas, numpy, matplotlib, and openpyxl
- Repository utility ``utils.paths.get_generated_dir`` for default path resolution
- Product B pickle cache containing ``values.pkl``, ``diffs.pkl``, ``units.pkl``,
    and ``fields.pkl``
- Optional Product A pickle cache with the same four files when using the
    10-year rolling overlay on the stitched 1000-year time series


Assumptions:
- values.pkl metric columns are already in TAF
- Product B block scenarios are named either n01..n10, or contain those tokens in the scenario name
- Non-storage metrics are monthly volumes that should be summed to water years
- Storage metrics are represented by end-of-September carryover values

Typical usage:

    python _productB_postproc.py ^
        --pickle-dir "data\GENERATED\postprocessing\calsim_runs\product_b\pickle_files" ^
        --benchmark-name Historical ^
        --out-dir "data\GENERATED\postprocessing\calsim_runs\product_b\output"
"""

from __future__ import annotations

import argparse
from pathlib import Path
import pickle
import re
from typing import Dict, Iterable, List, Sequence, Tuple

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.lines import Line2D
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


RUN_DIR = Path(__file__).resolve().parent
REPO_ROOT = RUN_DIR.parents[1]

import sys as _sys

_sys.path.insert(0, str(REPO_ROOT))
try:
    from utils.paths import get_generated_dir
except ImportError:
    def get_generated_dir() -> Path:
        return RUN_DIR

PICKLE_DIR = get_generated_dir() / "postprocessing" / "calsim_runs" / "product_b" / "pickle_files"
OUT_DIR = get_generated_dir() / "postprocessing" / "calsim_runs" / "product_b" / "output"
PRODUCT_A_PICKLE_DIR = (
    get_generated_dir()
    / "postprocessing"
    / "calsim_runs"
    / "product_a_modified"
    / "pickle_files"
)

FIXED_COLS = {"Date", "Scenario", "OctSeptYear", "MarFebYear", "Year", "Month", "JanDecYear"}
_BLOCK_RE = re.compile(r"(?<![A-Za-z0-9])n0*([1-9]|10)(?![A-Za-z0-9])", flags=re.IGNORECASE)

DEFAULT_DROUGHT_PERCENTILE = 5.0
HEATMAP_ROLLING_WINDOWS: Tuple[int, ...] = (5, 10)
# Exclude only the specific Cache Slough metric from heatmap outputs.
# Keep other Cache-related metrics (for example Cache Creek) in the heatmap.
HEATMAP_EXCLUDED_METRIC_KEYS: Tuple[str, ...] = ("C_CSL004A",)
HEATMAP_EXCLUDED_LABELS: Tuple[str, ...] = ("Cache Slough", "Delta: Cache Slough")
# Additional metrics excluded only from the block rolling-minima below-historical
# counts. The CVP and San Luis totals are already represented by their N/S and
# CVP/SWP component metrics, so counting them would double-count.
BLOCK_ROLLING_COUNT_EXCLUDED_METRIC_KEYS: Tuple[str, ...] = (
    "C_CSL004A",
    "DEL_CVP_TOTAL",
    "S_SLUIS_TOTAL",
)
COMPACT_SUMMARY_WINDOWS: Tuple[int, ...] = (2, 5, 10)
DEFAULT_SEQUENCE_WINDOWS: Tuple[int, ...] = (2, 5)
DEFAULT_SEQUENCE_FRAME_YEARS = 15
TWO_YEAR_SEQUENCE_FRAME_YEARS = 6


def load_pickles(pickle_dir: str | Path) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, str], Dict[str, str]]:
    pickle_path = Path(pickle_dir)
    with open(pickle_path / "values.pkl", "rb") as f:
        df_values = pickle.load(f)
    with open(pickle_path / "diffs.pkl", "rb") as f:
        df_diffs = pickle.load(f)
    with open(pickle_path / "units.pkl", "rb") as f:
        units = pickle.load(f)
    with open(pickle_path / "fields.pkl", "rb") as f:
        fields = pickle.load(f)
    return df_values, df_diffs, units, fields


def metric_groups_from_fields(fields: Dict[str, str]) -> Dict[str, str]:
    """Extract the group portion (text before ``:``) of each metric label."""
    out: Dict[str, str] = {}
    for key, label in fields.items():
        if isinstance(label, str) and ":" in label:
            out[key] = label.split(":", 1)[0].strip()
        else:
            out[key] = ""
    return out


def metric_label_from_fields(metric_key: str, fields: Dict[str, str]) -> str:
    """Return the display label (text after ``:``) for ``metric_key``."""
    raw = fields.get(metric_key, metric_key)
    return raw.split(":", 1)[1].strip() if ":" in raw else raw


def extract_block_index(scenario_name: str) -> int | None:
    """Return the 1-based block index parsed from a scenario name (e.g. ``n03`` -> 3)."""
    match = _BLOCK_RE.search(str(scenario_name))
    if not match:
        return None
    return int(match.group(1))


def block_label_from_index(block_index: int | float | None) -> str:
    """Format a block index as a zero-padded label (e.g. 3 -> ``"n03"``)."""
    if block_index is None or pd.isna(block_index):
        return ""
    return f"n{int(block_index):02d}"


def sort_block_labels(block_labels: Iterable[str]) -> List[str]:
    """Sort block labels by their embedded numeric index (``n01`` before ``n10``)."""
    def _key(label: str) -> Tuple[int, str]:
        digits = re.sub(r"\D", "", str(label))
        return (int(digits) if digits else 0, str(label))

    return sorted(block_labels, key=_key)


def make_safe_filename(text: str) -> str:
    """Sanitize ``text`` into a filesystem-safe filename stem."""
    safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(text))
    safe = safe.strip("_.")
    return safe or "plot"


# ---------------------------------------------------------------------------
# Annual aggregation and summary tables
# ---------------------------------------------------------------------------

def water_year_aggregate(df: pd.DataFrame, metric_key: str, group: str) -> pd.DataFrame:
    """
    Returns a DataFrame with columns:
      Scenario, OctSeptYear, WY_Value

    Rule for mixed metric types:
      - Storage group: end-of-September value (end of water year)
      - Else: sum monthly values to annual volume
    """

    needed = {"Scenario", "OctSeptYear", metric_key, "Date"}
    missing = needed - set(df.columns)
    if missing:
        raise ValueError(f"values.pkl missing columns: {sorted(missing)}")

    if group.strip().lower() == "storage":
        tmp = df[["Scenario", "Date", "OctSeptYear", metric_key]].copy()
        tmp = tmp.dropna(subset=[metric_key])
        tmp = tmp[tmp["Date"].dt.month == 9]
        wy = tmp.groupby(["Scenario", "OctSeptYear"], as_index=False)[metric_key].last()
    else:
        tmp = df[["Scenario", "OctSeptYear", metric_key]].copy()
        tmp = tmp.dropna(subset=[metric_key])
        wy = tmp.groupby(["Scenario", "OctSeptYear"], as_index=False)[metric_key].sum()

    return wy.rename(columns={metric_key: "WY_Value"})


def annualize_all_metrics(
    df_values: pd.DataFrame,
    metric_keys: Sequence[str],
    metric_groups: Dict[str, str],
    fields: Dict[str, str],
) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []

    for metric_key in metric_keys:
        group = metric_groups.get(metric_key, "")
        label = metric_label_from_fields(metric_key, fields)

        wy = water_year_aggregate(df=df_values, metric_key=metric_key, group=group)
        wy = wy.rename(columns={"OctSeptYear": "WY"})
        wy.insert(0, "Metric_Label", label)
        wy.insert(0, "Metric", metric_key)
        wy.insert(0, "Group", group)
        wy["Block_Index"] = wy["Scenario"].map(extract_block_index)
        wy["Block"] = wy["Block_Index"].map(block_label_from_index)
        non_block = wy["Block"] == ""
        wy.loc[non_block, "Block"] = wy.loc[non_block, "Scenario"]
        frames.append(wy)

    annual_long = pd.concat(frames, ignore_index=True)
    annual_long = annual_long.sort_values(["Metric", "Scenario", "WY"]).reset_index(drop=True)
    return annual_long


def load_product_a_annual(
    pickle_dir: str | Path,
    metric_keys: Sequence[str],
    metric_groups: Dict[str, str],
    fields: Dict[str, str],
    exclude_scenarios: Sequence[str] = ("Historical",),
) -> pd.DataFrame:
    """Load a Product A pickle cache and return a per-metric water-year DataFrame.

    The result has columns: Metric, WY, WY_Value. Only metrics that exist in
    metric_keys are returned, and only the non-benchmark scenario rows
    (the Product A scenario itself).
    """
    pickle_path = Path(pickle_dir)
    _required = ("values.pkl", "diffs.pkl", "units.pkl", "fields.pkl")
    if any(not (pickle_path / f).exists() for f in _required):
        return pd.DataFrame(columns=["Metric", "WY", "WY_Value"])

    df_values, _df_diffs, _units, _fields = load_pickles(pickle_path)
    df_values["Date"] = pd.to_datetime(df_values["Date"])

    available = [m for m in metric_keys if m in df_values.columns]
    if not available:
        return pd.DataFrame(columns=["Metric", "WY", "WY_Value"])

    pa_long = annualize_all_metrics(
        df_values=df_values,
        metric_keys=available,
        metric_groups=metric_groups,
        fields=fields,
    )
    excl = set(exclude_scenarios)
    pa_long = pa_long[~pa_long["Scenario"].isin(excl)].copy()
    if pa_long.empty:
        return pd.DataFrame(columns=["Metric", "WY", "WY_Value"])

    return pa_long[["Metric", "WY", "WY_Value"]].reset_index(drop=True)


def benchmark_summary_table(annual_long: pd.DataFrame, benchmark_name: str) -> pd.DataFrame:
    benchmark = annual_long[annual_long["Scenario"] == benchmark_name].copy()
    table = (
        benchmark.groupby(["Group", "Metric", "Metric_Label"], as_index=False)["WY_Value"]
        .agg(
            Historical_Years="count",
            Historical_Mean_WY_TAF="mean",
            Historical_Median_WY_TAF="median",
            Historical_Min_WY_TAF="min",
            Historical_Max_WY_TAF="max",
            Historical_Std_WY_TAF="std",
        )
    )
    return table.sort_values(["Group", "Metric"]).reset_index(drop=True)


def block_summary_table(annual_long: pd.DataFrame, benchmark_name: str) -> pd.DataFrame:
    benchmark_table = benchmark_summary_table(annual_long=annual_long, benchmark_name=benchmark_name)

    blocks = annual_long[annual_long["Block_Index"].notna()].copy()
    table = (
        blocks.groupby(
            ["Group", "Metric", "Metric_Label", "Scenario", "Block", "Block_Index"],
            as_index=False,
        )["WY_Value"]
        .agg(
            Block_Years="count",
            Mean_WY_TAF="mean",
            Median_WY_TAF="median",
            Min_WY_TAF="min",
            Max_WY_TAF="max",
            Std_WY_TAF="std",
        )
    )

    table = table.merge(
        benchmark_table,
        on=["Group", "Metric", "Metric_Label"],
        how="left",
        validate="many_to_one",
    )
    table["Mean_Diff_vs_Historical_TAF"] = table["Mean_WY_TAF"] - table["Historical_Mean_WY_TAF"]
    table["Mean_Diff_vs_Historical_pct"] = np.where(
        table["Historical_Mean_WY_TAF"].ne(0),
        table["Mean_Diff_vs_Historical_TAF"] / table["Historical_Mean_WY_TAF"] * 100.0,
        np.nan,
    )
    hist_cols = [c for c in table.columns if c.startswith("Historical_")]
    table = table.drop(columns=hist_cols)
    table = table.sort_values(["Group", "Metric", "Block_Index"]).reset_index(drop=True)
    table = table.drop(columns=["Scenario", "Block_Index"])
    return table


# ---------------------------------------------------------------------------
# Rolling-window helpers and tables
# ---------------------------------------------------------------------------

def _rolling_candidates(metric_block_df: pd.DataFrame, window_years: int) -> pd.DataFrame:
    metric_block_df = metric_block_df.sort_values("WY").reset_index(drop=True)
    if metric_block_df.empty or len(metric_block_df) < window_years:
        return pd.DataFrame(columns=["Window_Years", "RollingAvg_TAF", "WY_Start", "WY_End", "Seq_Start", "Seq_End"])

    rolling = metric_block_df["WY_Value"].rolling(window_years).mean()

    rows: List[dict] = []
    for idx in range(window_years - 1, len(metric_block_df)):
        value = rolling.iloc[idx]
        if pd.isna(value):
            continue
        wy_start = int(metric_block_df.loc[idx - window_years + 1, "WY"])
        wy_end = int(metric_block_df.loc[idx, "WY"])
        seq_start = int(idx - window_years + 2)  # 1-based within-scenario position
        seq_end = int(idx + 1)                    # 1-based within-scenario position
        rows.append(
            {
                "Window_Years": int(window_years),
                "RollingAvg_TAF": float(value),
                "WY_Start": wy_start,
                "WY_End": wy_end,
                "Seq_Start": seq_start,
                "Seq_End": seq_end,
            }
        )

    return pd.DataFrame(rows)


def _best_rolling_window(metric_block_df: pd.DataFrame, window_years: int) -> dict | None:
    candidates = _rolling_candidates(metric_block_df, window_years)
    if candidates.empty:
        return None

    best = (
        candidates.sort_values(["RollingAvg_TAF", "Seq_Start", "Seq_End", "WY_Start", "WY_End"])
        .iloc[0]
        .to_dict()
    )
    return best


def benchmark_rolling_minima_table(
    annual_long: pd.DataFrame,
    benchmark_name: str,
    min_window_years: int = 2,
    max_window_years: int = 10,
) -> pd.DataFrame:
    benchmark = annual_long[annual_long["Scenario"] == benchmark_name].copy()
    rows: List[dict] = []

    for (group, metric, label), metric_df in benchmark.groupby(["Group", "Metric", "Metric_Label"], sort=False):
        for window_years in range(min_window_years, max_window_years + 1):
            best = _best_rolling_window(metric_df, window_years)
            if best is None:
                continue

            rows.append(
                {
                    "Group": group,
                    "Metric": metric,
                    "Metric_Label": label,
                    "Window_Years": int(best["Window_Years"]),
                    "Historical_Min_RollingAvg_TAF": float(best["RollingAvg_TAF"]),
                    "Historical_WY_Start": int(best["WY_Start"]),
                    "Historical_WY_End": int(best["WY_End"]),
                    "Historical_Year_Index_Start": int(best["Seq_Start"]),
                    "Historical_Year_Index_End": int(best["Seq_End"]),
                    "Historical_Scenario": benchmark_name,
                }
            )

    table = pd.DataFrame(rows)
    if table.empty:
        return table

    return table.sort_values(["Group", "Metric", "Window_Years"]).reset_index(drop=True)


def stochastic_rolling_minima_table(
    annual_long: pd.DataFrame,
    min_window_years: int = 2,
    max_window_years: int = 10,
) -> pd.DataFrame:
    blocks = annual_long[annual_long["Block_Index"].notna()].copy()
    rows: List[dict] = []

    for (group, metric, label), metric_df in blocks.groupby(["Group", "Metric", "Metric_Label"], sort=False):
        for window_years in range(min_window_years, max_window_years + 1):
            candidate_rows: List[dict] = []

            for (scenario, block, block_index), block_df in metric_df.groupby(
                ["Scenario", "Block", "Block_Index"],
                sort=True,
            ):
                best = _best_rolling_window(block_df, window_years)
                if best is None:
                    continue
                best["Scenario"] = scenario
                best["Block"] = block
                best["Block_Index"] = int(block_index)
                candidate_rows.append(best)

            if not candidate_rows:
                continue

            candidates_all = pd.DataFrame(candidate_rows)
            best = (
                candidates_all.sort_values(["RollingAvg_TAF", "Block_Index", "Seq_Start", "Seq_End"])
                .iloc[0]
                .to_dict()
            )

            rows.append(
                {
                    "Group": group,
                    "Metric": metric,
                    "Metric_Label": label,
                    "Window_Years": int(best["Window_Years"]),
                    "Min_RollingAvg_TAF": float(best["RollingAvg_TAF"]),
                    "Block": best["Block"],
                    "Year_Index_Start": int(best["Seq_Start"]),
                    "Year_Index_End": int(best["Seq_End"]),
                }
            )

    table = pd.DataFrame(rows)
    if table.empty:
        return table

    return table.sort_values(["Group", "Metric", "Window_Years"]).reset_index(drop=True)


def build_block_rolling_below_historical_counts(
    annual_long: pd.DataFrame,
    benchmark_name: str,
    window_years: Sequence[int] = (2, 5, 10),
    exclude_metric_keys: Sequence[str] = HEATMAP_EXCLUDED_METRIC_KEYS,
    exclude_labels: Sequence[str] = HEATMAP_EXCLUDED_LABELS,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Per-block rolling minima vs historical rolling minimum.

    For each metric and each window in ``window_years``, compute the worst
    rolling-mean window for the benchmark scenario and for every Product B
    block (n01..n10), then flag blocks where the block's worst rolling
    average is strictly less than the historical worst rolling average.

    Excludes lower-priority metrics using the same policy as the heatmaps
    (exact metric keys and label matches).

    Returns
    -------
    (details, counts)
        details: one row per (Metric, Block, Window_Years) with both the
            block and historical rolling minima and the below-historical
            flag.
        counts: one row per (Block, Window_Years) with
            Count_Below_Historical, Total_Metrics, Pct_Below_Historical.
    """
    detail_cols = [
        "Group",
        "Metric",
        "Metric_Label",
        "Block",
        "Window_Years",
        "Block_Min_RollingAvg_TAF",
        "Historical_Min_RollingAvg_TAF",
        "Below_Historical",
    ]
    count_cols = [
        "Block",
        "Window_Years",
        "Count_Below_Historical",
        "Total_Metrics",
        "Pct_Below_Historical",
    ]

    if annual_long.empty:
        return (
            pd.DataFrame(columns=detail_cols),
            pd.DataFrame(columns=count_cols),
        )

    metric_key_set = {str(k).strip().upper() for k in exclude_metric_keys}
    label_set = {str(l).strip().casefold() for l in exclude_labels}

    df = annual_long.copy()
    if metric_key_set:
        df = df[~df["Metric"].astype(str).str.strip().str.upper().isin(metric_key_set)]
    if label_set:
        df = df[~df["Metric_Label"].astype(str).str.strip().str.casefold().isin(label_set)]

    if df.empty:
        return (
            pd.DataFrame(columns=detail_cols),
            pd.DataFrame(columns=count_cols),
        )

    benchmark_df = df[df["Scenario"] == benchmark_name]
    blocks_df = df[df["Block_Index"].notna()]

    # Historical worst rolling average per (Metric, Window_Years)
    hist_lookup: Dict[Tuple[str, int], float] = {}
    for metric, m_df in benchmark_df.groupby("Metric", sort=False):
        for w in window_years:
            best = _best_rolling_window(m_df, int(w))
            if best is not None:
                hist_lookup[(metric, int(w))] = float(best["RollingAvg_TAF"])

    detail_rows: List[dict] = []
    for (group, metric, label), m_df in blocks_df.groupby(
        ["Group", "Metric", "Metric_Label"], sort=False
    ):
        for block, b_df in m_df.groupby("Block", sort=True):
            for w in window_years:
                w_int = int(w)
                hist_val = hist_lookup.get((metric, w_int))
                best = _best_rolling_window(b_df, w_int)
                if best is None:
                    continue
                block_val = float(best["RollingAvg_TAF"])
                below = (
                    hist_val is not None
                    and not pd.isna(block_val)
                    and not pd.isna(hist_val)
                    and block_val < hist_val
                )
                detail_rows.append(
                    {
                        "Group": group,
                        "Metric": metric,
                        "Metric_Label": label,
                        "Block": block,
                        "Window_Years": w_int,
                        "Block_Min_RollingAvg_TAF": block_val,
                        "Historical_Min_RollingAvg_TAF": (
                            float(hist_val) if hist_val is not None else float("nan")
                        ),
                        "Below_Historical": bool(below),
                    }
                )

    details = pd.DataFrame(detail_rows, columns=detail_cols)
    if details.empty:
        return details, pd.DataFrame(columns=count_cols)

    grouped = details.groupby(["Block", "Window_Years"], sort=False)
    counts = grouped.agg(
        Count_Below_Historical=("Below_Historical", "sum"),
        Total_Metrics=("Below_Historical", "size"),
    ).reset_index()
    counts["Count_Below_Historical"] = counts["Count_Below_Historical"].astype(int)
    counts["Total_Metrics"] = counts["Total_Metrics"].astype(int)
    counts["Pct_Below_Historical"] = np.where(
        counts["Total_Metrics"] > 0,
        100.0 * counts["Count_Below_Historical"] / counts["Total_Metrics"],
        0.0,
    )

    block_order = sort_block_labels(counts["Block"].unique())
    block_rank = {b: i for i, b in enumerate(block_order)}
    counts["_block_rank"] = counts["Block"].map(block_rank)
    counts = counts.sort_values(["_block_rank", "Window_Years"]).drop(columns="_block_rank")
    counts = counts[count_cols].reset_index(drop=True)

    details["_block_rank"] = details["Block"].map(block_rank)
    details = details.sort_values(
        ["Group", "Metric", "_block_rank", "Window_Years"]
    ).drop(columns="_block_rank").reset_index(drop=True)

    return details, counts


def plot_block_rolling_below_historical_counts(
    counts: pd.DataFrame,
    out_png: str | Path,
    window_years: Sequence[int] = (2, 5, 10),
) -> Path | None:
    """Grouped bar chart of below-historical counts per block per window."""
    if counts.empty:
        return None

    block_order = sort_block_labels(counts["Block"].unique())
    if not block_order:
        return None

    pivot = counts.pivot_table(
        index="Block",
        columns="Window_Years",
        values="Count_Below_Historical",
        aggfunc="sum",
        fill_value=0,
    ).reindex(block_order)

    windows = [int(w) for w in window_years if int(w) in pivot.columns]
    if not windows:
        return None
    pivot = pivot[windows]

    n_blocks = len(block_order)
    n_windows = len(windows)
    x = np.arange(n_blocks)
    bar_width = 0.8 / max(n_windows, 1)

    style_colors = {
        "background": "#FFFFFF",
        "horizontal_gridlines": "#D9E2EA",
        "left_bottom_axis": "#888888",
        "top_right_border": "#F5F8FA",
        "axis_numbers": "#222222",
    }
    bar_colors = {
        2: "#D27E2A",
        5: "#8C8C8C",
        10: "#0B3D59",
    }
    fallback = ["#4C72B0", "#DD8452", "#55A467"]
    colors = {
        w: bar_colors.get(w, fallback[i % len(fallback)])
        for i, w in enumerate(windows)
    }

    fig, ax = plt.subplots(figsize=(8.0, 6.5), dpi=200)
    fig.patch.set_facecolor(style_colors["background"])
    ax.set_facecolor(style_colors["background"])
    for i, w in enumerate(windows):
        offsets = x - 0.4 + bar_width * (i + 0.5)
        values = pivot[w].to_numpy(dtype=float)
        bars = ax.bar(
            offsets,
            values,
            width=bar_width,
            color=colors[w],
            edgecolor=style_colors["top_right_border"],
            linewidth=0.5,
            label=f"{w}-year",
        )
        for bar, val in zip(bars, values):
            if val <= 0:
                continue
            ax.text(
                bar.get_x() + bar.get_width() / 2.0,
                bar.get_height(),
                f"{int(round(val))}",
                ha="center",
                va="bottom",
                fontsize=8,
                color=style_colors["axis_numbers"],
            )

    ax.set_xticks(x)
    ax.set_xticklabels(block_order)
    ax.set_xlabel("")
    ax.set_ylabel("")
    total_metrics_vals = (
        pd.to_numeric(counts["Total_Metrics"], errors="coerce").dropna().unique()
        if "Total_Metrics" in counts.columns
        else np.array([])
    )
    if total_metrics_vals.size == 1:
        title_suffix = f" (out of {int(total_metrics_vals[0])} metrics)"
    elif total_metrics_vals.size > 1:
        title_suffix = (
            f" (out of {int(total_metrics_vals.min())}-"
            f"{int(total_metrics_vals.max())} metrics)"
        )
    else:
        title_suffix = ""
    ax.set_title(
        "Count of Metrics with Block Rolling Minimum Below Historical"
        + title_suffix,
        color=bar_colors[10],
        fontweight="bold",
    )
    ax.tick_params(
        axis="both",
        colors=style_colors["axis_numbers"],
        labelcolor=style_colors["axis_numbers"],
        width=1.2,
        length=4,
    )
    ax.spines["left"].set_color(style_colors["left_bottom_axis"])
    ax.spines["left"].set_linewidth(1.5)
    ax.spines["bottom"].set_color(style_colors["left_bottom_axis"])
    ax.spines["bottom"].set_linewidth(1.5)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))
    ax.grid(
        axis="y",
        color=style_colors["horizontal_gridlines"],
        linestyle="-",
        linewidth=1.2,
        alpha=1.0,
    )
    ax.set_axisbelow(True)
    legend_handles = [
        Line2D(
            [0],
            [0],
            marker="s",
            linestyle="None",
            markersize=8,
            markerfacecolor=colors[w],
            markeredgecolor=colors[w],
            label=f"{w}-year",
        )
        for w in windows
    ]
    legend = ax.legend(
        handles=legend_handles,
        loc="upper center",
        bbox_to_anchor=(0.5, -0.12),
        ncol=3,
        frameon=False,
    )
    if legend is not None:
        plt.setp(legend.get_texts(), color=style_colors["axis_numbers"])

    fig.tight_layout()
    out_path = Path(out_png)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(
        out_path,
        dpi=300,
        bbox_inches="tight",
        facecolor=style_colors["background"],
    )
    plt.close(fig)
    return out_path


# ---------------------------------------------------------------------------
# Compact summary table (one row per metric)
# ---------------------------------------------------------------------------

def build_compact_summary_table(
    benchmark_summary: pd.DataFrame,
    block_summary: pd.DataFrame,
    benchmark_rolling: pd.DataFrame,
    stochastic_rolling: pd.DataFrame,
    rolling_windows: Sequence[int] = COMPACT_SUMMARY_WINDOWS,
) -> pd.DataFrame:
    """Build a wide one-row-per-metric summary with annual avg and rolling minima.

    Columns
    -------
    Group, Metric, Metric_Label
    Hist_Annual_Avg               -- historical mean annual value
    N1_10_Annual_Avg_Min          -- minimum block mean across N1-N10
    N1_10_Annual_Avg_Max          -- maximum block mean across N1-N10
    N1_10_Abs_Range               -- max block mean - min block mean
    N1_10_Pct_Range               -- abs range / |hist mean| * 100
    N1_10_Annual_Avg_Pct_Diff_Min -- min block mean pct diff vs historical mean
    N1_10_Annual_Avg_Pct_Diff_Max -- max block mean pct diff vs historical mean
    N1_10_Annual_Avg_Bracket      -- display string: min block mean - max block mean
    N1_10_Annual_Avg_Pct_Diff_Bracket -- display string: min pct diff - max pct diff
    Hist_Min                      -- historical worst single-year value (from benchmark_summary)
    N1_10_Min                     -- stochastic worst single-year value (worst across all blocks)
    Hist_{w}yr_Min                -- historical worst w-yr rolling avg
    N1_10_{w}yr_Min               -- stochastic worst w-yr rolling avg (single worst across all blocks)
    """
    table = benchmark_summary[["Group", "Metric", "Metric_Label", "Historical_Mean_WY_TAF"]].copy()
    table = table.rename(columns={"Historical_Mean_WY_TAF": "Hist_Annual_Avg"})

    # Range of block means across N1-N10
    block_range = (
        block_summary.groupby(["Group", "Metric", "Metric_Label"])["Mean_WY_TAF"]
        .agg(N1_10_Annual_Avg_Min="min", N1_10_Annual_Avg_Max="max")
        .reset_index()
    )
    block_range["N1_10_Abs_Range"] = block_range["N1_10_Annual_Avg_Max"] - block_range["N1_10_Annual_Avg_Min"]
    table = table.merge(block_range, on=["Group", "Metric", "Metric_Label"], how="left")
    table["N1_10_Pct_Range"] = np.where(
        table["Hist_Annual_Avg"].ne(0),
        table["N1_10_Abs_Range"] / table["Hist_Annual_Avg"].abs() * 100.0,
        np.nan,
    )
    table["N1_10_Annual_Avg_Pct_Diff_Min"] = np.where(
        table["Hist_Annual_Avg"].ne(0),
        (table["N1_10_Annual_Avg_Min"] - table["Hist_Annual_Avg"])
        / table["Hist_Annual_Avg"].abs() * 100.0,
        np.nan,
    )
    table["N1_10_Annual_Avg_Pct_Diff_Max"] = np.where(
        table["Hist_Annual_Avg"].ne(0),
        (table["N1_10_Annual_Avg_Max"] - table["Hist_Annual_Avg"])
        / table["Hist_Annual_Avg"].abs() * 100.0,
        np.nan,
    )

    def _value_bracket(row: pd.Series) -> str:
        vmin = row.get("N1_10_Annual_Avg_Min")
        vmax = row.get("N1_10_Annual_Avg_Max")
        if pd.isna(vmin) or pd.isna(vmax):
            return ""
        return f"{vmin:,.1f} - {vmax:,.1f}"

    def _pct_bracket(row: pd.Series) -> str:
        pmin = row.get("N1_10_Annual_Avg_Pct_Diff_Min")
        pmax = row.get("N1_10_Annual_Avg_Pct_Diff_Max")
        if pd.isna(pmin) or pd.isna(pmax):
            return ""
        return f"{pmin:+.1f}% - {pmax:+.1f}%"

    table["N1_10_Annual_Avg_Bracket"] = table.apply(_value_bracket, axis=1)
    table["N1_10_Annual_Avg_Pct_Diff_Bracket"] = table.apply(_pct_bracket, axis=1)

    # Single-year minimum: pull directly from summary tables (not rolling)
    hist_min_1yr = benchmark_summary[["Group", "Metric", "Metric_Label", "Historical_Min_WY_TAF"]].rename(
        columns={"Historical_Min_WY_TAF": "Hist_Min"}
    )
    table = table.merge(hist_min_1yr, on=["Group", "Metric", "Metric_Label"], how="left")

    stoch_min_1yr = (
        block_summary.groupby(["Group", "Metric", "Metric_Label"])["Min_WY_TAF"]
        .min()
        .reset_index()
        .rename(columns={"Min_WY_TAF": "N1_10_Min"})
    )
    table = table.merge(stoch_min_1yr, on=["Group", "Metric", "Metric_Label"], how="left")

    # Rolling minima columns
    for w in sorted(rolling_windows):
        # Historical
        hist_w = benchmark_rolling[benchmark_rolling["Window_Years"] == w][
            ["Group", "Metric", "Metric_Label", "Historical_Min_RollingAvg_TAF"]
        ].rename(columns={"Historical_Min_RollingAvg_TAF": f"Hist_{w}yr_Min"})
        table = table.merge(hist_w, on=["Group", "Metric", "Metric_Label"], how="left")

        # Stochastic (single worst across all blocks)
        stoch_w = stochastic_rolling[stochastic_rolling["Window_Years"] == w][
            ["Group", "Metric", "Metric_Label", "Min_RollingAvg_TAF"]
        ].rename(columns={"Min_RollingAvg_TAF": f"N1_10_{w}yr_Min"})
        table = table.merge(stoch_w, on=["Group", "Metric", "Metric_Label"], how="left")

    return table.sort_values(["Group", "Metric"]).reset_index(drop=True)


# ---------------------------------------------------------------------------
# 100-year block mean range vs historical
# ---------------------------------------------------------------------------

# Display label and canonical Metric_Label used for lookup. Labels are
# case/punctuation/space normalized at match time.
RANGE_FIGURE_GROUPS: Tuple[Tuple[str, Tuple[Tuple[str, Tuple[str, ...]], ...]], ...] = (
    (
        "DELTA",
        (
            ("Delta Inflow", ("Delta Inflow",)),
            ("Delta Outflow", ("Delta Outflow",)),
            ("SAC River at Freeport", ("SAC River at Freeport",)),
            ("Total Jones Exports", ("Total Jones Exports",)),
            ("Total Banks Exports", ("Total Banks Exports",)),
            ("San Joaquin River at Vernalis", ("San Joaquin River at Vernalis",)),
        ),
    ),
    (
        "DELIVERY",
        (
            ("SWP Total Delivery", ("SWP Total Delivery",)),
            ("CVP Total Delivery (North)", ("CVP North of Delta Delivery",)),
            ("CVP Total Delivery (South)", ("CVP South of Delta Delivery",)),
        ),
    ),
    (
        "STORAGE",
        (
            ("Oroville", ("Oroville Storage",)),
            ("Shasta", ("Shasta Storage",)),
            ("San Luis - SWP", ("San Luis Storage SWP",)),
            ("San Luis - CVP", ("San Luis Storage CVP",)),
        ),
    ),
)

# Groups of metric labels whose CDF y-axes should share the same scale and start at zero.
# Each inner tuple lists the canonical Metric_Label strings (as they appear in fields.pkl)
# for all metrics in the comparable group.
COMPARABLE_CDF_GROUPS: Tuple[Tuple[str, ...], ...] = (
    ("SWP Total Delivery", "CVP North of Delta Delivery", "CVP South of Delta Delivery","CVP Total Delivery"),
)

_RANGE_NORMALIZE_RE = re.compile(r"[^a-z0-9]+")


def _normalize_label(label: object) -> str:
    """Lowercase, strip punctuation, and compress whitespace for matching."""
    if label is None:
        return ""
    text = str(label).strip().lower()
    text = _RANGE_NORMALIZE_RE.sub(" ", text)
    return " ".join(text.split())


def _resolve_range_metric_rows(
    compact_summary: pd.DataFrame,
) -> List[Tuple[str, str, pd.Series]]:
    """Map requested display labels to rows in compact_summary in canonical order."""
    if "Metric_Label" not in compact_summary.columns:
        raise ValueError(
            "compact_summary is missing the 'Metric_Label' column required for the range figure."
        )

    norm_to_index: Dict[str, int] = {}
    for idx, raw in compact_summary["Metric_Label"].items():
        key = _normalize_label(raw)
        if key and key not in norm_to_index:
            norm_to_index[key] = idx

    resolved: List[Tuple[str, str, pd.Series]] = []
    missing: List[str] = []
    for group_display, metrics in RANGE_FIGURE_GROUPS:
        for display_label, aliases in metrics:
            row_idx = None
            for alias in aliases:
                key = _normalize_label(alias)
                if key in norm_to_index:
                    row_idx = norm_to_index[key]
                    break
            if row_idx is None:
                missing.append(display_label)
                continue
            resolved.append(
                (group_display, display_label, compact_summary.loc[row_idx])
            )

    if missing:
        available = sorted(
            {str(v) for v in compact_summary["Metric_Label"].dropna().unique()}
        )
        raise ValueError(
            "Could not resolve the following metrics for the 100-year block mean "
            "range figure: "
            + ", ".join(missing)
            + ". Available Metric_Label values: "
            + ", ".join(available)
        )
    return resolved


def _ensure_range_pct_diffs(row: pd.Series) -> Tuple[float, float]:
    """Return (pct_min, pct_max) from existing columns or compute from raw values."""
    pmin = row.get("N1_10_Annual_Avg_Pct_Diff_Min", np.nan)
    pmax = row.get("N1_10_Annual_Avg_Pct_Diff_Max", np.nan)
    if pd.isna(pmin) or pd.isna(pmax):
        hist = row.get("Hist_Annual_Avg", np.nan)
        vmin = row.get("N1_10_Annual_Avg_Min", np.nan)
        vmax = row.get("N1_10_Annual_Avg_Max", np.nan)
        if (
            not pd.isna(hist)
            and float(hist) != 0.0
            and not pd.isna(vmin)
            and not pd.isna(vmax)
        ):
            denom = abs(float(hist))
            pmin = (float(vmin) - float(hist)) / denom * 100.0
            pmax = (float(vmax) - float(hist)) / denom * 100.0
    return float(pmin), float(pmax)


def plot_100yr_block_mean_range_vs_historical(
    compact_summary: pd.DataFrame,
    out_dir: str | Path,
) -> Dict[str, str]:
    """Presentation-quality range/lollipop plot of 100-year block means vs historical.

    Each row spans the min-to-max percent difference of the ten Product B
    100-year block means relative to the historical mean. Rows where the
    range crosses below zero are colored amber/orange; rows entirely above
    zero are blue.

    Returns a dict with keys 'png' and 'svg' mapping to saved file paths.
    Files are written under ``<out_dir>/figures/annual_block_range/``.
    """
    if compact_summary is None or compact_summary.empty:
        raise ValueError("compact_summary is empty; cannot build range figure.")

    resolved = _resolve_range_metric_rows(compact_summary)

    records: List[dict] = []
    for group_display, display_label, row in resolved:
        pmin, pmax = _ensure_range_pct_diffs(row)
        records.append(
            {
                "group": group_display,
                "label": display_label,
                "pmin": pmin,
                "pmax": pmax,
            }
        )

    n = len(records)
    # First record at the top of the chart -> highest y value.
    y_positions = list(range(n, 0, -1))

    # Visual palette.
    color_blue = "#1F77B4"
    color_orange = "#D27E2A"
    text_color = "#222222"
    muted_text = "#5A6B7A"
    grid_color = "#D9E2EA"
    band_color = "#EEF3F7"
    fig = plt.figure(figsize=(15.5, 7.0), dpi=200)
    fig.patch.set_facecolor("white")
    ax = fig.add_axes([0.28, 0.10, 0.60, 0.80])
    ax.set_facecolor("white")

    # Determine x range from the data, with padding and tidy 5% rounding.
    pmins = [r["pmin"] for r in records if not pd.isna(r["pmin"])]
    pmaxs = [r["pmax"] for r in records if not pd.isna(r["pmax"])]
    if pmins and pmaxs:
        data_min = min(pmins)
        data_max = max(pmaxs)
    else:
        data_min, data_max = -10.0, 40.0
    span = max(data_max - data_min, 1.0)
    pad = max(span * 0.08, 2.0)
    x_lo = float(np.floor(min(data_min - pad, -5.0) / 5.0) * 5.0)

    right_pad = max(span * 0.22, 8.0)
    x_hi = float(np.ceil((data_max + right_pad) / 5.0) * 5.0)
    ax.set_xlim(x_lo, x_hi)
    ax.set_ylim(0.4, n + 0.6)

    # Group bands (light fill behind each category) + group labels.
    group_bounds: Dict[str, Tuple[int, int]] = {}
    for rec, y in zip(records, y_positions):
        g = rec["group"]
        if g not in group_bounds:
            group_bounds[g] = (y, y)
        else:
            lo, hi = group_bounds[g]
            group_bounds[g] = (min(lo, y), max(hi, y))

    band_x = x_lo - (x_hi - x_lo) * 0.40
    label_x = x_lo - (x_hi - x_lo) * 0.36
    for g, (lo_y, hi_y) in group_bounds.items():
        ax.add_patch(
            plt.Rectangle(
                (band_x, lo_y - 0.5),
                (x_hi - band_x),
                (hi_y - lo_y) + 1.0,
                facecolor=band_color,
                edgecolor="none",
                zorder=0,
                clip_on=False,
            )
        )
        center_y = (lo_y + hi_y) / 2.0
        ax.text(
            label_x,
            center_y,
            g,
            ha="center",
            va="center",
            rotation=90,
            fontsize=14,
            fontweight="bold",
            color=text_color,
            clip_on=False,
        )

    # Subtle horizontal separators between groups.
    seen: List[str] = []
    for rec in records:
        if rec["group"] not in seen:
            seen.append(rec["group"])
    for i in range(1, len(seen)):
        prev_group = seen[i - 1]
        boundary = group_bounds[prev_group][0] - 0.5
        ax.axhline(boundary, color="#C7D2DC", linewidth=0.8, zorder=1)

    # Vertical reference at 0% (historical mean).
    ax.axvline(0.0, color="#3C4A57", linewidth=1.0, zorder=2)

    # X-axis grid + ticks formatted as percentages.
    ax.xaxis.set_major_locator(mticker.MultipleLocator(10))
    ax.xaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _p: f"{int(round(v))}%")
    )
    ax.grid(axis="x", color=grid_color, linewidth=0.8, alpha=1.0, zorder=1)
    ax.set_axisbelow(True)

    # Range segments + endpoint markers + annotations.
    for rec, y in zip(records, y_positions):
        pmin = rec["pmin"]
        pmax = rec["pmax"]
        if pd.isna(pmin) or pd.isna(pmax):
            continue
        c = color_orange if pmin < 0.0 else color_blue
        ax.plot(
            [pmin, pmax],
            [y, y],
            color=c,
            linewidth=3.0,
            solid_capstyle="round",
            zorder=4,
        )
        ax.scatter(
            [pmin, pmax],
            [y, y],
            s=70,
            color=c,
            edgecolors="white",
            linewidths=1.2,
            zorder=5,
        )
        ax.text(
            pmax + (x_hi - x_lo) * 0.012,
            y,
            f"{pmin:+.0f}% to {pmax:+.0f}%",
            ha="left",
            va="center",
            fontsize=12,
            color=text_color,
        )

    # Y tick labels = metric display labels.
    ax.set_yticks(y_positions)
    ax.set_yticklabels(
        [r["label"] for r in records], fontsize=13, color=text_color
    )
    ax.tick_params(axis="y", length=0, pad=6, labelsize=13)
    ax.tick_params(axis="x", colors=text_color, labelcolor=text_color, labelsize=12)
    for side in ("top", "right", "left"):
        ax.spines[side].set_visible(False)
    ax.spines["bottom"].set_color("#9AA7B3")
    ax.spines["bottom"].set_linewidth(1.0)

    ax.set_xlabel(
        "Range of 100-year block means relative to historical mean (%)",
        fontsize=13,
        color=muted_text,
        labelpad=8,
    )

    # Figure-level title.
    fig.text(
        0.04,
        0.955,
        "Range of 100-Year Block Means vs Historical",
        fontsize=22,
        fontweight="bold",
        color="#0B3D59",
        ha="left",
        va="center",
    )

    # Outputs.
    out_root = Path(out_dir)
    fig_dir = out_root / "figures" / "annual_block_range"
    fig_dir.mkdir(parents=True, exist_ok=True)
    png_path = fig_dir / "range_100yr_block_means_vs_historical.png"
    svg_path = fig_dir / "range_100yr_block_means_vs_historical.svg"
    fig.savefig(png_path, dpi=200, facecolor="white", bbox_inches="tight", pad_inches=0.15)
    try:
        fig.savefig(svg_path, facecolor="white", bbox_inches="tight", pad_inches=0.15)
        svg_out = str(svg_path)
    except Exception:
        # SVG export is optional; PNG is the primary output. Swallow any
        # backend-specific failure and report an empty SVG path instead.
        svg_out = ""
    plt.close(fig)
    return {"png": str(png_path), "svg": svg_out}


# ---------------------------------------------------------------------------
# Heatmap: metric-by-block % diff vs benchmark
# ---------------------------------------------------------------------------

def _percentile(x: Iterable[float], percentile: float) -> float:
    values = pd.Series(list(x), dtype=float).dropna().to_numpy(dtype=float)
    if values.size == 0:
        return float("nan")
    return float(np.percentile(values, percentile))


def build_heatmap_data(
    annual_long: pd.DataFrame,
    benchmark_name: str,
    min_window_years: int = 2,
    max_window_years: int = 10,
    drought_percentile: float = DEFAULT_DROUGHT_PERCENTILE,
    rolling_window_years: Sequence[int] = HEATMAP_ROLLING_WINDOWS,
) -> pd.DataFrame:
    """Build a long-form DataFrame with % diff vs benchmark for key stats.

    Stats:
    - Mean Annual
    - P5 Annual
    - Worst 5-year Rolling Avg
    - Worst 10-year Rolling Avg
    """
    benchmark = annual_long[annual_long["Scenario"] == benchmark_name].copy()
    blocks = annual_long[annual_long["Block_Index"].notna()].copy()

    bm_mean = benchmark.groupby(["Group", "Metric", "Metric_Label"])["WY_Value"].mean()
    bm_low = benchmark.groupby(["Group", "Metric", "Metric_Label"])["WY_Value"].agg(
        lambda s: _percentile(s, drought_percentile)
    )

    blk_mean = blocks.groupby(["Group", "Metric", "Metric_Label", "Block"])["WY_Value"].mean()
    blk_low = blocks.groupby(["Group", "Metric", "Metric_Label", "Block"])["WY_Value"].agg(
        lambda s: _percentile(s, drought_percentile)
    )

    def _pct_diff(block_stat: pd.Series, bm_stat: pd.Series) -> pd.DataFrame:
        df = block_stat.reset_index().rename(columns={"WY_Value": "Block_Value"})
        bm = bm_stat.reset_index().rename(columns={"WY_Value": "BM_Value"})
        df = df.merge(bm, on=["Group", "Metric", "Metric_Label"], how="left", validate="many_to_one")
        df["Pct_Diff"] = np.where(
            df["BM_Value"].ne(0),
            (df["Block_Value"] - df["BM_Value"]) / df["BM_Value"].abs() * 100.0,
            np.nan,
        )
        return df

    mean_diff = _pct_diff(blk_mean, bm_mean)
    mean_diff["Stat"] = "Mean Annual"

    drought_label = f"P{drought_percentile:g} Annual"
    drought_diff = _pct_diff(blk_low, bm_low)
    drought_diff["Stat"] = drought_label

    rolling_frames: List[pd.DataFrame] = []
    for window_years in sorted({int(w) for w in rolling_window_years if int(w) > 0}):
        if window_years < min_window_years or window_years > max_window_years:
            continue

        bm_rolling: Dict[Tuple[str, str, str], float] = {}
        for (group, metric, label), mdf in benchmark.groupby(["Group", "Metric", "Metric_Label"], sort=False):
            best = _best_rolling_window(mdf, window_years)
            bm_rolling[(group, metric, label)] = float(best["RollingAvg_TAF"]) if best is not None else np.nan

        blk_rows: List[dict] = []
        for (group, metric, label, scenario, block, block_idx), bdf in blocks.groupby(
            ["Group", "Metric", "Metric_Label", "Scenario", "Block", "Block_Index"],
            sort=False,
        ):
            best = _best_rolling_window(bdf, window_years)
            blk_min = float(best["RollingAvg_TAF"]) if best is not None else np.nan
            bm_val = bm_rolling.get((group, metric, label), np.nan)
            pct = ((blk_min - bm_val) / abs(bm_val) * 100.0) if bm_val != 0 and not np.isnan(bm_val) else np.nan
            blk_rows.append(
                {
                    "Group": group,
                    "Metric": metric,
                    "Metric_Label": label,
                    "Block": block,
                    "Block_Value": blk_min,
                    "BM_Value": bm_val,
                    "Pct_Diff": pct,
                    "Stat": f"Worst {window_years}-yr Rolling Avg",
                }
            )

        if blk_rows:
            rolling_frames.append(pd.DataFrame(blk_rows))

    all_frames = [mean_diff, drought_diff] + rolling_frames
    if not all_frames:
        return pd.DataFrame(columns=["Group", "Metric", "Metric_Label", "Block", "Stat", "Pct_Diff", "Block_Value", "BM_Value"])

    heatmap_long = pd.concat(all_frames, ignore_index=True)
    heatmap_long = heatmap_long[["Group", "Metric", "Metric_Label", "Block", "Stat", "Pct_Diff", "Block_Value", "BM_Value"]]
    return heatmap_long.sort_values(["Stat", "Group", "Metric", "Block"]).reset_index(drop=True)


def filter_heatmap_metrics(
    heatmap_long: pd.DataFrame,
    excluded_metric_keys: Sequence[str] = HEATMAP_EXCLUDED_METRIC_KEYS,
    excluded_labels: Sequence[str] = HEATMAP_EXCLUDED_LABELS,
) -> pd.DataFrame:
    """Remove lower-priority variables from heatmap outputs only.

    The annual, rolling, and compact summary tables keep all metrics. This
    filter is deliberately applied only to the heatmap workbook and heatmap
    figures. Use exact metric keys / labels.
    """
    if heatmap_long.empty:
        return heatmap_long.copy()

    exclude = pd.Series(False, index=heatmap_long.index)

    metric_key_set = {str(key).strip().upper() for key in excluded_metric_keys}
    label_set = {str(label).strip().casefold() for label in excluded_labels}

    if "Metric" in heatmap_long.columns and metric_key_set:
        metrics = heatmap_long["Metric"].astype(str).str.strip().str.upper()
        exclude |= metrics.isin(metric_key_set)

    if "Metric_Label" in heatmap_long.columns and label_set:
        labels = heatmap_long["Metric_Label"].astype(str).str.strip().str.casefold()
        exclude |= labels.isin(label_set)

    return heatmap_long.loc[~exclude].copy()


def plot_heatmap(
    heatmap_long: pd.DataFrame,
    out_dir: str | Path,
    benchmark_name: str = "Historical",
) -> Dict[str, str]:
    """Generate one heatmap PNG per stat.

    Positive values are intentionally mapped to blue.
    """
    import matplotlib.colors as mcolors

    out_dir = Path(out_dir)
    fig_dir = out_dir / "figures" / "heatmap"
    fig_dir.mkdir(parents=True, exist_ok=True)

    outputs: Dict[str, str] = {}

    if heatmap_long.empty:
        return outputs

    for stat_name, stat_df in heatmap_long.groupby("Stat", sort=False):
        pivot = stat_df.pivot_table(
            index="Metric_Label",
            columns="Block",
            values="Pct_Diff",
            aggfunc="first",
        )
        if pivot.empty:
            continue

        block_cols = sort_block_labels(list(pivot.columns))
        pivot = pivot[block_cols]

        values = pivot.to_numpy(dtype=float)
        finite_values = values[np.isfinite(values)]
        if finite_values.size == 0:
            continue

        n_rows, n_cols = pivot.shape
        fig_height = max(4.0, 0.4 * n_rows + 1.5)
        fig_width = max(6.0, 0.8 * n_cols + 3.0)
        fig, ax = plt.subplots(figsize=(fig_width, fig_height))

        vmax = max(float(np.max(np.abs(finite_values))), 1.0)
        norm = mcolors.TwoSlopeNorm(vmin=-vmax, vcenter=0, vmax=vmax)
        cmap = plt.cm.RdBu  # negative = red, positive = blue

        im = ax.imshow(values, aspect="auto", cmap=cmap, norm=norm)

        ax.set_xticks(range(n_cols))
        ax.set_xticklabels(block_cols, fontsize=8)
        ax.set_yticks(range(n_rows))
        ax.set_yticklabels(pivot.index, fontsize=7)

        for i in range(n_rows):
            for j in range(n_cols):
                val = values[i, j]
                if np.isfinite(val):
                    color = "white" if abs(val) > vmax * 0.6 else "black"
                    ax.text(j, i, f"{val:+.1f}%", ha="center", va="center", fontsize=6, color=color)

        ax.set_title(
            f"{stat_name} - Percent Difference from {benchmark_name}",
            fontsize=10,
            fontweight="bold",
        )
        ax.set_xlabel("Product B Block")
        cbar = fig.colorbar(im, ax=ax, shrink=0.8, pad=0.02)
        cbar.set_label(f"% Diff vs {benchmark_name}", fontsize=8)

        fig.tight_layout()
        safe_name = make_safe_filename(str(stat_name).replace("%", "pct").lower())
        out_png = fig_dir / f"heatmap_{safe_name}.png"
        fig.savefig(out_png, dpi=300, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        outputs[str(stat_name)] = str(out_png)

    return outputs


# ---------------------------------------------------------------------------
# CDF plots
# ---------------------------------------------------------------------------

def empirical_cdf(x: Iterable[float]) -> Tuple[np.ndarray, np.ndarray]:
    values = np.asarray(list(x), dtype=float)
    values = values[np.isfinite(values)]
    if values.size == 0:
        return np.array([]), np.array([])
    xs = np.sort(values)
    p = np.arange(1, xs.size + 1) / xs.size * 100.0
    return p, xs


def plot_annual_cdf(
    annual_long: pd.DataFrame,
    metric_key: str,
    metric_label: str,
    benchmark_name: str,
    out_png: str | Path,
    unit: str = "TAF",
    y_lim: tuple[float, float] | None = None,
) -> None:
    df_metric = annual_long[annual_long["Metric"] == metric_key].copy()

    fig, ax = plt.subplots(figsize=(6.5, 4.0))

    benchmark = df_metric[df_metric["Scenario"] == benchmark_name]
    p, xs = empirical_cdf(benchmark["WY_Value"].to_numpy(dtype=float))
    ax.plot(p, xs, linewidth=2.0, label=benchmark_name, color="black")

    block_order = (
        df_metric[df_metric["Block_Index"].notna()][["Scenario", "Block", "Block_Index"]]
        .drop_duplicates()
        .sort_values("Block_Index")
    )

    for _, row in block_order.iterrows():
        scenario = row["Scenario"]
        block_label = row["Block"]
        sub = df_metric[df_metric["Scenario"] == scenario]
        p, xs = empirical_cdf(sub["WY_Value"].to_numpy(dtype=float))
        ax.plot(p, xs, linewidth=1.1, alpha=0.9, label=block_label)

    ax.set_xlabel("Non-Exceedance Probability (%)")
    ax.set_ylabel(unit)
    ax.set_xlim(0, 100)
    if y_lim is not None:
        ax.set_ylim(y_lim)
    ax.set_title(f"{metric_label} - Annual CDF")
    ax.grid(True, linewidth=0.35, alpha=0.4)
    ax.legend(ncol=3, fontsize=7, frameon=False, loc="lower right")
    fig.tight_layout()

    out_path = Path(out_png)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=300, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def plot_monthly_cdf(
    df_values: pd.DataFrame,
    metric_key: str,
    metric_label: str,
    benchmark_name: str,
    out_png: str | Path,
    unit: str = "TAF",
    y_lim: tuple[float, float] | None = None,
) -> None:
    """Plot a monthly CDF using all monthly values for each scenario.

    This mirrors the annual CDF styling but does not aggregate to water years;
    each scenario contributes its complete monthly time series for the metric.
    """
    if metric_key not in df_values.columns or "Scenario" not in df_values.columns:
        return

    df_metric = df_values[["Scenario", metric_key]].dropna().copy()
    if df_metric.empty:
        return

    fig, ax = plt.subplots(figsize=(6.5, 4.0))

    benchmark = df_metric[df_metric["Scenario"] == benchmark_name]
    p, xs = empirical_cdf(benchmark[metric_key].to_numpy(dtype=float))
    if xs.size:
        ax.plot(p, xs, linewidth=2.0, label=benchmark_name, color="black")

    block_order = (
        df_metric[["Scenario"]]
        .drop_duplicates()
        .assign(Block_Index=lambda d: d["Scenario"].map(extract_block_index))
        .dropna(subset=["Block_Index"])
        .sort_values("Block_Index")
    )

    for _, row in block_order.iterrows():
        scenario = row["Scenario"]
        block_label = block_label_from_index(row["Block_Index"])
        sub = df_metric[df_metric["Scenario"] == scenario]
        p, xs = empirical_cdf(sub[metric_key].to_numpy(dtype=float))
        if xs.size:
            ax.plot(p, xs, linewidth=1.1, alpha=0.9, label=block_label)

    ax.set_xlabel("Non-Exceedance Probability (%)")
    ax.set_ylabel(unit)
    ax.set_xlim(0, 100)
    if y_lim is not None:
        ax.set_ylim(y_lim)
    ax.set_title(f"{metric_label} - Monthly CDF")
    ax.grid(True, linewidth=0.35, alpha=0.4)
    ax.legend(ncol=3, fontsize=7, frameon=False, loc="lower right")
    fig.tight_layout()

    out_path = Path(out_png)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=300, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def plot_summary_boxplots(
    annual_long: pd.DataFrame,
    fields: Dict[str, str],
    out_dir: str | Path,
    benchmark_name: str = "Historical",
    unit: str = "TAF",
) -> Dict[str, str]:
    """Create per-metric boxplots with one historical box and one box per Product B block.

    Each box contains annual water-year values.
    """
    out_dir = Path(out_dir)
    fig_dir = out_dir / "figures" / "block_boxplots"
    fig_dir.mkdir(parents=True, exist_ok=True)

    outputs: Dict[str, str] = {}
    if annual_long.empty:
        return outputs

    # Plot color palette
    _BLUE = "#003D6B"
    _LIGHT_BLUE = "#B0C4DE"
    _GRAY = "#5A5A5A"
    _HIST_FILL = "#F2F2F2"

    for metric_key, mdf in annual_long.groupby("Metric", sort=False):
        blocks_df = mdf[mdf["Block_Index"].notna()].copy()
        bench_df = mdf[mdf["Scenario"] == benchmark_name]
        if blocks_df.empty and bench_df.empty:
            continue

        block_labels = sort_block_labels(blocks_df["Block"].unique())
        hist_vals = bench_df["WY_Value"].dropna().to_numpy(dtype=float)
        hist_included = hist_vals.size > 0

        block_data = [
            blocks_df.loc[blocks_df["Block"] == bl, "WY_Value"].dropna().to_numpy(dtype=float)
            for bl in block_labels
        ]

        data: List[np.ndarray] = []
        plot_labels: List[str] = []
        if hist_included:
            data.append(hist_vals)
            plot_labels.append("Hist")
        data.extend(block_data)
        plot_labels.extend(block_labels)

        if not data or all(d.size == 0 for d in data):
            continue

        metric_label = metric_label_from_fields(metric_key, fields)
        fig, ax = plt.subplots(figsize=(10.5, 5))

        positions = list(range(1, len(plot_labels) + 1))
        bp = ax.boxplot(
            data,
            positions=positions,
            widths=0.5,
            showfliers=True,
            patch_artist=True,
            showmeans=True,
            medianprops=dict(color=_GRAY, linewidth=1.8),
            meanprops=dict(
                marker="D",
                markerfacecolor=_BLUE,
                markeredgecolor=_BLUE,
                markersize=5,
            ),
            whiskerprops=dict(color=_BLUE, linewidth=1.0),
            capprops=dict(color=_BLUE, linewidth=1.0),
            flierprops=dict(
                marker="o",
                markerfacecolor=_GRAY,
                markeredgecolor=_GRAY,
                markersize=3,
                alpha=0.5,
            ),
        )

        for i, box in enumerate(bp["boxes"]):
            if hist_included and i == 0:
                box.set(facecolor=_HIST_FILL, edgecolor=_GRAY, linewidth=1.3)
            else:
                box.set(facecolor=_LIGHT_BLUE, edgecolor=_BLUE, linewidth=1.0)

        if hist_included:
            hist_mean = float(np.nanmean(hist_vals))
            ax.axhline(
                hist_mean,
                color=_GRAY,
                linestyle="--",
                linewidth=1.4,
                label=f"{benchmark_name} Mean ({hist_mean:,.0f} {unit})",
            )

        ax.set_xticks(positions)
        ax.set_xticklabels(plot_labels, fontsize=11, fontweight="medium")
        ax.set_xlabel("Historical / Product B Block", fontsize=12, fontweight="bold", labelpad=8)
        ax.set_ylabel(f"Annual Value ({unit})", fontsize=12, fontweight="bold", labelpad=8)
        ax.set_title(
            f"{metric_label} - Product B Block Distribution",
            fontsize=14,
            fontweight="bold",
            pad=12,
        )
        ax.tick_params(axis="both", labelsize=11)
        ax.grid(True, axis="y", linewidth=0.3, alpha=0.5, color="#CCCCCC")
        ax.set_axisbelow(True)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        if hist_included:
            ax.legend(
                frameon=True,
                loc="upper right",
                fontsize=10,
                edgecolor="#CCCCCC",
                fancybox=False,
                framealpha=0.9,
            )

        fig.tight_layout()
        out_png = fig_dir / f"{make_safe_filename(metric_key)}.png"
        fig.savefig(out_png, dpi=300, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        outputs[str(metric_key)] = str(out_png)

    return outputs


def _build_centered_sequence_frame(
    metric_df: pd.DataFrame,
    window_years: int,
    frame_years: int,
) -> dict | None:
    if frame_years < window_years:
        raise ValueError(
            f"sequence_frame_years ({frame_years}) must be >= window_years ({window_years})."
        )

    ordered = metric_df.sort_values("WY").reset_index(drop=True)
    best = _best_rolling_window(ordered, window_years)
    if best is None:
        return None

    window_start_pos = (frame_years - window_years) // 2 + 1  # 1-based within display frame
    before_slots = window_start_pos - 1
    after_slots = frame_years - window_years - before_slots

    source_start_pos = int(best["Seq_Start"]) - before_slots

    frame_values = np.full(frame_years, np.nan)
    frame_wy: List[int | None] = [None] * frame_years

    for frame_pos in range(1, frame_years + 1):
        source_pos = source_start_pos + frame_pos - 1  # still 1-based
        if 1 <= source_pos <= len(ordered):
            row = ordered.iloc[source_pos - 1]
            frame_values[frame_pos - 1] = float(row["WY_Value"])
            frame_wy[frame_pos - 1] = int(row["WY"])

    return {
        "frame_x": np.arange(1, frame_years + 1, dtype=int),
        "frame_values": frame_values,
        "frame_wy": frame_wy,
        "window_pos_start": window_start_pos,
        "window_pos_end": window_start_pos + window_years - 1,
        "seq_start": int(best["Seq_Start"]),
        "seq_end": int(best["Seq_End"]),
        "wy_start": int(best["WY_Start"]),
        "wy_end": int(best["WY_End"]),
        "rolling_avg": float(best["RollingAvg_TAF"]),
    }


def _plot_sequence_trace(
    ax: plt.Axes,
    sequence_frame: dict,
    label: str,
    color: str | tuple | None = None,
    full_linewidth: float = 1.0,
    window_linewidth: float = 2.4,
    full_alpha: float = 0.35,
    window_alpha: float = 0.95,
    zorder: float = 2.0,
    marker: str = "o",
    markersize: float = 4.0,
) -> None:
    """Plot a faint full trace and put the legend entry on the highlighted window."""
    x = sequence_frame["frame_x"]
    y = sequence_frame["frame_values"]
    ax.plot(
        x,
        y,
        label="_nolegend_",
        color=color,
        linewidth=full_linewidth,
        alpha=full_alpha,
        zorder=zorder,
    )

    mask = (
        (x >= int(sequence_frame["window_pos_start"]))
        & (x <= int(sequence_frame["window_pos_end"]))
        & np.isfinite(y)
    )
    if np.any(mask):
        ax.plot(
            x[mask],
            y[mask],
            label=label,
            color=color,
            linewidth=window_linewidth,
            alpha=window_alpha,
            marker=marker,
            markersize=markersize,
            zorder=zorder + 0.1,
        )


def plot_worst_window_sequences(
    annual_long: pd.DataFrame,
    fields: Dict[str, str],
    out_dir: str | Path,
    benchmark_name: str = "Historical",
    units: Dict[str, str] | None = None,
    sequence_window_years: Sequence[int] = DEFAULT_SEQUENCE_WINDOWS,
    sequence_frame_years: int = DEFAULT_SEQUENCE_FRAME_YEARS,
) -> Dict[str, str]:
    """Create per-metric overlays of the worst historical and all block drought windows.

    A 2-year window is plotted in a compact 6-year frame with the middle two years
    highlighted. Other windows use ``sequence_frame_years`` with the window centered.
    """
    units = units or {}
    out_dir = Path(out_dir)
    fig_root = out_dir / "figures" / "worst_window_sequences"
    fig_root.mkdir(parents=True, exist_ok=True)

    outputs: Dict[str, str] = {}
    window_years_list = sorted({int(w) for w in sequence_window_years if int(w) > 0})

    def _frame_years_for_window(window_years: int) -> int:
        return TWO_YEAR_SEQUENCE_FRAME_YEARS if window_years == 2 else sequence_frame_years

    for window_years in window_years_list:
        frame_years = _frame_years_for_window(window_years)
        if frame_years < window_years:
            raise ValueError(
                f"sequence frame length ({frame_years}) must be >= window length ({window_years})."
            )

        window_dir = fig_root / f"{window_years}yr"
        window_dir.mkdir(parents=True, exist_ok=True)

        for metric_key in annual_long["Metric"].drop_duplicates():
            df_metric = annual_long[annual_long["Metric"] == metric_key].copy()
            if df_metric.empty:
                continue

            benchmark_df = df_metric[df_metric["Scenario"] == benchmark_name].copy()
            benchmark_frame = _build_centered_sequence_frame(
                benchmark_df,
                window_years=window_years,
                frame_years=frame_years,
            )

            block_frames: List[Tuple[str, int, dict]] = []
            block_order = (
                df_metric[df_metric["Block_Index"].notna()][["Scenario", "Block", "Block_Index"]]
                .drop_duplicates()
                .sort_values("Block_Index")
            )
            for _, row in block_order.iterrows():
                block_df = df_metric[df_metric["Scenario"] == row["Scenario"]].copy()
                block_frame = _build_centered_sequence_frame(
                    block_df,
                    window_years=window_years,
                    frame_years=frame_years,
                )
                if block_frame is None:
                    continue
                block_frames.append((str(row["Block"]), int(row["Block_Index"]), block_frame))

            if benchmark_frame is None and not block_frames:
                continue

            fig_width = 9.0 if frame_years <= 6 else 12.0
            fig, ax = plt.subplots(figsize=(fig_width, 5.5))

            # Shaded critical window with subtle edge lines
            window_pos_start = (frame_years - window_years) // 2 + 1
            window_pos_end = window_pos_start + window_years - 1
            ax.axvspan(window_pos_start - 0.5, window_pos_end + 0.5,
                       color="#E8EDF2", alpha=0.95, zorder=0)
            ax.axvline(window_pos_start - 0.5, color="#B0B0B0", linewidth=0.6,
                       linestyle="--", alpha=0.6, zorder=0.5)
            ax.axvline(window_pos_end + 0.5, color="#B0B0B0", linewidth=0.6,
                       linestyle="--", alpha=0.6, zorder=0.5)

            # Historical benchmark trace (bold, square markers)
            if benchmark_frame is not None:
                hist_name = "Hist" if benchmark_name.lower().startswith("hist") else benchmark_name
                hist_label = (
                    f"{hist_name} (WY {benchmark_frame['wy_start']}"
                    f"-{benchmark_frame['wy_end']}"
                    f", avg {benchmark_frame['rolling_avg']:,.0f})"
                )
                _plot_sequence_trace(
                    ax,
                    benchmark_frame,
                    label=hist_label,
                    color="black",
                    full_linewidth=1.6,
                    window_linewidth=3.0,
                    full_alpha=0.5,
                    window_alpha=1.0,
                    zorder=3.0,
                    marker="s",
                    markersize=4.5,
                )

            # Identify the driest block as the one with the lowest rolling_avg
            # over its worst window. All other blocks are drawn in gray.
            worst_block_index = None
            if block_frames:
                worst_block_index = min(
                    block_frames,
                    key=lambda bf: bf[2]["rolling_avg"],
                )[1]

            # Block traces: gray for non-highlight, crimson for the driest block.
            # Only the driest block is added to the legend; non-highlight gray
            # blocks are collapsed into a single proxy entry below.
            for block_label, block_index, block_frame in block_frames:
                is_worst = block_index == worst_block_index
                color = "crimson" if is_worst else "0.65"
                if is_worst:
                    label = (
                        f"{block_label} (driest, Yr {block_frame['seq_start']}"
                        f"-{block_frame['seq_end']}"
                        f", avg {block_frame['rolling_avg']:,.0f})"
                    )
                else:
                    # "_nolegend_" prefix prevents this trace from appearing
                    # in get_legend_handles_labels(); the proxy handle below
                    # represents all gray blocks.
                    label = "_nolegend_"
                _plot_sequence_trace(
                    ax,
                    block_frame,
                    label=label,
                    color=color,
                    full_linewidth=1.4 if is_worst else 0.9,
                    window_linewidth=2.6 if is_worst else 1.6,
                    full_alpha=0.6 if is_worst else 0.25,
                    window_alpha=1.0 if is_worst else 0.6,
                    zorder=2.5 if is_worst else 1.5,
                    marker="o",
                    markersize=4.0 if is_worst else 3.0,
                )

            metric_label = metric_label_from_fields(metric_key, fields)
            unit = units.get(metric_key, "TAF")
            ax.set_title(
                f"{metric_label} - Worst {window_years}-Year Sequence Comparison",
                fontsize=14,
                fontweight="bold",
                pad=12,
            )
            ax.set_xlabel(f"Year (relative to worst {window_years}-yr window)",
                          fontsize=12, fontweight="bold", labelpad=8)
            ax.set_ylabel(f"Annual Value ({unit})", fontsize=12, fontweight="bold", labelpad=8)
            ax.tick_params(axis="both", labelsize=11)
            ax.set_xlim(0.5, frame_years + 0.5)
            ax.set_xticks(np.arange(1, frame_years + 1, dtype=int))
            ax.grid(True, axis="both", linewidth=0.3, alpha=0.4, color="#CCCCCC")
            ax.set_axisbelow(True)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)

            # Annotation for shaded region is now represented by a legend patch.

            # Build a compact custom legend:
            #   - Historical (black square trace, label includes WY range/avg)
            #   - Driest block (crimson, label includes Yr range/avg)
            #   - Other blocks (n=N) (gray proxy line)
            #   - Critical window (light gray patch)
            from matplotlib.lines import Line2D
            from matplotlib.patches import Patch

            handles, labels = ax.get_legend_handles_labels()
            n_other_blocks = sum(
                1 for _, idx, _ in block_frames if idx != worst_block_index
            )
            if n_other_blocks > 0:
                handles.append(
                    Line2D(
                        [0], [0],
                        color="0.65", linewidth=1.6, alpha=0.6,
                        marker="o", markersize=3.0,
                    )
                )
                labels.append("Other blocks")
            handles.append(
                Patch(facecolor="#E8EDF2", edgecolor="#B0B0B0", linewidth=0.6)
            )
            labels.append(f"{window_years}-yr critical window")

            if handles:
                ax.legend(
                    handles,
                    labels,
                    ncol=1,
                    fontsize=9,
                    frameon=True,
                    edgecolor="#CCCCCC",
                    fancybox=False,
                    framealpha=0.95,
                    loc="upper right",
                )

            fig.tight_layout()
            out_png = window_dir / f"{make_safe_filename(metric_key)}_worst_{window_years}yr.png"
            fig.savefig(out_png, dpi=300, bbox_inches="tight", facecolor="white")
            plt.close(fig)
            outputs[f"{metric_key}:{window_years}"] = str(out_png)

    return outputs


# ---------------------------------------------------------------------------
# 1000-year stitched time series
# ---------------------------------------------------------------------------

def plot_1000yr_timeseries(
    annual_long: pd.DataFrame,
    fields: Dict[str, str],
    out_dir: str | Path,
    benchmark_name: str = "Historical",
    unit: str = "TAF",
    product_a_annual: pd.DataFrame | None = None,
    product_a_label: str = "Product A",
    product_a_start_wy: int | None = None,
) -> Dict[str, str]:
    """Create a stitched 1000-year time series plot for each metric.

    The historical benchmark trace is plotted first, then blocks n01-n10
    are concatenated sequentially with a continuous sequence index.
    The historical benchmark mean is shown as a horizontal reference line.

    If ``product_a_annual`` is provided (columns: Metric, WY, WY_Value), an
    additional trace is overlaid in the historical region. Each Product A WY
    is aligned to the historical x-axis by water year (so a 50-year Product A
    sequence appears at the same x-position as the matching historical years).
    """
    out_dir = Path(out_dir)
    fig_dir = out_dir / "figures" / "timeseries_1000yr"
    fig_dir.mkdir(parents=True, exist_ok=True)

    _BLUE = "#003D6B"
    _PRODUCT_B_TRACE = "#8FBAD6"
    _HIST_LINE = "#3F3F3F"
    _HIST_TRACE = "#B8B8B8"
    _BLOCK_BORDER = "#B3B3B3"
    _HIST_BAND = "#F3F3F3"
    _PRODUCT_A_TRACE = "#C0392B"

    has_product_a = (
        product_a_annual is not None
        and not product_a_annual.empty
        and {"Metric", "WY", "WY_Value"}.issubset(product_a_annual.columns)
    )

    outputs: Dict[str, str] = {}
    if annual_long.empty:
        return outputs

    for metric_key, mdf in annual_long.groupby("Metric", sort=False):
        blocks_df = mdf[mdf["Block_Index"].notna()].copy()
        bench_df = mdf[mdf["Scenario"] == benchmark_name]
        if blocks_df.empty:
            continue

        block_labels = sort_block_labels(blocks_df["Block"].unique())

        hist_df = bench_df.sort_values("WY").dropna(subset=["WY_Value"])
        hist_vals = hist_df["WY_Value"].to_numpy(dtype=float)
        hist_wys = hist_df["WY"].astype(int).to_numpy()
        hist_years = len(hist_vals)
        hist_x_arr = np.arange(1, hist_years + 1, dtype=int)

        # Stitch blocks sequentially
        seq_x: List[int] = []
        seq_y: List[float] = []
        block_boundaries: List[int] = []
        offset = hist_years

        for bl in block_labels:
            bdata = blocks_df[blocks_df["Block"] == bl].sort_values("WY")
            vals = bdata["WY_Value"].to_numpy(dtype=float)
            n = len(vals)
            seq_x.extend(range(offset + 1, offset + n + 1))
            seq_y.extend(vals)
            offset += n
            block_boundaries.append(offset)

        if not seq_x:
            continue

        seq_x_arr = np.array(seq_x)
        seq_y_arr = np.array(seq_y)
        stochastic_years = len(seq_y_arr)
        total_years = hist_years + stochastic_years

        metric_label = metric_label_from_fields(metric_key, fields)
        fig, ax = plt.subplots(figsize=(16, 5.2))

        if hist_years:
            ax.axvspan(0.5, hist_years + 0.5, color=_HIST_BAND,
                       alpha=1.0, linewidth=0, zorder=0)

        if hist_years:
            ax.plot(hist_x_arr, hist_vals, color=_HIST_TRACE, linewidth=0.8,
                    alpha=0.75, zorder=3, label=f"{benchmark_name} annual")

            if hist_years >= 10:
                hist_rolling = pd.Series(hist_vals).rolling(10, min_periods=10).mean().to_numpy()
                ax.plot(hist_x_arr, hist_rolling, color=_HIST_LINE, linewidth=1.5,
                        alpha=0.9, zorder=4, label=f"{benchmark_name} 10-yr rolling")

            ax.axvline(hist_years + 0.5, color=_HIST_LINE, linewidth=1.0,
                       linestyle="--", alpha=0.55, zorder=1)

        # Product A overlay (aligned to historical water years)
        if has_product_a and hist_years:
            pa_metric = product_a_annual[product_a_annual["Metric"] == metric_key]
            pa_metric = pa_metric.dropna(subset=["WY", "WY_Value"]).sort_values("WY")
            if not pa_metric.empty:
                pa_wys = pa_metric["WY"].astype(int).to_numpy()
                pa_vals = pa_metric["WY_Value"].to_numpy(dtype=float)
                if product_a_start_wy is not None:
                    keep = pa_wys >= int(product_a_start_wy)
                    pa_wys = pa_wys[keep]
                    pa_vals = pa_vals[keep]
                # Map each PA water year to its x position via the explicit
                # hist_wys -> hist_x_arr index. This is robust to gaps in the
                # benchmark series (hist_df drops null annual values, so WYs
                # are not necessarily contiguous).
                wy_to_x = {int(wy): int(x) for wy, x in zip(hist_wys, hist_x_arr)}
                pa_pairs = [(wy_to_x[w], v) for w, v in zip(pa_wys, pa_vals) if int(w) in wy_to_x]
                if pa_pairs:
                    pa_x = np.array([p[0] for p in pa_pairs], dtype=int)
                    pa_vals = np.array([p[1] for p in pa_pairs], dtype=float)
                else:
                    pa_x = np.empty(0, dtype=int)
                    pa_vals = np.empty(0, dtype=float)
                if pa_x.size >= 10:
                    pa_rolling = pd.Series(pa_vals).rolling(10, min_periods=10).mean().to_numpy()
                    ax.plot(pa_x, pa_rolling, color=_PRODUCT_A_TRACE, linewidth=1.5,
                            alpha=0.95, zorder=5,
                            label=f"{product_a_label} 10-yr rolling ({pa_x.size}-yr)")

        # Block boundary lines
        for i, boundary in enumerate(block_boundaries[:-1]):
            ax.axvline(boundary + 0.5, color=_BLOCK_BORDER, linewidth=0.6,
                       linestyle=":", alpha=0.6, zorder=1)

        # Annual time series
        ax.plot(seq_x_arr, seq_y_arr, color=_PRODUCT_B_TRACE, linewidth=0.55,
                alpha=0.78, zorder=2, label="Product B annual")

        # 10-year rolling average
        if stochastic_years >= 10:
            rolling = pd.Series(seq_y_arr).rolling(10, min_periods=10).mean().to_numpy()
            ax.plot(seq_x_arr, rolling, color=_BLUE, linewidth=1.5,
                    alpha=0.95, zorder=3, label="Product B 10-yr rolling")

        if stochastic_years:
            product_b_mean = float(np.nanmean(seq_y_arr))
            ax.hlines(product_b_mean, xmin=hist_years + 0.5, xmax=total_years + 0.5,
                      color=_BLUE, linestyle="--", linewidth=1.2,
                      alpha=0.9, label=f"Product B mean ({product_b_mean:,.0f} {unit})", zorder=4)

        # Historical benchmark mean
        if not bench_df.empty:
            hist_mean = bench_df["WY_Value"].mean()
            ax.hlines(hist_mean, xmin=1, xmax=total_years,
                      color=_HIST_LINE, linestyle=":", linewidth=1.3,
                      alpha=0.9, label=f"{benchmark_name} mean ({hist_mean:,.0f} {unit})", zorder=4)

        ax.set_xlim(0.5, total_years + 0.5)
        ax.set_xlabel("Historical water year / Product B sequence index", fontsize=11, fontweight="bold", labelpad=8)
        ax.set_ylabel(f"Annual value ({unit})", fontsize=11, fontweight="bold", labelpad=8)
        ax.set_title(metric_label, fontsize=14, fontweight="bold", loc="center", pad=26)
        ax.text(0.5, 1.015,
            f"Historical water years followed by {stochastic_years}-year Product B stochastic sequence",
                transform=ax.transAxes, ha="center", va="bottom",
                fontsize=9, color="#555555")

        tick_positions: List[int] = []
        tick_labels: List[str] = []
        if hist_years:
            tick_positions.append(1)
            tick_labels.append(str(hist_wys[0]))
            if hist_years > 1:
                tick_positions.append(hist_years)
                tick_labels.append(str(hist_wys[-1]))
        for product_b_index in range(100, stochastic_years + 1, 100):
            tick_positions.append(hist_years + product_b_index)
            tick_labels.append(str(product_b_index))
        ax.set_xticks(tick_positions)
        ax.set_xticklabels(tick_labels)

        minor_positions = [
            hist_years + product_b_index
            for product_b_index in range(50, stochastic_years + 1, 50)
            if product_b_index % 100 != 0
        ]
        ax.xaxis.set_minor_locator(mticker.FixedLocator(minor_positions))
        ax.tick_params(axis="both", labelsize=9, color="#555555", labelcolor="#333333")
        ax.yaxis.set_major_formatter(mticker.StrMethodFormatter("{x:,.0f}"))
        ax.grid(True, axis="y", linewidth=0.35, alpha=0.45, color="#D0D0D0")
        ax.set_axisbelow(True)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_color("#777777")
        ax.spines["bottom"].set_color("#777777")
        ax.margins(y=0.08)
        ax.legend(fontsize=7.5, frameon=True, loc="upper right",
                ncol=1, handlelength=2.0, columnspacing=0.9,
                handletextpad=0.45, borderpad=0.35,
                edgecolor="#D0D0D0", fancybox=False, framealpha=0.92)

        # Block labels at top of each block region
        if hist_years:
            y_top = ax.get_ylim()[1]
            y_range = y_top - ax.get_ylim()[0]
            ax.text((1 + hist_years) / 2, y_top - y_range * 0.03, benchmark_name,
                    ha="center", va="top", fontsize=7, color="#555555", alpha=0.95,
                    fontweight="bold")

        prev = hist_years
        for bl, boundary in zip(block_labels, block_boundaries):
            mid = (prev + boundary) / 2
            y_top = ax.get_ylim()[1]
            y_range = y_top - ax.get_ylim()[0]
            ax.text(mid, y_top - y_range * 0.03, bl,
                    ha="center", va="top", fontsize=7, color="#777777", alpha=0.85)
            prev = boundary

        fig.tight_layout(rect=[0, 0, 1, 0.96])
        out_png = fig_dir / f"{make_safe_filename(metric_key)}.png"
        fig.savefig(out_png, dpi=300, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        outputs[str(metric_key)] = str(out_png)

    return outputs


# ---------------------------------------------------------------------------
# Excel formatting
# ---------------------------------------------------------------------------

def format_excel_workbook(path: str | Path) -> None:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    centered = Alignment(horizontal="center", vertical="center")
    left_aligned = Alignment(horizontal="left", vertical="center")

    integer_headers = {
        "wy",
        "wy_start",
        "wy_end",
        "block_index",
        "window_years",
        "historical_years",
        "block_years",
        "historical_year_index_start",
        "historical_year_index_end",
        "year_index",
        "year_index_start",
        "year_index_end",
    }

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered

        column_widths: Dict[str, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                text = "" if cell.value is None else str(cell.value)
                width = min(max(len(text) + 2, 10), 42)
                column_widths[cell.column_letter] = max(column_widths.get(cell.column_letter, 0), width)

        for letter, width in column_widths.items():
            ws.column_dimensions[letter].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = str(ws.cell(row=1, column=cell.column).value or "").lower()
                if header in {"group", "metric", "metric_label", "scenario", "block"}:
                    cell.alignment = left_aligned
                else:
                    cell.alignment = centered

                if "pct" in header:
                    cell.number_format = "0.0"
                elif header in integer_headers or header.startswith("frame_pos_"):
                    cell.number_format = "0"
                elif any(token in header for token in ["taf", "mean", "median", "min", "max", "std", "rollingavg"]):
                    cell.number_format = "#,##0.0"

    wb.save(workbook_path)


# ---------------------------------------------------------------------------
# Package runner
# ---------------------------------------------------------------------------

def run_post_processing_package(
    pickle_dir: str | Path,
    benchmark_name: str,
    out_dir: str | Path,
    min_window_years: int = 2,
    max_window_years: int = 10,
    drought_percentile: float = DEFAULT_DROUGHT_PERCENTILE,
    sequence_window_years: Sequence[int] = DEFAULT_SEQUENCE_WINDOWS,
    sequence_frame_years: int = DEFAULT_SEQUENCE_FRAME_YEARS,
    product_a_pickle_dir: str | Path | None = None,
    product_a_label: str = "Product A",
    product_a_start_wy: int | None = 1972,
) -> Dict[str, str]:
    df_values, _df_diffs, units, fields = load_pickles(pickle_dir)
    df_values["Date"] = pd.to_datetime(df_values["Date"])

    metric_groups = metric_groups_from_fields(fields)
    metric_keys = [col for col in df_values.columns if col not in FIXED_COLS]

    annual_long = annualize_all_metrics(
        df_values=df_values,
        metric_keys=metric_keys,
        metric_groups=metric_groups,
        fields=fields,
    )

    benchmark_summary = benchmark_summary_table(annual_long=annual_long, benchmark_name=benchmark_name)
    block_summary = block_summary_table(annual_long=annual_long, benchmark_name=benchmark_name)

    benchmark_rolling = benchmark_rolling_minima_table(
        annual_long=annual_long,
        benchmark_name=benchmark_name,
        min_window_years=min_window_years,
        max_window_years=max_window_years,
    )
    stochastic_rolling = stochastic_rolling_minima_table(
        annual_long=annual_long,
        min_window_years=min_window_years,
        max_window_years=max_window_years,
    )

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    # -- Compact summary table (one row per metric) --
    compact_summary = build_compact_summary_table(
        benchmark_summary=benchmark_summary,
        block_summary=block_summary,
        benchmark_rolling=benchmark_rolling,
        stochastic_rolling=stochastic_rolling,
    )
    compact_summary_xlsx = out_path / "compact_summary.xlsx"
    with pd.ExcelWriter(compact_summary_xlsx, engine="openpyxl") as writer:
        compact_summary.to_excel(writer, sheet_name="summary", index=False)
    format_excel_workbook(compact_summary_xlsx)

    # -- 100-year block mean range vs historical (presentation figure) --
    try:
        range_figure_paths = plot_100yr_block_mean_range_vs_historical(
            compact_summary=compact_summary,
            out_dir=out_path,
        )
    except ValueError as exc:
        print(f"[range_100yr_block_means] skipped: {exc}")
        range_figure_paths = {"png": "", "svg": ""}

    # -- Heatmap: metric-by-block % diff vs benchmark --
    heatmap_long = build_heatmap_data(
        annual_long=annual_long,
        benchmark_name=benchmark_name,
        min_window_years=min_window_years,
        max_window_years=max_window_years,
        drought_percentile=drought_percentile,
        rolling_window_years=HEATMAP_ROLLING_WINDOWS,
    )
    heatmap_long = filter_heatmap_metrics(heatmap_long)

    heatmap_xlsx = out_path / "heatmap_block_summary.xlsx"
    with pd.ExcelWriter(heatmap_xlsx, engine="openpyxl") as writer:
        wrote_heatmap_sheet = False
        for stat_name, stat_df in heatmap_long.groupby("Stat", sort=False):
            pivot = stat_df.pivot_table(
                index="Metric_Label",
                columns="Block",
                values="Pct_Diff",
                aggfunc="first",
            )
            if pivot.empty:
                continue
            block_cols = sort_block_labels(list(pivot.columns))
            pivot = pivot[block_cols]
            safe_sheet = f"{stat_name} % Diff"[:31]
            pivot.to_excel(writer, sheet_name=safe_sheet)
            wrote_heatmap_sheet = True
        if not wrote_heatmap_sheet:
            pd.DataFrame({"Note": ["No heatmap metrics remained after filtering."]}).to_excel(
                writer, sheet_name="heatmap", index=False
            )
    format_excel_workbook(heatmap_xlsx)

    heatmap_pngs = plot_heatmap(heatmap_long=heatmap_long, out_dir=out_path, benchmark_name=benchmark_name)

    # -- Annual summary Excel --
    annual_summary_xlsx = out_path / "annual_block_summary.xlsx"
    with pd.ExcelWriter(annual_summary_xlsx, engine="openpyxl") as writer:
        benchmark_summary.to_excel(writer, sheet_name="historical_summary", index=False)
        block_summary.to_excel(writer, sheet_name="block_summary", index=False)
        annual_export = annual_long.drop(columns=["Scenario", "Block_Index"]).copy()
        annual_export["Year_Index"] = (
            annual_export.groupby(["Metric", "Block"]).cumcount() + 1
        )
        col_order = ["Group", "Metric", "Metric_Label", "Block", "Year_Index", "WY_Value"]
        annual_export = annual_export[[c for c in col_order if c in annual_export.columns]]
        annual_export.to_excel(writer, sheet_name="annual_values_long", index=False)
    format_excel_workbook(annual_summary_xlsx)

    # -- Rolling minima Excel --
    rolling_minima_xlsx = out_path / "rolling_minima.xlsx"
    with pd.ExcelWriter(rolling_minima_xlsx, engine="openpyxl") as writer:
        benchmark_rolling.to_excel(writer, sheet_name="historical_rolling_minima", index=False)
        stochastic_rolling.to_excel(writer, sheet_name="stochastic_rolling_minima", index=False)
    format_excel_workbook(rolling_minima_xlsx)

    # -- Block rolling minima below historical: counts and grouped bar chart --
    issue_windows: Tuple[int, ...] = (2, 5, 10)
    issue_details, issue_counts = build_block_rolling_below_historical_counts(
        annual_long=annual_long,
        benchmark_name=benchmark_name,
        window_years=issue_windows,
        exclude_metric_keys=BLOCK_ROLLING_COUNT_EXCLUDED_METRIC_KEYS,
    )
    block_rolling_issue_counts_xlsx = out_path / "rolling_minima_vs_historical_counts.xlsx"
    if not issue_details.empty:
        metrics_included = (
            issue_details[["Group", "Metric", "Metric_Label"]]
            .drop_duplicates()
            .sort_values(["Group", "Metric"])
            .reset_index(drop=True)
        )
    else:
        metrics_included = pd.DataFrame(columns=["Group", "Metric", "Metric_Label"])
    excluded_metric_keys_df = pd.DataFrame(
        {"Excluded_Metric_Key": list(BLOCK_ROLLING_COUNT_EXCLUDED_METRIC_KEYS)}
    )
    excluded_labels_df = pd.DataFrame(
        {"Excluded_Label": list(HEATMAP_EXCLUDED_LABELS)}
    )
    with pd.ExcelWriter(block_rolling_issue_counts_xlsx, engine="openpyxl") as writer:
        issue_counts.to_excel(writer, sheet_name="counts_by_block", index=False)
        issue_details.to_excel(writer, sheet_name="details", index=False)
        metrics_included.to_excel(writer, sheet_name="metrics_included", index=False)
        excluded_metric_keys_df.to_excel(
            writer, sheet_name="excluded_metric_keys", index=False
        )
        excluded_labels_df.to_excel(
            writer, sheet_name="excluded_labels", index=False
        )
    format_excel_workbook(block_rolling_issue_counts_xlsx)

    block_rolling_issue_counts_figure_dir = (
        out_path / "figures" / "rolling_minima_vs_historical_counts"
    )
    # Bar chart variants: user-selectable window subsets.
    # Default: one with all three windows (2, 5, 10), one without the 5-year window.
    issue_count_chart_variants: Sequence[Tuple[str, Tuple[int, ...]]] = (
        ("block_rolling_minima_below_historical_counts.png", (2, 5, 10)),
        ("block_rolling_minima_below_historical_counts_no5yr.png", (2, 10)),
    )
    block_rolling_issue_counts_figures: List[str] = []
    for fname, windows_subset in issue_count_chart_variants:
        fig_path = block_rolling_issue_counts_figure_dir / fname
        plot_block_rolling_below_historical_counts(
            counts=issue_counts,
            out_png=fig_path,
            window_years=windows_subset,
        )
        block_rolling_issue_counts_figures.append(str(fig_path))
    # Keep primary (all-windows) figure path for backward compatibility.
    block_rolling_issue_counts_figure = (
        block_rolling_issue_counts_figure_dir
        / issue_count_chart_variants[0][0]
    )

    # -- Annual CDFs --
    # Build per-metric y_lim overrides for comparable delivery groups.
    # fields values look like "GROUP: Display Label"; match on the display label
    # (text after the first colon) so COMPARABLE_CDF_GROUPS entries align.
    _label_to_key: Dict[str, str] = {
        metric_label_from_fields(k, fields): k for k in fields
    }
    _annual_y_lims: Dict[str, tuple[float, float]] = {}
    _monthly_y_lims: Dict[str, tuple[float, float]] = {}
    for _group_labels in COMPARABLE_CDF_GROUPS:
        _group_keys = [_label_to_key[lbl] for lbl in _group_labels if lbl in _label_to_key]
        if len(_group_keys) < 2:
            continue
        # Annual: find max WY_Value across all scenarios and all metrics in the group.
        _ann_vals = annual_long[annual_long["Metric"].isin(_group_keys)]["WY_Value"].dropna()
        _ann_ymax = float(_ann_vals.max()) if not _ann_vals.empty else None
        if _ann_ymax is not None and _ann_ymax > 0:
            for _k in _group_keys:
                _annual_y_lims[_k] = (0.0, _ann_ymax)
        # Monthly: find max raw monthly value across all scenarios and all metrics in the group.
        _mon_cols = [k for k in _group_keys if k in df_values.columns]
        if _mon_cols:
            _mon_ymax = float(df_values[_mon_cols].max().max())
            if _mon_ymax > 0:
                for _k in _group_keys:
                    _monthly_y_lims[_k] = (0.0, _mon_ymax)

    cdf_dir = out_path / "figures" / "annual_cdf"
    for metric_key in metric_keys:
        plot_annual_cdf(
            annual_long=annual_long,
            metric_key=metric_key,
            metric_label=metric_label_from_fields(metric_key, fields),
            benchmark_name=benchmark_name,
            out_png=cdf_dir / f"{make_safe_filename(metric_key)}.png",
            unit=units.get(metric_key, "TAF"),
            y_lim=_annual_y_lims.get(metric_key),
        )

    # -- Monthly CDFs: all monthly values on one CDF per metric --
    monthly_cdf_dir = out_path / "figures" / "monthly_cdf"
    for metric_key in metric_keys:
        plot_monthly_cdf(
            df_values=df_values,
            metric_key=metric_key,
            metric_label=metric_label_from_fields(metric_key, fields),
            benchmark_name=benchmark_name,
            out_png=monthly_cdf_dir / f"{make_safe_filename(metric_key)}.png",
            unit=units.get(metric_key, "TAF"),
            y_lim=_monthly_y_lims.get(metric_key),
        )

    # -- Boxplot summary figures --
    boxplot_pngs = plot_summary_boxplots(
        annual_long=annual_long,
        fields=fields,
        out_dir=out_path,
        benchmark_name=benchmark_name,
        unit="TAF",
    )

    # -- Worst-window sequence overlays --
    sequence_pngs = plot_worst_window_sequences(
        annual_long=annual_long,
        fields=fields,
        out_dir=out_path,
        benchmark_name=benchmark_name,
        units=units,
        sequence_window_years=sequence_window_years,
        sequence_frame_years=sequence_frame_years,
    )

    # -- 1000-year stitched time series --
    product_a_annual = None
    if product_a_pickle_dir:
        pa_dir = Path(product_a_pickle_dir)
        _required = ("values.pkl", "diffs.pkl", "units.pkl", "fields.pkl")
        _missing = [f for f in _required if not (pa_dir / f).exists()]
        if _missing:
            print(
                f"[plot_1000yr_timeseries] Product A pickle(s) not found at {pa_dir} "
                f"({', '.join(_missing)}); skipping overlay."
            )
        else:
            product_a_annual = load_product_a_annual(
                pickle_dir=pa_dir,
                metric_keys=metric_keys,
                metric_groups=metric_groups,
                fields=fields,
                exclude_scenarios=(benchmark_name,),
            )

    timeseries_pngs = plot_1000yr_timeseries(
        annual_long=annual_long,
        fields=fields,
        out_dir=out_path,
        benchmark_name=benchmark_name,
        unit="TAF",
        product_a_annual=product_a_annual,
        product_a_label=product_a_label,
        product_a_start_wy=product_a_start_wy,
    )

    return {
        "compact_summary_xlsx": str(compact_summary_xlsx),
        "annual_block_range_figure": range_figure_paths.get("png", ""),
        "annual_block_range_figure_svg": range_figure_paths.get("svg", ""),
        "heatmap_xlsx": str(heatmap_xlsx),
        "heatmap_figures": str(out_path / "figures" / "heatmap"),
        "annual_summary_xlsx": str(annual_summary_xlsx),
        "rolling_minima_xlsx": str(rolling_minima_xlsx),
        "block_rolling_issue_counts_xlsx": str(block_rolling_issue_counts_xlsx),
        "block_rolling_issue_counts_figure": str(block_rolling_issue_counts_figure),
        "block_rolling_issue_counts_figures": block_rolling_issue_counts_figures,
        "annual_cdf_dir": str(cdf_dir),
        "monthly_cdf_dir": str(monthly_cdf_dir),
        "block_boxplot_dir": str(out_path / "figures" / "block_boxplots"),
        "worst_window_sequence_dir": str(out_path / "figures" / "worst_window_sequences"),
        "timeseries_1000yr_dir": str(out_path / "figures" / "timeseries_1000yr"),
        "n_heatmap_figures": str(len(heatmap_pngs)),
        "n_boxplot_figures": str(len(boxplot_pngs)),
        "n_sequence_figures": str(len(sequence_pngs)),
        "n_timeseries_figures": str(len(timeseries_pngs)),
        "n_monthly_cdf_figures": str(len(metric_keys)),
        "n_metrics": str(len(metric_keys)),
        "n_rows_annual_values": str(len(annual_long)),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Run Product B stochastic post-processing.")
    parser.add_argument(
        "--pickle-dir",
        default=str(PICKLE_DIR),
        help="Directory containing values.pkl / diffs.pkl / units.pkl / fields.pkl",
    )
    parser.add_argument("--benchmark-name", default="Historical", help="Benchmark scenario name in values.pkl")
    parser.add_argument(
        "--out-dir",
        default=str(OUT_DIR),
        help="Output directory for Product B post-processing",
    )
    parser.add_argument("--min-window-years", type=int, default=2, help="Minimum rolling window in years")
    parser.add_argument("--max-window-years", type=int, default=10, help="Maximum rolling window in years")
    parser.add_argument(
        "--drought-percentile",
        type=float,
        default=DEFAULT_DROUGHT_PERCENTILE,
        help="Drought percentile used for the annual low-flow heatmap stat (default: 5)",
    )
    parser.add_argument(
        "--sequence-frame-years",
        type=int,
        default=DEFAULT_SEQUENCE_FRAME_YEARS,
        help="Frame length used for non-2-year worst-window sequence overlays (default: 15; 2-year uses 6)",
    )
    parser.add_argument(
        "--sequence-window-years",
        nargs="*",
        type=int,
        default=list(DEFAULT_SEQUENCE_WINDOWS),
        help="Rolling-window lengths to plot for the worst-window sequence overlays (default: 2 5)",
    )
    parser.add_argument(
        "--product-a-pickle-dir",
        default=str(PRODUCT_A_PICKLE_DIR),
        help=(
            "Directory containing Product A pickles "
            "(values.pkl, diffs.pkl, units.pkl, fields.pkl). "
            "Used to overlay the Product A trace on the 1000-yr timeseries figure. "
            "Pass an empty string to disable the overlay."
        ),
    )
    parser.add_argument(
        "--product-a-label",
        default="Product A",
        help="Legend label for the Product A overlay trace.",
    )
    parser.add_argument(
        "--product-a-start-wy",
        type=int,
        default=1972,
        help=(
            "Earliest water year to include from the Product A series when overlaying "
            "on the 1000-yr timeseries (default: 1972). Use 0 to disable filtering."
        ),
    )
    args = parser.parse_args()

    outputs = run_post_processing_package(
        pickle_dir=args.pickle_dir,
        benchmark_name=args.benchmark_name,
        out_dir=args.out_dir,
        min_window_years=args.min_window_years,
        max_window_years=args.max_window_years,
        drought_percentile=args.drought_percentile,
        sequence_window_years=args.sequence_window_years,
        sequence_frame_years=args.sequence_frame_years,
        product_a_pickle_dir=args.product_a_pickle_dir or None,
        product_a_label=args.product_a_label,
        product_a_start_wy=(args.product_a_start_wy if args.product_a_start_wy and args.product_a_start_wy > 0 else None),
    )

    print("Created:")
    for key, value in outputs.items():
        print(f"  {key}: {value}")


if __name__ == "__main__":
    main()
