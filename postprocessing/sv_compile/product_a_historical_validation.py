# %% ── Compile Product A Historical Validation DSS ───────────────────────────
"""
Consolidated Product A historical-validation compiler.

Workflow
--------
1. Scan each module's ``output/_product_a_validation/`` for data CSVs
2. Copy those CSVs into ``compiled_input_files/<module>/``
3. For each module, build an intermediate DSS by overriding the baseline
   ``__calsim_sv_default__.dss`` with the CSV values
4. Merge all intermediate DSS records into a single ``ProductA_Historical_Validation_SV.dss``
5. Cross-reference against the master inventory to report:
   - Expected (Part B, C) that were successfully modified
   - Expected (Part B, C) that are MISSING (not provided by any module)
   - Unexpected (Part B, C) modifications (not in the inventory)

Inventory flags
---------------
- **Missing = T** → SV does not exist in the baseline; skip entirely
- **Used in DCR = F** → SV not used in the DCR study; skip entirely
- **Constant/Rept = T** → repeat the last 12 WY months of the baseline across
  the full overwrite window (auto-generated, no CSV needed)

Diagnostic outputs  (all written to ``_product_a_validation/``)
------------------------------------------------------------------------
- ``compiled_input_files/<module>/*.csv``  — local copies of source CSVs
- ``<module>__productA.dss``               — per-module intermediate DSS
- ``_constant_rept__productA.dss``         — auto-filled Constant/Rept DSS
- ``ProductA_Historical_Validation_SV.dss``           — final compiled DSS
- ``paths_modified.csv``                   — every (Part B/C, module) modified
- ``paths_unchanged.csv``                  — baseline paths untouched by any module
- ``paths_not_in_baseline.csv``            — module paths absent from the baseline
- ``inventory_expected_modified.csv``      — inventory SVs successfully modified
- ``inventory_expected_missing.csv``       — inventory SVs that should have been
                                             modified but were not
- ``inventory_constant_rept.csv``          — auto-filled Constant/Rept SVs
- ``inventory_skipped_missing.csv``        — SVs skipped (Missing = T)
- ``inventory_skipped_not_in_dcr.csv``     — SVs skipped (Used in DCR = F)
- ``inventory_unexpected.csv``             — modifications not in the inventory
- ``modification_statistics.csv``          — per-path monthly & annual comparison
                                             stats (base vs modified) when enabled
- ``modification_statistics_report.txt``   — summary report by input category
- ``figures/r2_nse_by_category.png``       — R²/NSE box plots by category
- ``figures/pctchange_monthly_by_category.png`` — monthly % change heatmap
- ``figures/abschange_annual_by_category.png``  — annual abs-change box plots
- ``figures/r2_nse_scatter.png``           — R² vs NSE scatter by category
- ``figures/all-terms/``                   — per-variable plots (regular terms)
- ``figures/all-terms-constant-rept/``     — per-variable plots (12-month repeat)
- ``figures/all-terms-constant/``          — per-variable plots (constant value)
- ``compilation_summary.txt``              — full statistics

CLI flags
---------
- ``--compute-stats``  Skip compilation; compute per-path modification
  statistics from an existing compiled DSS and ``paths_modified.csv``.
  Errors out if the compiled DSS has not been created yet.
- ``--stats-report``   Generate summary report and figures from an existing
  ``modification_statistics.csv``.  Can be combined with ``--compute-stats``
  or run independently after stats have already been computed.
- ``--no-term-plots``  Skip per-variable diagnostic plots (time series + scatter).
"""

import os
import sys
import time
import shutil
import glob
import argparse
import subprocess
import atexit
import warnings
import numpy as np
import pandas as pd
from pathlib import Path
from collections import OrderedDict

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from utils.paths import get_base_dir, get_generated_dir, get_module_generated_dir, get_inventory_dir

# Suppress noisy warnings from empty slices (e.g. all-NaN columns in stats)
warnings.filterwarnings("ignore", message="Mean of empty slice")
warnings.filterwarnings("ignore", message="All-NaN slice encountered")
from pydsstools.heclib.dss import HecDss
from pydsstools.core import TimeSeriesContainer


# ══════════════════════════════════════════════════════════════════════════════
# Configuration
# ══════════════════════════════════════════════════════════════════════════════
_base = get_base_dir()
_gen  = get_module_generated_dir("postprocessing/sv_compile")

BASELINE_DSS = _base / "CalSim3" / "__calsim_sv_default__.dss"
OUTPUT_DIR   = _gen / "product_a_validation"
OUTPUT_DSS   = OUTPUT_DIR / "ProductA_Historical_Validation_SV.dss"
COMPILED_CSV = OUTPUT_DIR / "compiled_input_files"

INVENTORY_XLSX = get_inventory_dir() / "_MASTER_INVENTORY_FOR_STOCHASTIC_INPUT_GENERATION_.xlsx"

DSS_PATTERN = "/*/*/*/*/1MON/*"

# Overwrite window (inclusive end-of-month timestamps)
OVERWRITE_START = pd.Timestamp(1971, 10, 31)
OVERWRITE_END   = pd.Timestamp(2018,  9, 30)

# -- Junction helper for long DSS paths ----------------------------------------
# The Fortran HEC-DSS library inside pydsstools limits path names to 256 chars.
# The data directory may live on OneDrive with a very long path, so we create a
# temporary Windows directory junction under the repo root to shorten it.
# All GENERATED paths in this script are under OUTPUT_DIR, so a single junction
# covers all long paths.  BASELINE_DSS (under BASE/) is short enough as-is.

_REPO_ROOT = Path(__file__).resolve().parents[2]
_DSS_LINK = _REPO_ROOT / "_dss_link"
_PATH_LIMIT = 200  # conservative limit vs Fortran's 256-char CNAME


def _create_junction(target_dir):
    """Create (or re-create) a directory junction at _DSS_LINK -> target_dir."""
    if _DSS_LINK.exists():
        subprocess.run(["cmd", "/c", "rmdir", str(_DSS_LINK)], capture_output=True)
    subprocess.run(
        ["cmd", "/c", "mklink", "/J", str(_DSS_LINK), str(target_dir)],
        check=True, capture_output=True,
    )


def _remove_junction():
    """Remove the _DSS_LINK junction (does not affect target directory)."""
    if _DSS_LINK.exists():
        subprocess.run(["cmd", "/c", "rmdir", str(_DSS_LINK)], capture_output=True)


_USE_JUNCTION = len(str(OUTPUT_DSS)) > _PATH_LIMIT
if _USE_JUNCTION:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    _create_junction(OUTPUT_DIR)
    atexit.register(_remove_junction)


def _dss_str(path):
    """Return a DSS-safe string path, shortened via junction if active."""
    s = str(path)
    if not _USE_JUNCTION or len(s) <= _PATH_LIMIT:
        return s
    try:
        rel = Path(path).relative_to(OUTPUT_DIR)
        return str(_DSS_LINK / rel)
    except ValueError:
        return s

# ── CLI arguments ────────────────────────────────────────────────────────────
_parser = argparse.ArgumentParser(
    description="Product A Historical Validation — DSS Compilation",
)
_parser.add_argument(
    "--compute-stats", action="store_true", default=False,
    help="Skip compilation; compute per-path modification statistics from "
         "an existing compiled DSS and paths_modified.csv.",
)
_parser.add_argument(
    "--stats-report", action="store_true", default=False,
    help="Generate summary report and figures from modification_statistics.csv. "
         "Can be combined with --compute-stats or run standalone.",
)
_parser.add_argument(
    "--no-term-plots", action="store_true", default=False,
    help="Skip per-variable diagnostic plots (time series + scatter). "
         "By default these are generated as part of the stats report.",
)
CLI_ARGS = _parser.parse_args()

# ── Module definitions ───────────────────────────────────────────────────────
# label → (absolute path to validation dir, inventory Input_Category name)
_gen_dir = get_generated_dir()
MODULE_CONFIG = OrderedDict([
    ("calsimhydro",              (_gen_dir / "mod_hydrology/calsimhydro/output/_3_postprocess_product_a/_product_a_validation", 
        "CalSimHydro")),
    ("calsimhydro_ee",           (_gen_dir / "mod_hydrology/calsimhydro_ee/output/_2_postprocess_product_a/_product_a_validation",  
        "CalSimHydroEE")),
    ("evaporation",              (_gen_dir / "mod_reservoir/evaporation/output/_2_run_reservoir_evap/_product_a_validation",   
        "Reservoir Evaporation")),
    ("rim_inflow",               (_gen_dir / "mod_hydrology/rim_inflow/output/_2_qmap_historical_validation/_product_a_validation", 
        "Rim Inflow")),
    ("delta_channel_depletion",  (_gen_dir / "mod_hydrology/delta_channel_depletion/output/_2_postprocess_product_a/_product_a_validation", 
        "Delta Channel Depletion")),
    ("small_watersheds",         (_gen_dir / "mod_hydrology/small_watersheds/output/_2_postprocess_product_a/_product_a_validation", 
        "Small Watersheds")),
    ("storage_curves",           (_gen_dir / "mod_reservoir/storage_curves/output/_product_a_validation",  
        "Reservoir Storage Curves")),
    ("instream_flows",           (_gen_dir / "mod_other/instream_flows/output/_product_a_validation", 
        "Instream Flows")),
    ("tulare_gw_terms",          (_gen_dir / "mod_hydrology/tulare_gw_terms/output/_1_wyt_monthly_avg/product_a/_product_a_validation", 
        "Tulare Groundwater Terms")),
    ("climate",                  (_gen_dir / "mod_forcing/climate/output/_product_a_validation",
        "Climate")),
    ("miscellaneous",            (_gen_dir / "mod_other/miscellaneous/output/_product_a_validation", 
            "Other")),
    ("upper_watershed",          (_gen_dir / "mod_other/upper_watershed/output/_product_a_validation",
        "Upper Watershed Modules")),
])

# Inventory categories that are NOT expected to have validation modules
CATEGORIES_WITHOUT_VALIDATION = {"Closure Terms", "Day-Volume Fraction", "Salinity"}


# ══════════════════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════════════════
def excel_to_part(name: str) -> str:
    """Normalise a Part B/C value to upper-case, spaces → underscores."""
    return str(name).upper().replace(" ", "_")


def ym_to_eom(year: int, month: int) -> pd.Timestamp:
    """Year + Month → end-of-month timestamp."""
    return pd.Timestamp(year=int(year), month=int(month), day=1).to_period("M").to_timestamp("M")


def dss_eom(ts_pytimes) -> pd.DatetimeIndex:
    """Convert DSS monthly timestamps to end-of-month (shifted back one month)."""
    return (pd.to_datetime(ts_pytimes).to_period("M") - 1).to_timestamp("M")


def safe_write_ts(dss_out, pathname: str, ts_obj):
    """Write a TimeSeriesContainer, tolerating API differences."""
    ts_obj.pathname = pathname
    if hasattr(dss_out, "put_ts"):
        dss_out.put_ts(ts_obj)
    elif hasattr(dss_out, "write_ts"):
        dss_out.write_ts(ts_obj)


def path_key(pathname: str) -> tuple:
    """(Part B, Part C) upper-cased from a DSS pathname."""
    parts = pathname.strip("/").split("/")
    return (parts[1].upper(), parts[2].upper())


def is_data_csv(filepath: str) -> bool:
    """Return True if *filepath* looks like a data CSV (not a report/diagnostic)."""
    name = os.path.basename(filepath).lower()
    if "parts_not_found" in name or "replaced_months_report" in name:
        return False
    return name.endswith(".csv")


def collect_data_csvs(src_dir: Path) -> list:
    """Return sorted list of data-CSV paths (excluding reports) in *src_dir*."""
    if not src_dir.is_dir():
        return []
    csvs = sorted(src_dir.glob("*.csv"))
    return [c for c in csvs if is_data_csv(str(c))]


import functools

@functools.lru_cache(maxsize=1)
def read_master_inventory() -> pd.DataFrame:
    """Read the MASTER sheet from the inventory workbook.

    Returns DataFrame with columns:
        Part_B, Part_C, Input_Category, Missing, Constant_Rept, Used_in_DCR
    Boolean flag columns are True/False (blanks treated as False).
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(str(INVENTORY_XLSX), read_only=True, data_only=True)
    except PermissionError:
        sys.exit(
            f"ERROR: Cannot read the master inventory — the file may be open in Excel.\n"
            f"  Close the file and re-run the script:\n"
            f"  {INVENTORY_XLSX}"
        )
    ws = wb["MASTER"]
    rows = list(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True))
    wb.close()
    records = []
    for r in rows:
        b, c, cat = r[2], r[3], r[8]   # columns C, D, I (0-indexed: 2, 3, 8)
        if b and c and cat:
            records.append({
                "Part_B":          str(b).upper().replace(" ", "_"),
                "Part_C":          str(c).upper().replace(" ", "_"),
                "Input_Category":  str(cat).strip(),
                "Missing":         str(r[10]).strip().upper() == "T",
                "Constant_Rept":   str(r[11]).strip().upper() == "T",
                "Used_in_DCR":     str(r[12]).strip().upper() != "F",  # T or blank → True
            })
    return pd.DataFrame(records)


def _extract_wy_pattern(ser_clean):
    """Extract a 12-month repeating pattern from non-missing baseline data.

    Strategy (in order of preference):
      1. Find the last **complete** water year (Oct–Sep) and use its 12 values.
      2. If no complete WY exists, build the pattern from the last occurrence
         of each calendar month across the entire series.

    Returns
    -------
    pattern : dict[int, float]
        Mapping  {month_number: value}  (1-12).  May have fewer than 12 keys
        if some months are entirely absent from the series.
    complete : bool
        True if all 12 months are represented.
    """
    WY_MONTHS = [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]

    # Assign each timestamp a water year (Oct = start of next WY)
    wy = ser_clean.index.year + (ser_clean.index.month >= 10).astype(int)

    # Try to find the last complete WY (has all 12 months)
    for water_year in sorted(wy.unique(), reverse=True):
        wy_slice = ser_clean[wy == water_year]
        months_present = set(wy_slice.index.month)
        if months_present == set(WY_MONTHS):
            pattern = {}
            for dt, val in zip(wy_slice.index, wy_slice.values):
                pattern[dt.month] = val
            return pattern, True

    # Fallback: last occurrence of each month across the entire series
    pattern = {}
    for m in WY_MONTHS:
        month_vals = ser_clean[ser_clean.index.month == m]
        if not month_vals.empty:
            pattern[m] = month_vals.iloc[-1]

    return pattern, len(pattern) == 12


def build_constant_rept_overrides(
    dss_in, baseline_bucket, part_key, overwrite_start, overwrite_end
):
    """For a Constant/Rept SV, extract a repeating annual pattern from the
    baseline and tile it across the entire overwrite window.

    The pattern is sourced from the last complete water year (Oct–Sep) in the
    baseline.  If no complete WY exists, falls back to the last available value
    for each calendar month.

    IMPORTANT: All D-part blocks for a given (Part B, Part C) are merged into
    a single series *before* extracting the pattern, so that one consistent
    12-month pattern is applied across the entire overwrite window.

    Returns a dict  {pathname: np.ndarray}  of modified full-length value arrays,
    or an empty dict if the key is not in the baseline.
    """
    if part_key not in baseline_bucket:
        return {}

    # ── Step 1: Read all D-part blocks and merge into one unified series ──
    block_data = []  # list of (pathname, eom_pd, ts_vals)
    merged = {}      # {datetime: value} across all blocks
    for pathname in baseline_bucket[part_key]:
        ts = dss_in.read_ts(pathname, trim_missing=False)
        eom_idx = dss_eom(ts.pytimes)
        eom_pd  = pd.DatetimeIndex(eom_idx)
        ts_vals = np.array(ts.values, dtype=float).copy()
        block_data.append((pathname, eom_pd, ts_vals))
        for i, dt in enumerate(eom_pd):
            if ts_vals[i] > -900:
                merged[dt] = ts_vals[i]

    if not merged:
        return {pn: tv for pn, _, tv in block_data}

    # ── Step 2: Extract ONE pattern from the merged series ───────────────
    merged_dates = sorted(merged.keys())
    merged_ser = pd.Series(
        [merged[d] for d in merged_dates],
        index=pd.DatetimeIndex(merged_dates),
    )
    pattern, complete = _extract_wy_pattern(merged_ser)

    if not complete:
        missing_months = sorted(set(range(1, 13)) - set(pattern.keys()))
        month_names = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
                       7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
        missing_str = ", ".join(month_names[m] for m in missing_months)
        print(f"  WARNING: Constant/Rept pattern for {part_key} is missing "
              f"months: {missing_str}")

    # ── Step 3: Apply the single pattern to every block ──────────────────
    result = {}
    for pathname, eom_pd, ts_vals in block_data:
        mask = (eom_pd >= overwrite_start) & (eom_pd <= overwrite_end)
        for i in np.where(mask)[0]:
            m = eom_pd[i].month
            if m in pattern:
                ts_vals[i] = pattern[m]
        result[pathname] = ts_vals

    return result


def generate_per_variable_plots(stats_csv: Path, cached_series: dict = None):
    """Generate diagnostic figures for every modified (Part B, Part C).

    If *cached_series* is provided (a dict from ``compute_modification_statistics``),
    DSS files are **not** re-read — a major speed-up.

    Three figure sets are produced per variable:

    1. **Monthly** — ``figures/all-terms/<Cat>/<B>__<C>_r2=<NN>.png``
       Left: monthly time-series, Right: monthly scatter with R²/NSE.

    2. **Annual** — ``figures/all-terms-annual/<Cat>/<B>__<C>_r2=<NN>.png``
       Left: water-year annual time-series, Right: annual scatter with R²/NSE.

    3. **Monthly Average** — ``figures/all-terms-mavg/<Cat>/<B>__<C>.png``
       Left: 12-month climatology bar chart, Right: climatology scatter.

    Terms are separated into three output tiers:

    - **all-terms** / **all-terms-annual** / **all-terms-mavg** — regular
      reconstructed terms.
    - **all-terms-constant-rept** / **-annual** / **-mavg** — terms whose
      modified values are a 12-month repeating pattern (inventory
      ``Constant/Rept = T``).
    - **all-terms-constant** / **-annual** / **-mavg** — terms whose modified
      values are a single constant (zero or otherwise).
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MaxNLocator

    _MEAN_UNITS = {"CFS", "MG/L", "FAHRENHE", "KPA", "NONE", "DAYS"}

    FIG_W     = 6.5      # inches — maximum width
    FIG_H     = 2.5      # inches — height (controls scatter squareness)
    FONT_SIZE = 7

    plt.rcParams.update({
        "font.size":        FONT_SIZE,
        "axes.titlesize":   FONT_SIZE,
        "axes.labelsize":   FONT_SIZE,
        "xtick.labelsize":  FONT_SIZE,
        "ytick.labelsize":  FONT_SIZE,
        "legend.fontsize":  FONT_SIZE,
        "figure.dpi":       300,
        "savefig.dpi":      300,
        "figure.facecolor": "white",
    })

    _MONTH_LABELS = ["Oct", "Nov", "Dec", "Jan", "Feb", "Mar",
                     "Apr", "May", "Jun", "Jul", "Aug", "Sep"]
    _MONTH_NUMS   = [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]

    # ── Output directories ───────────────────────────────────────────────
    # Regular reconstructed terms
    dir_monthly = stats_csv.parent / "figures" / "all-terms"
    dir_annual  = stats_csv.parent / "figures" / "all-terms-annual"
    dir_mavg    = stats_csv.parent / "figures" / "all-terms-mavg"
    # 12-month repeating pattern (Constant/Rept = T)
    dir_monthly_rept = stats_csv.parent / "figures" / "all-terms-constant-rept"
    dir_annual_rept  = stats_csv.parent / "figures" / "all-terms-annual-constant-rept"
    dir_mavg_rept    = stats_csv.parent / "figures" / "all-terms-mavg-constant-rept"
    # Single constant value (zero or otherwise)
    dir_monthly_const = stats_csv.parent / "figures" / "all-terms-constant"
    dir_annual_const  = stats_csv.parent / "figures" / "all-terms-annual-constant"
    dir_mavg_const    = stats_csv.parent / "figures" / "all-terms-mavg-constant"
    for d in (dir_monthly, dir_annual, dir_mavg,
              dir_monthly_rept, dir_annual_rept, dir_mavg_rept,
              dir_monthly_const, dir_annual_const, dir_mavg_const):
        d.mkdir(parents=True, exist_ok=True)

    # ── R² / NSE / Category lookup from the stats CSV ────────────────────
    stats_df = pd.read_csv(stats_csv)

    inv = read_master_inventory()
    cat_map = {}
    rept_map = {}   # (Part_B, Part_C) → bool  (Constant/Rept flag)
    for _, r in inv.iterrows():
        cat_map[(r["Part_B"], r["Part_C"])] = r["Input_Category"]
        rept_map[(r["Part_B"], r["Part_C"])] = bool(r.get("Constant_Rept", False))

    # Build Input_Category → numbered folder name lookup from MODULE_CONFIG
    _cat_to_folder = {}
    for label, (_rel, inv_cat) in MODULE_CONFIG.items():
        _cat_to_folder[inv_cat] = label          # e.g. "CalSimHydro" → "_05_CalSimHydro"
    # Categories without dedicated validation modules — map to their directory numbers
    _NO_VALIDATION_FOLDER = {
        "Closure Terms":      "_07_ClosureTerms",
        "Day-Volume Fraction": "_09_DayVolumeFraction",
        "Salinity":           "_12_Salinity",
    }
    for cat_name in CATEGORIES_WITHOUT_VALIDATION:
        if cat_name not in _cat_to_folder:
            _cat_to_folder[cat_name] = _NO_VALIDATION_FOLDER.get(
                cat_name, f"_XX_{cat_name.replace(' ', '')}"
            )

    stats_lookup = {}
    for _, row in stats_df.iterrows():
        k = (row["Part_B"], row["Part_C"])
        # Skip all-zero paths — nothing useful to plot
        if row.get("AllZero", False):
            continue
        stats_lookup[k] = {
            "R2":       row.get("R2-Monthly", np.nan),
            "NSE":      row.get("NSE-Monthly", np.nan),
            "R2_Ann":   row.get("R2-Ann", np.nan),
            "NSE_Ann":  row.get("NSE-Ann", np.nan),
            "Units":    row.get("Units", ""),
            "Category": cat_map.get(k, "Unknown"),
        }

    # ── Catalog baseline DSS ─────────────────────────────────────────────
    with HecDss.Open(str(BASELINE_DSS), version=6, catalog_flag=True) as _dss:
        _bp = _dss.getPathnameList(DSS_PATTERN)
    baseline_bucket = {}
    for p in _bp:
        k = path_key(p)
        baseline_bucket.setdefault(k, []).append(p)

    # ── Helper: R²/NSE computation ───────────────────────────────────────
    import warnings

    def _r2(obs, sim):
        if len(obs) < 2:
            return np.nan
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", RuntimeWarning)
            c = np.corrcoef(obs, sim)[0, 1]
        return c ** 2 if np.isfinite(c) else np.nan

    def _nse(obs, sim):
        if len(obs) < 2:
            return np.nan
        ss_res = np.sum((sim - obs) ** 2)
        ss_tot = np.sum((obs - np.mean(obs)) ** 2)
        return 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan

    # ── Helper: plot time-series + scatter panel ─────────────────────────
    def _plot_ts_scatter(b_x, b_y, m_x, m_y, title, units, r2_val, nse_val,
                         out_path, x_is_date=True):
        """Shared plotting logic for monthly / annual figures."""
        fig = plt.figure(figsize=(FIG_W, FIG_H))
        gs  = fig.add_gridspec(1, 2,
                                width_ratios=[FIG_W - FIG_H, FIG_H],
                                wspace=0.25)
        ax_ts = fig.add_subplot(gs[0])
        ax_sc = fig.add_subplot(gs[1])

        # Left: time series
        ax_ts.plot(b_x, b_y, color="#1f77b4", lw=0.5, label="Baseline", alpha=0.8)
        ax_ts.plot(m_x, m_y, color="#d62728", lw=0.5, label="Product A", alpha=0.8)
        ax_ts.set_ylabel(units if units else "Value")
        ax_ts.set_title(title)
        ax_ts.legend(loc="best", framealpha=0.7)
        if x_is_date:
            for lbl in ax_ts.get_xticklabels():
                lbl.set_rotation(30)
                lbl.set_ha("right")

        # Right: scatter
        all_v  = np.concatenate([np.asarray(b_y), np.asarray(m_y)])
        vmin   = np.nanmin(all_v)
        vmax_v = np.nanmax(all_v)
        margin = (vmax_v - vmin) * 0.05 if vmax_v != vmin else 1.0
        lims   = [vmin - margin, vmax_v + margin]

        ax_sc.scatter(b_y, m_y, s=6, alpha=0.45, color="#1f77b4", edgecolors="none")
        ax_sc.plot(lims, lims, "k--", lw=0.8, alpha=0.5)
        ax_sc.set_xlim(lims)
        ax_sc.set_ylim(lims)
        ax_sc.set_box_aspect(1)

        ax_sc.xaxis.set_major_locator(MaxNLocator(nbins=5))
        fig.canvas.draw()
        xticks = ax_sc.get_xticks()
        ax_sc.set_xticks(xticks)
        ax_sc.set_yticks(xticks)
        ax_sc.set_xlim(lims)
        ax_sc.set_ylim(lims)

        ax_sc.set_xlabel(f"Baseline ({units})" if units else "Baseline")
        ax_sc.set_ylabel(f"Product A ({units})" if units else "Product A")

        r2_str  = f"R\u00b2 = {r2_val:.3f}"  if np.isfinite(r2_val)  else "R\u00b2 = n/a"
        nse_str = f"NSE = {nse_val:.3f}" if np.isfinite(nse_val) else "NSE = n/a"
        ax_sc.text(0.05, 0.95, f"{r2_str}\n{nse_str}",
                   transform=ax_sc.transAxes, va="top", ha="left",
                   fontsize=FONT_SIZE,
                   bbox=dict(boxstyle="round,pad=0.3",
                             facecolor="white", alpha=0.85,
                             edgecolor="gray", linewidth=0.5))

        fig.savefig(out_path, bbox_inches="tight")
        plt.close(fig)

    # ── Helper: monthly-average bar + scatter ────────────────────────────
    def _plot_mavg(b_mavg, m_mavg, title, units, out_path):
        """12-month climatology bar chart + scatter."""
        fig = plt.figure(figsize=(FIG_W, FIG_H))
        gs  = fig.add_gridspec(1, 2,
                                width_ratios=[FIG_W - FIG_H, FIG_H],
                                wspace=0.25)
        ax_bar = fig.add_subplot(gs[0])
        ax_sc  = fig.add_subplot(gs[1])

        x_pos = np.arange(12)
        w = 0.35
        ax_bar.bar(x_pos - w/2, b_mavg, w, color="#1f77b4", alpha=0.8, label="Baseline")
        ax_bar.bar(x_pos + w/2, m_mavg, w, color="#d62728", alpha=0.8, label="Product A")
        ax_bar.set_xticks(x_pos)
        ax_bar.set_xticklabels(_MONTH_LABELS)
        ax_bar.set_ylabel(units if units else "Value")
        ax_bar.set_title(title)
        ax_bar.legend(loc="best", framealpha=0.7)

        # Scatter
        all_v  = np.concatenate([b_mavg, m_mavg])
        vmin   = np.nanmin(all_v)
        vmax_v = np.nanmax(all_v)
        margin = (vmax_v - vmin) * 0.05 if vmax_v != vmin else 1.0
        lims   = [vmin - margin, vmax_v + margin]

        ax_sc.scatter(b_mavg, m_mavg, s=12, alpha=0.7, color="#1f77b4", edgecolors="none")
        ax_sc.plot(lims, lims, "k--", lw=0.8, alpha=0.5)
        ax_sc.set_xlim(lims)
        ax_sc.set_ylim(lims)
        ax_sc.set_box_aspect(1)

        from matplotlib.ticker import MaxNLocator as _MNL
        ax_sc.xaxis.set_major_locator(_MNL(nbins=5))
        fig.canvas.draw()
        xticks = ax_sc.get_xticks()
        ax_sc.set_xticks(xticks)
        ax_sc.set_yticks(xticks)
        ax_sc.set_xlim(lims)
        ax_sc.set_ylim(lims)

        ax_sc.set_xlabel(f"Baseline ({units})" if units else "Baseline")
        ax_sc.set_ylabel(f"Product A ({units})" if units else "Product A")

        r2_val  = _r2(b_mavg, m_mavg)
        nse_val = _nse(b_mavg, m_mavg)
        r2_str  = f"R\u00b2 = {r2_val:.3f}"  if np.isfinite(r2_val)  else "R\u00b2 = n/a"
        nse_str = f"NSE = {nse_val:.3f}" if np.isfinite(nse_val) else "NSE = n/a"
        ax_sc.text(0.05, 0.95, f"{r2_str}\n{nse_str}",
                   transform=ax_sc.transAxes, va="top", ha="left",
                   fontsize=FONT_SIZE,
                   bbox=dict(boxstyle="round,pad=0.3",
                             facecolor="white", alpha=0.85,
                             edgecolor="gray", linewidth=0.5))

        fig.savefig(out_path, bbox_inches="tight")
        plt.close(fig)

    # ── Helper: classify a term ──────────────────────────────────────────
    def _classify_term(pk, m_ser):
        """Return 'constant-rept', 'constant', or 'regular'.

        - 'constant-rept' — inventory Constant/Rept = T
        - 'constant'      — modified series is a single constant value
                            (std ≈ 0, including all-zero)
        - 'regular'       — everything else
        """
        if rept_map.get(pk, False):
            # Check if it is *also* a flat constant (all same value)
            if m_ser.std() < 1e-10:
                return "constant"        # truly constant, even if flagged rept
            return "constant-rept"
        # Not flagged as Constant/Rept — check for flat constant
        if m_ser.std() < 1e-10:
            return "constant"
        return "regular"

    # Map classification → (monthly_dir, annual_dir, mavg_dir)
    _CLASS_DIRS = {
        "regular":       (dir_monthly,       dir_annual,       dir_mavg),
        "constant-rept": (dir_monthly_rept,  dir_annual_rept,  dir_mavg_rept),
        "constant":      (dir_monthly_const, dir_annual_const, dir_mavg_const),
    }

    # ── Generate figures per variable ────────────────────────────────────
    n_plotted = 0
    n_by_class = {"regular": 0, "constant-rept": 0, "constant": 0}
    all_keys  = sorted(stats_lookup.keys())

    # If cached_series is available, use it directly (no DSS reads needed).
    # Otherwise, fall back to reading DSS files.
    _use_cache = cached_series is not None and len(cached_series) > 0

    def _iter_series():
        """Yield (pk, b_ser, m_ser, units) for each variable."""
        if _use_cache:
            for pk in all_keys:
                entry = cached_series.get(pk)
                if entry is None:
                    continue
                yield pk, entry["b_ser"], entry["m_ser"], entry["units"]
        else:
            # Fallback: open DSS and merge blocks (slow path)
            with HecDss.Open(str(BASELINE_DSS), version=6, catalog_flag=True) as dss_base, \
                 HecDss.Open(_dss_str(OUTPUT_DSS),   version=6, catalog_flag=True) as dss_mod:
                for pk in all_keys:
                    if pk not in baseline_bucket:
                        continue
                    base_dict, mod_dict = {}, {}
                    for pathname in baseline_bucket[pk]:
                        try:
                            ts_b = dss_base.read_ts(pathname, trim_missing=False)
                            ts_m = dss_mod.read_ts(pathname, trim_missing=False)
                        except Exception:
                            continue
                        eom = dss_eom(ts_b.pytimes)
                        bv  = np.array(ts_b.values, dtype=float)
                        mv  = np.array(ts_m.values, dtype=float)
                        for i, dt in enumerate(eom):
                            if bv[i] > -900:
                                base_dict[dt] = bv[i]
                            if mv[i] > -900:
                                mod_dict[dt] = mv[i]
                    if not base_dict or not mod_dict:
                        continue
                    common = sorted(set(base_dict) & set(mod_dict))
                    common = [d for d in common
                              if OVERWRITE_START <= d <= OVERWRITE_END]
                    if not common:
                        continue
                    idx   = pd.DatetimeIndex(common)
                    b_ser = pd.Series([base_dict[d] for d in common], index=idx)
                    m_ser = pd.Series([mod_dict[d]  for d in common], index=idx)
                    units = stats_lookup.get(pk, {}).get("Units", "")
                    yield pk, b_ser, m_ser, units

    if _use_cache:
        print("    (using cached series — DSS reads skipped)")

    for pk, b_ser, m_ser, units in _iter_series():
        info     = stats_lookup.get(pk, {})
        r2_val   = info.get("R2",  np.nan)
        nse_val  = info.get("NSE", np.nan)
        r2_ann   = info.get("R2_Ann",  np.nan)
        nse_ann  = info.get("NSE_Ann", np.nan)
        category = info.get("Category", "Unknown")

        # Classify:  regular | constant-rept | constant
        classification = _classify_term(pk, m_ser)
        _dir_m, _dir_a, _dir_v = _CLASS_DIRS[classification]
        n_by_class[classification] += 1

        # Category subfolder name
        cat_folder = _cat_to_folder.get(category, f"_XX_{category}")
        cat_safe = cat_folder.replace("/", "_").replace("\\", "_").strip()

        # ── 1) Monthly time-series + scatter ─────────────────────────
        cat_dir_m = _dir_m / cat_safe
        cat_dir_m.mkdir(parents=True, exist_ok=True)
        fname_m = f"{pk[0]}__{pk[1]}.png".replace("/", "_")
        _plot_ts_scatter(
            b_ser.index, b_ser.values, m_ser.index, m_ser.values,
            f"{pk[0]} / {pk[1]}", units, r2_val, nse_val,
            cat_dir_m / fname_m, x_is_date=True,
        )

        # ── 2) Annual time-series + scatter ──────────────────────────
        idx = b_ser.index
        wy_labels = np.where(idx.month >= 10, idx.year + 1, idx.year)
        df_wy = pd.DataFrame({"base": b_ser.values, "mod": m_ser.values},
                             index=wy_labels)
        wy_counts = df_wy.groupby(level=0).size()
        complete  = wy_counts[wy_counts == 12].index
        df_wy     = df_wy.loc[df_wy.index.isin(complete)]

        if not df_wy.empty:
            use_mean = any(u in units.upper() for u in _MEAN_UNITS)
            agg = "mean" if use_mean else "sum"
            ann = df_wy.groupby(level=0).agg(agg)

            cat_dir_a = _dir_a / cat_safe
            cat_dir_a.mkdir(parents=True, exist_ok=True)
            fname_a = f"{pk[0]}__{pk[1]}.png".replace("/", "_")
            _plot_ts_scatter(
                ann.index.values, ann["base"].values,
                ann.index.values, ann["mod"].values,
                f"{pk[0]} / {pk[1]}  (WY Annual)", units, r2_ann, nse_ann,
                cat_dir_a / fname_a, x_is_date=False,
            )

        # ── 3) Monthly average climatology ───────────────────────────
        b_mavg = np.array([b_ser[b_ser.index.month == m].mean()
                           for m in _MONTH_NUMS])
        m_mavg = np.array([m_ser[m_ser.index.month == m].mean()
                           for m in _MONTH_NUMS])

        if np.any(np.isfinite(b_mavg)):
            cat_dir_v = _dir_v / cat_safe
            cat_dir_v.mkdir(parents=True, exist_ok=True)
            fname_v = f"{pk[0]}__{pk[1]}.png".replace("/", "_")
            _plot_mavg(b_mavg, m_mavg,
                       f"{pk[0]} / {pk[1]}  (Monthly Avg)", units,
                       cat_dir_v / fname_v)

        n_plotted += 1
        if n_plotted % 100 == 0:
            print(f"    ... {n_plotted} variables plotted")

    print(f"  Per-variable plots:  {n_plotted} variables")
    print(f"    regular:       {n_by_class['regular']:>4d}  → all-terms/")
    print(f"    constant-rept: {n_by_class['constant-rept']:>4d}  → all-terms-constant-rept/")
    print(f"    constant:      {n_by_class['constant']:>4d}  → all-terms-constant/")
    return n_plotted


def generate_stats_report(stats_csv: Path, cached_series: dict = None):
    """Read ``modification_statistics.csv``, join with inventory categories,
    and produce a summary text report plus diagnostic figures.

    Outputs (written to the same directory as *stats_csv*):
        ``modification_statistics_report.txt``
        ``figures/r2_nse_by_category.png``
        ``figures/pctchange_monthly_by_category.png``
        ``figures/abschange_annual_by_category.png``
        ``figures/r2_nse_scatter.png``
        ``figures/all-terms/<PartB>__<PartC>.png``  (one per variable)
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MaxNLocator

    _FS = 7
    plt.rcParams.update({
        "font.size":        _FS,
        "axes.titlesize":   _FS,
        "axes.labelsize":   _FS,
        "xtick.labelsize":  _FS,
        "ytick.labelsize":  _FS,
        "legend.fontsize":  _FS,
        "figure.dpi":       300,
        "savefig.dpi":      300,
        "figure.facecolor": "white",
    })

    out_dir = stats_csv.parent
    fig_dir = out_dir / "figures"
    fig_dir.mkdir(exist_ok=True)

    # ── Load stats & join inventory category ─────────────────────────────
    df = pd.read_csv(stats_csv)
    inv = read_master_inventory()

    # Build (Part_B, Part_C) → Input_Category lookup  &  Constant_Rept flag
    cat_map = {}
    rept_map = {}
    for _, r in inv.iterrows():
        cat_map[(r["Part_B"], r["Part_C"])]  = r["Input_Category"]
        rept_map[(r["Part_B"], r["Part_C"])] = r.get("Constant_Rept", False)

    df["Input_Category"] = df.apply(
        lambda r: cat_map.get((r["Part_B"], r["Part_C"]), "Unknown"), axis=1
    )
    df["Constant_Rept"] = df.apply(
        lambda r: rept_map.get((r["Part_B"], r["Part_C"]), False), axis=1
    )

    # Exclude paths flagged as all-zero / trivial (both baseline & modified ≈ 0)
    n_allzero = 0
    if "AllZero" in df.columns:
        n_allzero = df["AllZero"].sum()
        df = df[~df["AllZero"]].copy()

    # Drop categories where every path has zero change (e.g. Constant/Rept-only)
    # A category is "unmodified" if all its AbsChange monthly averages are zero or NaN
    _abs_cols = [f"AbsChange-{m}-Avg" for m in
                 ["Oct","Nov","Dec","Jan","Feb","Mar",
                  "Apr","May","Jun","Jul","Aug","Sep"]
                 if f"AbsChange-{m}-Avg" in df.columns]
    if _abs_cols:
        _grp = df.groupby("Input_Category")[_abs_cols].apply(
            lambda g: g.abs().sum().sum()
        )
        _active_cats = _grp[_grp > 0].index.tolist()
        df = df[df["Input_Category"].isin(_active_cats)].copy()

    # Category order by module number (all modules including non-validation ones)
    # Full ordered list: MODULE_CONFIG categories + non-validation categories
    # interleaved at their correct directory number positions.
    _ALL_CATS_ORDERED = [
        # 05          06              07                08
        "CalSimHydro", "CalSimHydroEE", "Closure Terms", "Reservoir Evaporation",
        # 09                   10            11
        "Day-Volume Fraction", "Rim Inflow", "Delta Channel Depletion",
        # 12         13                  14
        "Salinity", "Small Watersheds", "Reservoir Storage Curves",
        # 15              16                        17
        "Instream Flows", "Tulare Groundwater Terms", "Climate",
        # 18      19
        "Other", "Upper Watershed Modules",
    ]
    _present = set(df["Input_Category"].unique())
    cat_order = [c for c in _ALL_CATS_ORDERED if c in _present]
    cat_order += sorted(_present - set(cat_order))
    n_total = len(df)

    # Subset excluding Constant/Rept terms — used for summary figures only
    df_plots = df[~df["Constant_Rept"]].copy()
    n_rept = n_total - len(df_plots)
    # Keep only categories that still have rows after filtering
    _present_plots = set(df_plots["Input_Category"].unique())
    cat_order_plots = [c for c in cat_order if c in _present_plots]

    _MONTHS = ["Oct", "Nov", "Dec", "Jan", "Feb", "Mar",
               "Apr", "May", "Jun", "Jul", "Aug", "Sep"]

    # ══════════════════════════════════════════════════════════════════════
    # 1.  Text report
    # ══════════════════════════════════════════════════════════════════════
    lines = []
    lines.append("=" * 72)
    lines.append("  Modification Statistics — Summary Report")
    lines.append("=" * 72)
    lines.append("")
    lines.append(f"  Total paths analysed:  {n_total}")
    if n_allzero:
        lines.append(f"  Excluded (all-zero):   {n_allzero}")
    if n_rept:
        lines.append(f"  Constant/Rept terms:   {n_rept}  (excluded from summary figures)")
    lines.append(f"  Input categories:      {len(cat_order)}")
    lines.append("")

    for cat in cat_order:
        sub = df[df["Input_Category"] == cat]
        n = len(sub)
        lines.append("─" * 72)
        lines.append(f"  {cat}  ({n} path{'s' if n != 1 else ''})")
        lines.append("─" * 72)

        # Overall quality metrics
        r2m  = sub["R2-Monthly"].dropna()
        nsem = sub["NSE-Monthly"].dropna()
        r2a  = sub["R2-Ann"].dropna()
        nsea = sub["NSE-Ann"].dropna()

        lines.append("")
        lines.append("  Goodness-of-fit                   Median     Mean      Min       Max")
        for label, s in [("R² (monthly)", r2m), ("NSE (monthly)", nsem),
                         ("R² (annual)", r2a), ("NSE (annual)", nsea)]:
            if s.empty:
                lines.append(f"    {label:28s}     —         —         —         —")
            else:
                lines.append(f"    {label:28s} {s.median():9.4f} {s.mean():9.4f} "
                             f"{s.min():9.4f} {s.max():9.4f}")

        # Monthly median percent change (across paths in this category)
        lines.append("")
        lines.append("  Median % change by month (across paths):")
        line_vals = "    "
        for m in _MONTHS:
            col = f"PctChange-{m}-Avg"
            if col in sub.columns:
                v = sub[col].median()
                line_vals += f"{m}: {v:+7.1f}%  "
            else:
                line_vals += f"{m}:     —  "
        lines.append(line_vals)

        # Annual summary
        if "AbsChange-Ann-Avg" in sub.columns:
            ann_chg = sub["AbsChange-Ann-Avg"].dropna()
            ann_pct = sub["PctChange-Ann-Avg"].dropna()
            lines.append("")
            lines.append(f"  Annual change (WY):  "
                         f"abs median = {ann_chg.median():+.2f}  "
                         f"| pct median = {ann_pct.median():+.1f}%  "
                         f"| range [{ann_pct.min():+.1f}%, {ann_pct.max():+.1f}%]")

        # Worst paths (lowest NSE-Monthly)
        if not nsem.empty:
            worst = sub.nsmallest(min(5, n), "NSE-Monthly")[
                ["Part_B", "Part_C", "R2-Monthly", "NSE-Monthly"]
            ]
            lines.append("")
            lines.append("  Lowest NSE-Monthly paths:")
            for _, w in worst.iterrows():
                lines.append(f"    {w['Part_B']:30s} {w['Part_C']:20s}  "
                             f"R²={w['R2-Monthly']:.4f}  NSE={w['NSE-Monthly']:.4f}")

        # Paths excluded from box plots (NaN R2-Monthly → constant, zero, or missing)
        nan_r2 = sub[sub["R2-Monthly"].isna()][["Part_B", "Part_C"]]
        if not nan_r2.empty:
            lines.append("")
            lines.append(f"  Excluded from box plots (NaN R²-Monthly — constant/zero/missing):")
            for _, w in nan_r2.iterrows():
                lines.append(f"    {w['Part_B']:30s} {w['Part_C']:20s}")

        lines.append("")

    # Write text report
    rpt_path = out_dir / "modification_statistics_report.txt"
    rpt_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Report:  {rpt_path.name}")

    # ══════════════════════════════════════════════════════════════════════
    # 2.  Figures
    # ══════════════════════════════════════════════════════════════════════
    plt.rcParams.update({"figure.dpi": 300, "font.size": 7,
                         "axes.titlesize": 7, "axes.labelsize": 7,
                         "xtick.labelsize": 7, "ytick.labelsize": 7,
                         "legend.fontsize": 7, "figure.facecolor": "white"})

    n_cats = len(cat_order_plots)

    # ── Figs 1-4: Individual R²/NSE box plots by category ───────────────
    #    (Constant/Rept terms excluded; n= reflects non-NaN count per metric)
    for col, title, fname in [
        ("R2-Monthly",  "R² (Monthly)",  "r2_monthly_by_category.png"),
        ("NSE-Monthly", "NSE (Monthly)", "nse_monthly_by_category.png"),
        ("R2-Ann",      "R² (Annual)",   "r2_annual_by_category.png"),
        ("NSE-Ann",     "NSE (Annual)",  "nse_annual_by_category.png"),
    ]:
        fig, ax = plt.subplots(figsize=(6.5, 4.0))
        # Filter to rows where R2-Monthly is non-NaN, then extract the metric values
        box_data = [df_plots.loc[(df_plots["Input_Category"] == c) & df_plots["R2-Monthly"].notna(), col].values
                    for c in cat_order_plots]
        # n= reflects non-NaN R2-Monthly count (consistent across all metric plots)
        r2m_counts = {c: len(df_plots.loc[(df_plots["Input_Category"] == c) & df_plots["R2-Monthly"].notna()])
                      for c in cat_order_plots}
        box_labels = [f"{c} (n={r2m_counts[c]})" for c in cat_order_plots]
        bp = ax.boxplot(box_data, vert=True, patch_artist=True, widths=0.6)
        for patch in bp["boxes"]:
            patch.set_facecolor("#5B9BD5")
            patch.set_alpha(0.7)
        ax.set_xticklabels(box_labels, rotation=45, ha="right")
        ax.set_ylabel(title)
        ax.set_title(f"{title} by Input Category  (excl. Constant/Rept)")
        ax.axhline(0.0, color="red", ls="--", lw=0.6, alpha=0.5)

        if col.startswith("NSE"):
            _NSE_YMIN = -1.05
            ax.set_ylim(_NSE_YMIN, 1.05)
            ax.axhline(1.0, color="red", ls="--", lw=0.6, alpha=0.5)
            for idx_b, bd in enumerate(box_data, start=1):
                if len(bd) > 0 and np.nanmin(bd) < _NSE_YMIN:
                    ax.text(idx_b, _NSE_YMIN + 0.05, "*", ha="center",
                            va="bottom", fontsize=_FS, fontweight="bold",
                            color="red", clip_on=False)
        else:
            ax.axhline(1.0, color="green", ls="--", lw=0.6, alpha=0.5)

        fig.savefig(fig_dir / fname, bbox_inches="tight")
        plt.close(fig)
        print(f"  Figure:  figures/{fname}")

    # ── Fig 5: Monthly % change heatmap by category ─────────────────────
    #    (Constant/Rept terms excluded)
    pct_cols = [f"PctChange-{m}-Avg" for m in _MONTHS]
    existing_cols = [c for c in pct_cols if c in df_plots.columns]
    if existing_cols:
        heat = df_plots.groupby("Input_Category")[existing_cols].median()
        heat = heat.reindex(cat_order_plots)
        heat.columns = [c.split("-")[1] for c in heat.columns]

        fig_h = max(2.5, 0.4 * n_cats)
        fig, ax = plt.subplots(figsize=(6.5, fig_h))
        im = ax.imshow(heat.values, aspect="auto", cmap="RdBu",
                        vmin=-25, vmax=25)
        ax.set_xticks(range(len(heat.columns)))
        ax.set_xticklabels(heat.columns)
        ax.set_yticks(range(len(heat.index)))
        # Build labels with per-category non-NaN counts for % change
        _heat_counts = df_plots.groupby("Input_Category")[existing_cols].apply(
            lambda g: g.dropna(how="all").shape[0]
        )
        heat_labels = [f"{c} (n={_heat_counts.get(c, 0)})" for c in cat_order_plots]
        ax.set_yticklabels(heat_labels)
        # Annotate cells
        for i in range(heat.shape[0]):
            for j in range(heat.shape[1]):
                v = heat.iloc[i, j]
                if np.isfinite(v):
                    txt = f"{v:+.0f}" if abs(v) >= 1 else f"{v:+.1f}"
                    ax.text(j, i, txt, ha="center", va="center",
                            fontsize=5, color="black" if abs(v) < 60 else "white")
        fig.colorbar(im, ax=ax, label="Median % Change")
        ax.set_title("Median Monthly % Change (Mod vs Base) by Category  (excl. Constant/Rept)")
        fig.tight_layout()
        fig.savefig(fig_dir / "pctchange_monthly_by_category.png", bbox_inches="tight")
        plt.close(fig)
        print(f"  Figure:  figures/pctchange_monthly_by_category.png")

    # ── Figs 6+: Scatter — |Annual Avg Baseline| vs Avg Annual % Change ─
    #    One plot per category + one "All" plot.  Constant/Rept excluded.
    _scatter_dir = fig_dir / "annual_scatter"
    _scatter_dir.mkdir(exist_ok=True)

    _has_cols = ("Base-Ann-Avg" in df_plots.columns and
                 "Mod-Ann-Avg" in df_plots.columns)
    if _has_cols:

        def _scatter_base_vs_mod(sub, title, out_path):
            """Scatter: Mean Annual Baseline (x) vs Mean Annual Product A (y).
            Linear axes, 1:1 line plus ±10% reference lines."""
            x = sub["Base-Ann-Avg"].values.astype(float)
            y = sub["Mod-Ann-Avg"].values.astype(float)
            mask = np.isfinite(x) & np.isfinite(y)
            x, y = x[mask], y[mask]
            if len(x) == 0:
                return
            fig, ax = plt.subplots(figsize=(6.5, 4.0))

            ax.scatter(x, y, s=8, alpha=0.55,
                       color="#1f77b4", edgecolors="none", zorder=2)

            # ── Reference lines ──────────────────────────────────────
            all_v = np.concatenate([x, y])
            v_lo  = float(np.nanmin(all_v))
            v_hi  = float(np.nanmax(all_v))
            margin = (v_hi - v_lo) * 0.05 if v_hi != v_lo else 1.0
            ref_lo = v_lo - margin
            ref_hi = v_hi + margin
            ref = np.array([ref_lo, ref_hi])

            # 1:1 line
            ax.plot(ref, ref, "k-", lw=0.8, alpha=0.5, zorder=1, label="1:1")
            # ±10 % deviation lines
            ax.plot(ref, ref * 1.10, color="#d62728", ls="--", lw=0.7,
                    alpha=0.6, zorder=1, label="\u00b110%")
            ax.plot(ref, ref * 0.90, color="#d62728", ls="--", lw=0.7,
                    alpha=0.6, zorder=1)

            ax.set_xlabel("Mean Annual Baseline")
            ax.set_ylabel("Mean Annual Product A")
            ax.set_title(f"{title} (n={len(x)})")
            ax.legend(loc="upper left", framealpha=0.7)

            ax.set_xlim(ref_lo, ref_hi)
            ax.set_ylim(ref_lo, ref_hi)

            # Grid
            ax.grid(True, which="major", ls="-", lw=0.4, alpha=0.3)

            fig.savefig(out_path, bbox_inches="tight")
            plt.close(fig)

        # One per category, split by units
        _n_scatter = 0
        for cat in cat_order_plots:
            cat_sub = df_plots[df_plots["Input_Category"] == cat]
            safe_cat = cat.replace(" ", "_").replace("/", "_")
            # Get distinct units in this category
            _units_in_cat = sorted(cat_sub["Units"].dropna().unique())
            for unit in (_units_in_cat or [""]):
                unit_sub = cat_sub[cat_sub["Units"] == unit] if unit else cat_sub
                safe_unit = unit.replace("/", "_").replace(" ", "_")
                suffix = f"_{safe_unit}" if len(_units_in_cat) > 1 else ""
                unit_label = f" [{unit}]" if unit else ""
                _scatter_base_vs_mod(
                    unit_sub,
                    f"{cat}{unit_label} — Baseline vs Product A",
                    _scatter_dir / f"{safe_cat}{suffix}.png",
                )
                _n_scatter += 1

        print(f"  Scatter plots:  figures/annual_scatter/  "
              f"({_n_scatter} plots across {len(cat_order_plots)} categories)")

    # ══════════════════════════════════════════════════════════════════════
    # 3.  Per-variable diagnostic plots
    # ══════════════════════════════════════════════════════════════════════
    print("  Generating per-variable diagnostic plots ...")
    if not CLI_ARGS.no_term_plots:
        generate_per_variable_plots(stats_csv, cached_series=cached_series)
    else:
        print("    (skipped — --no-term-plots flag set)")

    return rpt_path


def compute_modification_statistics(all_modified_keys, baseline_bucket):
    """Compare baseline vs. modified DSS for every modified (Part B, Part C).

    Returns ``(n_paths, cached_series)`` where *cached_series* is a dict
    ``{(Part_B, Part_C): {"b_ser": pd.Series, "m_ser": pd.Series, "units": str}}``
    that can be passed directly to ``generate_per_variable_plots`` so DSS files
    do not need to be re-read.
    """
    # Units that should use MEAN for annual aggregation (rates, intensities,
    # concentrations, temperatures, dimensionless).  Everything else uses SUM.
    _MEAN_UNITS = {"CFS", "MG/L", "FAHRENHE", "KPA", "NONE", "DAYS"}
    _MONTHS_WY = [
        ("Oct", 10), ("Nov", 11), ("Dec", 12),
        ("Jan", 1),  ("Feb", 2),  ("Mar", 3),
        ("Apr", 4),  ("May", 5),  ("Jun", 6),
        ("Jul", 7),  ("Aug", 8),  ("Sep", 9),
    ]

    import warnings

    def _r2(obs, sim):
        if len(obs) < 2:
            return np.nan
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", RuntimeWarning)
            c = np.corrcoef(obs, sim)[0, 1]
        return c ** 2 if np.isfinite(c) else np.nan

    def _nse(obs, sim):
        if len(obs) < 2:
            return np.nan
        ss_res = np.sum((sim - obs) ** 2)
        ss_tot = np.sum((obs - np.mean(obs)) ** 2)
        return 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan

    stat_rows = []
    cached_series = {}   # (B,C) → {"b_ser", "m_ser", "units"}
    n_read_errors = 0

    # Look up Input_Category for each (Part_B, Part_C)
    _inv = read_master_inventory()
    _cat_lookup = {}
    for _, _ir in _inv.iterrows():
        _cat_lookup[(_ir["Part_B"], _ir["Part_C"])] = _ir["Input_Category"]

    with HecDss.Open(str(BASELINE_DSS), version=6, catalog_flag=True) as dss_base_r, \
         HecDss.Open(_dss_str(OUTPUT_DSS), version=6, catalog_flag=True) as dss_mod_r:

        for pk in sorted(all_modified_keys):
            if pk not in baseline_bucket:
                continue

            # Merge ALL D-part blocks for this (B,C) into single series
            base_dict = {}
            mod_dict  = {}
            units     = ""
            for pathname in baseline_bucket[pk]:
                try:
                    ts_base = dss_base_r.read_ts(pathname, trim_missing=False)
                    ts_mod  = dss_mod_r.read_ts(pathname, trim_missing=False)
                except Exception:
                    n_read_errors += 1
                    continue
                eom = dss_eom(ts_base.pytimes)
                b_vals = np.array(ts_base.values, dtype=float)
                m_vals = np.array(ts_mod.values, dtype=float)
                if not units:
                    units = (ts_base.units or "").strip().upper()
                for i, dt in enumerate(eom):
                    if b_vals[i] > -900:
                        base_dict[dt] = b_vals[i]
                    if m_vals[i] > -900:
                        mod_dict[dt] = m_vals[i]

            if not base_dict or not mod_dict:
                continue

            # Build aligned series within the overwrite window
            all_dates = sorted(set(base_dict.keys()) & set(mod_dict.keys()))
            all_dates = [d for d in all_dates
                         if OVERWRITE_START <= d <= OVERWRITE_END]
            if not all_dates:
                continue

            eom_v = pd.DatetimeIndex(all_dates)
            b_v = pd.Series([base_dict[d] for d in all_dates], index=eom_v)
            m_v = pd.Series([mod_dict[d] for d in all_dates], index=eom_v)

            # Cache the merged series for later use by plot generation
            cached_series[pk] = {"b_ser": b_v, "m_ser": m_v, "units": units}

            chg = m_v - b_v
            pct = (chg / b_v.replace(0, np.nan)) * 100.0

            # Flag paths where both baseline & modified are effectively zero
            _ZERO_TOL = 1e-6
            all_zero = (np.max(np.abs(b_v.values)) < _ZERO_TOL and
                        np.max(np.abs(m_v.values)) < _ZERO_TOL)

            row = {"Part_B": pk[0], "Part_C": pk[1], "Units": units,
                   "Pathname": baseline_bucket[pk][0],
                   "Input_Category": _cat_lookup.get(pk, "Unknown"),
                   "AllZero": all_zero}

            for mname, mnum in _MONTHS_WY:
                mm = b_v.index.month == mnum
                if mm.sum() == 0:
                    continue
                bm = b_v[mm]; mv = m_v[mm]; cm = chg[mm]; pm = pct[mm]
                row[f"Base-{mname}-Avg"]      = bm.mean()
                row[f"Base-{mname}-Min"]      = bm.min()
                row[f"Base-{mname}-Max"]      = bm.max()
                row[f"Mod-{mname}-Avg"]       = mv.mean()
                row[f"Mod-{mname}-Min"]       = mv.min()
                row[f"Mod-{mname}-Max"]       = mv.max()
                row[f"AbsChange-{mname}-Avg"] = cm.mean()
                row[f"AbsChange-{mname}-Min"] = cm.min()
                row[f"AbsChange-{mname}-Max"] = cm.max()
                row[f"PctChange-{mname}-Avg"] = pm.mean()
                row[f"PctChange-{mname}-Min"] = pm.min()
                row[f"PctChange-{mname}-Max"] = pm.max()

            row["R2-Monthly"]  = _r2(b_v.values, m_v.values)
            row["NSE-Monthly"] = _nse(b_v.values, m_v.values)

            wy_labels = np.where(eom_v.month >= 10,
                                 eom_v.year + 1, eom_v.year)
            df_wy = pd.DataFrame({"base": b_v.values, "mod": m_v.values},
                                 index=wy_labels)
            wy_counts = df_wy.groupby(level=0).size()
            complete  = wy_counts[wy_counts == 12].index
            df_wy     = df_wy.loc[df_wy.index.isin(complete)]

            if not df_wy.empty:
                use_mean = any(u in units for u in _MEAN_UNITS)
                agg = "mean" if use_mean else "sum"
                ann = df_wy.groupby(level=0).agg(agg)
                ann["change"] = ann["mod"] - ann["base"]
                ann["pct"]    = (ann["change"]
                                 / ann["base"].replace(0, np.nan)) * 100.0
                row["Base-Ann-Avg"]      = ann["base"].mean()
                row["Base-Ann-Min"]      = ann["base"].min()
                row["Base-Ann-Max"]      = ann["base"].max()
                row["Mod-Ann-Avg"]       = ann["mod"].mean()
                row["Mod-Ann-Min"]       = ann["mod"].min()
                row["Mod-Ann-Max"]       = ann["mod"].max()
                row["AbsChange-Ann-Avg"] = ann["change"].mean()
                row["AbsChange-Ann-Min"] = ann["change"].min()
                row["AbsChange-Ann-Max"] = ann["change"].max()
                row["PctChange-Ann-Avg"] = ann["pct"].mean()
                row["PctChange-Ann-Min"] = ann["pct"].min()
                row["PctChange-Ann-Max"] = ann["pct"].max()
                row["R2-Ann"]            = _r2(ann["base"].values,
                                               ann["mod"].values)
                row["NSE-Ann"]           = _nse(ann["base"].values,
                                                ann["mod"].values)
                row["Ann_Aggregation"]   = agg

            stat_rows.append(row)

    _id_cols = ["Part_B", "Part_C", "Units", "Input_Category", "Pathname"]
    _metric_cols = []
    for mname, _ in _MONTHS_WY:
        for metric in ["Base", "Mod", "AbsChange", "PctChange"]:
            for stat in ["Avg", "Min", "Max"]:
                _metric_cols.append(f"{metric}-{mname}-{stat}")
    _metric_cols += ["R2-Monthly", "NSE-Monthly"]
    for metric in ["Base", "Mod", "AbsChange", "PctChange"]:
        for stat in ["Avg", "Min", "Max"]:
            _metric_cols.append(f"{metric}-Ann-{stat}")
    _metric_cols += ["R2-Ann", "NSE-Ann", "Ann_Aggregation"]

    stats_df = pd.DataFrame(stat_rows)
    ordered  = [c for c in _id_cols + _metric_cols if c in stats_df.columns]
    stats_df = stats_df[ordered]

    fp = OUTPUT_DIR / "modification_statistics.csv"
    stats_df.to_csv(fp, index=False)

    if n_read_errors:
        print(f"    (note: {n_read_errors} individual DSS block reads failed)")

    return len(stats_df), cached_series


# ══════════════════════════════════════════════════════════════════════════════
# STEP 0 — Pre-flight
# ══════════════════════════════════════════════════════════════════════════════

# ── Compute-stats-only mode ──────────────────────────────────────────────────
if CLI_ARGS.compute_stats:
    print("=" * 72)
    print("  Product A Historical Validation — Compute Statistics Only")
    print("=" * 72)

    if not OUTPUT_DSS.exists():
        sys.exit(
            f"ERROR: Compiled DSS not found — run the full compilation first.\n"
            f"  Expected: {OUTPUT_DSS}"
        )
    if not BASELINE_DSS.exists():
        sys.exit(f"ERROR: Baseline DSS not found:\n  {BASELINE_DSS}")

    paths_mod_csv = OUTPUT_DIR / "paths_modified.csv"
    if not paths_mod_csv.exists():
        sys.exit(
            f"ERROR: paths_modified.csv not found — run the full compilation first.\n"
            f"  Expected: {paths_mod_csv}"
        )

    # Reconstruct all_modified_keys from the diagnostic CSV
    _pm = pd.read_csv(paths_mod_csv)
    all_modified_keys = set(
        zip(_pm["Part B"].str.upper(), _pm["Part C"].str.upper())
    )
    print(f"  Modified (Part B,C) from paths_modified.csv: {len(all_modified_keys):,}")

    # Reconstruct baseline_bucket from the baseline DSS
    print("  Cataloging baseline DSS ...")
    with HecDss.Open(str(BASELINE_DSS), version=6, catalog_flag=True) as _dss_b:
        _bp = _dss_b.getPathnameList(DSS_PATTERN)
    baseline_bucket = {}
    for p in _bp:
        k = path_key(p)
        baseline_bucket.setdefault(k, []).append(p)
    print(f"  Baseline unique (Part B, Part C): {len(baseline_bucket):,}")
    print()

    # Run stats and exit
    print("Computing baseline vs. modified statistics ...")
    t_stats = time.time()
    n_stats, _cached = compute_modification_statistics(all_modified_keys, baseline_bucket)
    fp = OUTPUT_DIR / "modification_statistics.csv"
    print(f"  {fp.name:45s}  {n_stats:>6,} paths  ({time.time()-t_stats:.1f}s)")
    print()

    if CLI_ARGS.stats_report:
        print("Generating summary report & figures ...")
        generate_stats_report(fp, cached_series=_cached)

    print(f"\nDone.  Outputs in: {OUTPUT_DIR}")
    sys.exit(0)


# ── Stats-report-only mode (no compilation, no stats recompute) ──────────────
if CLI_ARGS.stats_report:
    print("=" * 72)
    print("  Product A Historical Validation — Generate Stats Report")
    print("=" * 72)

    stats_csv = OUTPUT_DIR / "modification_statistics.csv"
    if not stats_csv.exists():
        sys.exit(
            f"ERROR: modification_statistics.csv not found — "
            f"run --compute-stats first.\n  Expected: {stats_csv}"
        )

    print("Generating summary report & figures ...")
    generate_stats_report(stats_csv)
    print(f"\nDone.  Outputs in: {OUTPUT_DIR}")
    sys.exit(0)


print("=" * 72)
print("  Product A Historical Validation — DSS Compilation")
print("=" * 72)

if not BASELINE_DSS.exists():
    sys.exit(f"ERROR: Baseline DSS not found:\n  {BASELINE_DSS}")
if not INVENTORY_XLSX.exists():
    sys.exit(f"ERROR: Master inventory not found:\n  {INVENTORY_XLSX}")

# Warn early if the inventory Excel is locked (open in Excel) so the user
# can close it and press Enter to continue — before the long run begins.
while True:
    try:
        with open(INVENTORY_XLSX, "r+b"):
            pass
        break  # file is accessible
    except PermissionError:
        print(
            f"\nWARNING: The master inventory appears to be open in Excel.\n"
            f"  {INVENTORY_XLSX}\n"
            f"  Close the file, then press Enter to continue (or Ctrl+C to abort)..."
        )
        try:
            input()
        except KeyboardInterrupt:
            sys.exit("\nAborted.")

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
COMPILED_CSV.mkdir(parents=True, exist_ok=True)

# ── Locate source CSVs per module ────────────────────────────────────────────
available_modules = OrderedDict()   # label → list[Path] of data CSVs
missing_modules   = []
module_csv_mtimes = OrderedDict()   # label → {csv_name: mtime_str}

for label, (src_dir, _inv_cat) in MODULE_CONFIG.items():
    csvs = collect_data_csvs(src_dir)
    if csvs:
        available_modules[label] = csvs
        mtimes = {}
        for c in csvs:
            mt = time.strftime("%Y-%m-%d %H:%M", time.localtime(os.path.getmtime(c)))
            mtimes[c.name] = mt
        module_csv_mtimes[label] = mtimes
        most_recent = max(mtimes.values())
        print(f"  [OK]   {label:40s}  {len(csvs)} CSV(s)  (latest: {most_recent})")
        for c in csvs:
            print(f"           {c.name:50s}  modified: {mtimes[c.name]}")
    else:
        missing_modules.append(label)
        print(f"  [SKIP] {label:40s}  (no data CSVs found)")

if not available_modules:
    sys.exit("ERROR: No data CSVs found in any module. Nothing to compile.")

print(f"\nModules available: {len(available_modules)} / {len(MODULE_CONFIG)}")
if missing_modules:
    print(f"Modules skipped:  {', '.join(missing_modules)}")
print()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Copy CSVs into compiled_input_files/<module>/
# ══════════════════════════════════════════════════════════════════════════════
print("Step 1: Copying source CSVs into compiled_input_files/ ...")
for label, csvs in available_modules.items():
    dest_dir = COMPILED_CSV / label
    dest_dir.mkdir(parents=True, exist_ok=True)
    for csv_path in csvs:
        shutil.copy2(csv_path, dest_dir / csv_path.name)
    print(f"  {label:40s}  → {len(csvs)} file(s) copied")
print()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Catalog baseline DSS
# ══════════════════════════════════════════════════════════════════════════════
print("Step 2: Cataloging baseline DSS ...")
with HecDss.Open(str(BASELINE_DSS), version=6, catalog_flag=True) as dss_base:
    baseline_paths = dss_base.getPathnameList(DSS_PATTERN)

baseline_keys = set()
baseline_bucket = {}   # (B,C) → [pathnames]
for p in baseline_paths:
    k = path_key(p)
    baseline_keys.add(k)
    baseline_bucket.setdefault(k, []).append(p)

print(f"  Baseline monthly paths:   {len(baseline_paths):,}")
print(f"  Unique (Part B, Part C):  {len(baseline_keys):,}")
print()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Build per-module intermediate DSS files
# ══════════════════════════════════════════════════════════════════════════════
print("Step 3: Building intermediate DSS per module ...")
print("-" * 72)

module_dss_paths   = OrderedDict()   # label → Path to intermediate DSS
module_modified    = {}              # label → set of (B,C) keys modified
module_not_in_base = []              # records of (B,C) not found in baseline

for label, csvs in available_modules.items():
    t0 = time.time()

    # ── 3a. Read & combine CSVs for this module ─────────────────────────────
    frames = []
    for f in csvs:
        df = pd.read_csv(f)
        df.columns = [c.strip() for c in df.columns]
        frames.append(df)
    csv_df = pd.concat(frames, ignore_index=True)

    required = {"Part B", "Part C", "Year", "Month", "Value"}
    missing_cols = required - set(csv_df.columns)
    if missing_cols:
        print(f"  [WARN] {label}: CSV missing columns {sorted(missing_cols)} — skipping")
        continue

    csv_df["PARTB"] = csv_df["Part B"].map(excel_to_part)
    csv_df["PARTC"] = csv_df["Part C"].map(excel_to_part)
    csv_df["DATE"]  = [ym_to_eom(y, m) for y, m in zip(csv_df["Year"], csv_df["Month"])]
    csv_df["Value"] = pd.to_numeric(csv_df["Value"], errors="coerce")

    override_dict = {}
    for (partb, partc), g in csv_df.groupby(["PARTB", "PARTC"]):
        s = pd.Series(g["Value"].values, index=pd.DatetimeIndex(g["DATE"]))
        s = s.sort_index().dropna()
        if len(s) > 0:
            override_dict[(partb, partc)] = s

    # ── 3b. Create intermediate DSS ─────────────────────────────────────────
    int_dss = OUTPUT_DIR / f"{label}__productA.dss"
    if int_dss.exists():
        int_dss.unlink()
    for ext in [".dsd", ".dsk", ".dsc"]:
        companion = int_dss.with_suffix(ext)
        if companion.exists():
            companion.unlink()

    n_written   = 0
    n_not_found = 0
    mod_keys_modified = set()
    missing_parts = []

    with HecDss.Open(str(BASELINE_DSS), version=6, catalog_flag=True) as dss_in, \
         HecDss.Open(_dss_str(int_dss), version=6) as dss_out:

        for part_key in sorted(override_dict):
            if part_key not in baseline_bucket:
                missing_parts.append(part_key)
                n_not_found += 1
                continue

            partb, partc = part_key
            overrides = override_dict[part_key]
            # Clip to overwrite window
            overrides = overrides.loc[
                (overrides.index >= OVERWRITE_START) & (overrides.index <= OVERWRITE_END)
            ]
            if overrides.empty:
                continue

            override_index  = overrides.index
            override_values = overrides.values.astype(float)

            for pathname in baseline_bucket[part_key]:
                ts = dss_in.read_ts(pathname, trim_missing=False)
                eom_idx = dss_eom(ts.pytimes)
                eom_pd  = pd.DatetimeIndex(eom_idx)

                locs  = eom_pd.get_indexer(override_index)
                valid = locs >= 0

                ts_vals = np.array(ts.values, dtype=float).copy()

                if np.any(valid):
                    ts_vals[locs[valid]] = override_values[valid]

                tsc = TimeSeriesContainer()
                tsc.pathname      = pathname
                tsc.startDateTime = ts.pytimes[0].strftime("%d%b%Y %H:%M")
                tsc.numberValues  = len(ts_vals)
                tsc.units         = ts.units
                tsc.type          = ts.type
                tsc.interval      = ts.interval
                tsc.values        = ts_vals

                safe_write_ts(dss_out, pathname, tsc)
                n_written += 1

            mod_keys_modified.add(part_key)

    module_dss_paths[label] = int_dss
    module_modified[label]  = mod_keys_modified

    for pk in missing_parts:
        module_not_in_base.append({
            "Part B": pk[0], "Part C": pk[1], "Module": label,
        })

    nf_tag = f"  ({n_not_found} not in baseline)" if n_not_found else ""
    print(f"  {label:40s}  {n_written:>5} paths  |  "
          f"{len(mod_keys_modified):>4} (B,C) modified  ({time.time()-t0:.1f}s){nf_tag}")

print("-" * 72)
print()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3b — Auto-fill Constant/Rept SVs from baseline last-12-month pattern
# ══════════════════════════════════════════════════════════════════════════════
print("Step 3b: Auto-filling Constant/Rept SVs ...")

# Read inventory early to identify Constant/Rept keys
inventory_df = read_master_inventory()
print(f"  Inventory rows loaded: {len(inventory_df):,}")

# Category → module label mapping
inv_category_map = {}
for label, (_rel, inv_cat) in MODULE_CONFIG.items():
    inv_category_map[inv_cat] = label

# Identify Constant/Rept SVs that belong to a configured module category
const_rept_svs = inventory_df[
    (inventory_df["Constant_Rept"] == True)
    & (~inventory_df["Missing"])
    & (inventory_df["Used_in_DCR"])
    & (inventory_df["Input_Category"].isin(inv_category_map.keys()))
].copy()
const_rept_keys = set(zip(const_rept_svs["Part_B"], const_rept_svs["Part_C"]))

# Also include Constant/Rept SVs from categories WITHOUT dedicated modules
const_rept_other = inventory_df[
    (inventory_df["Constant_Rept"] == True)
    & (~inventory_df["Missing"])
    & (inventory_df["Used_in_DCR"])
    & (~inventory_df["Input_Category"].isin(inv_category_map.keys()))
].copy()
const_rept_keys |= set(zip(const_rept_other["Part_B"], const_rept_other["Part_C"]))

# Remove any that were already handled by module CSVs
already_from_csv = set()
for keys in module_modified.values():
    already_from_csv |= keys
const_rept_to_fill = const_rept_keys - already_from_csv

const_rept_dss = OUTPUT_DIR / "_constant_rept__productA.dss"
if const_rept_dss.exists():
    const_rept_dss.unlink()
for ext in [".dsd", ".dsk", ".dsc"]:
    companion = const_rept_dss.with_suffix(ext)
    if companion.exists():
        companion.unlink()

n_const_written = 0
const_rept_filled = set()

if const_rept_to_fill:
    with HecDss.Open(str(BASELINE_DSS), version=6, catalog_flag=True) as dss_in, \
         HecDss.Open(_dss_str(const_rept_dss), version=6) as dss_out:

        for pk in sorted(const_rept_to_fill):
            overrides = build_constant_rept_overrides(
                dss_in, baseline_bucket, pk, OVERWRITE_START, OVERWRITE_END
            )
            for pathname, vals in overrides.items():
                ts = dss_in.read_ts(pathname, trim_missing=False)
                tsc = TimeSeriesContainer()
                tsc.pathname      = pathname
                tsc.startDateTime = ts.pytimes[0].strftime("%d%b%Y %H:%M")
                tsc.numberValues  = len(vals)
                tsc.units         = ts.units
                tsc.type          = ts.type
                tsc.interval      = ts.interval
                tsc.values        = vals
                safe_write_ts(dss_out, pathname, tsc)
                n_const_written += 1
            if overrides:
                const_rept_filled.add(pk)

    module_dss_paths["_constant_rept"] = const_rept_dss
    module_modified["_constant_rept"]  = const_rept_filled

print(f"  Constant/Rept SVs to auto-fill:  {len(const_rept_to_fill):,}")
print(f"  Successfully written:            {n_const_written:,} paths ({len(const_rept_filled)} B,C)")
print(f"  Already provided by module CSVs: {len(const_rept_keys & already_from_csv):,}")
print()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — Compile final DSS from intermediate DSS files
# ══════════════════════════════════════════════════════════════════════════════
print("Step 4: Compiling final ProductA_Historical_Validation_SV.dss ...")
t_start = time.time()

# Start from a copy of the baseline
if OUTPUT_DSS.exists():
    OUTPUT_DSS.unlink()
for ext in [".dsd", ".dsk", ".dsc"]:
    companion = OUTPUT_DSS.with_suffix(ext)
    if companion.exists():
        companion.unlink()

shutil.copy2(BASELINE_DSS, OUTPUT_DSS)
for ext in [".dsd", ".dsk", ".dsc"]:
    src = BASELINE_DSS.with_suffix(ext)
    if src.exists():
        shutil.copy2(src, OUTPUT_DSS.with_suffix(ext))

# Layer each intermediate DSS onto the output
modified_records  = []
all_modified_keys = set()

for label, int_dss in module_dss_paths.items():
    with HecDss.Open(_dss_str(int_dss), version=6, catalog_flag=True) as dss_mod, \
         HecDss.Open(_dss_str(OUTPUT_DSS), version=6) as dss_out:

        mod_paths = dss_mod.getPathnameList(DSS_PATTERN)
        for mp in mod_paths:
            mk = path_key(mp)
            try:
                ts = dss_mod.read_ts(mp, trim_missing=False)
            except Exception as e:
                print(f"    WARNING: Could not read {mp} from {label}: {e}")
                continue

            tsc = TimeSeriesContainer()
            tsc.pathname      = mp
            tsc.startDateTime = ts.pytimes[0].strftime("%d%b%Y %H:%M")
            tsc.numberValues  = len(ts.values)
            tsc.units         = ts.units
            tsc.type          = ts.type
            tsc.interval      = ts.interval
            tsc.values        = np.array(ts.values, dtype=float)

            safe_write_ts(dss_out, mp, tsc)

            modified_records.append({
                "Part B": mk[0], "Part C": mk[1],
                "Module": label, "Pathname": mp,
            })
            all_modified_keys.add(mk)

print(f"  Final DSS built in {time.time()-t_start:.1f}s")
print(f"  Total paths written from modules: {len(modified_records):,}")
print()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5 — Inventory cross-reference (flag-aware)
# ══════════════════════════════════════════════════════════════════════════════
print("Step 5: Cross-referencing against master inventory ...")

# ── Classify inventory SVs by flags ──────────────────────────────────────────
# Skipped: Missing=T
inv_skipped_missing = inventory_df[inventory_df["Missing"] == True].copy()
skipped_missing_keys = set(zip(inv_skipped_missing["Part_B"], inv_skipped_missing["Part_C"]))

# Skipped: Used_in_DCR=F (and not already Missing)
inv_skipped_not_dcr = inventory_df[
    (~inventory_df["Missing"]) & (~inventory_df["Used_in_DCR"])
].copy()
skipped_not_dcr_keys = set(zip(inv_skipped_not_dcr["Part_B"], inv_skipped_not_dcr["Part_C"]))

# Constant/Rept (already handled in 3b)
inv_const_rept_all = inventory_df[
    (inventory_df["Constant_Rept"] == True)
    & (~inventory_df["Missing"])
    & (inventory_df["Used_in_DCR"])
].copy()
const_rept_all_keys = set(zip(inv_const_rept_all["Part_B"], inv_const_rept_all["Part_C"]))

# Active SVs: not missing, used in DCR, and NOT constant/rept
# These are the ones that MUST come from module CSVs
active_svs = inventory_df[
    (~inventory_df["Missing"])
    & (inventory_df["Used_in_DCR"])
    & (~inventory_df["Constant_Rept"])
    & (inventory_df["Input_Category"].isin(
        set(inv_category_map.keys()) | CATEGORIES_WITHOUT_VALIDATION
    ))
].copy()
active_svs["Module"] = active_svs["Input_Category"].map(inv_category_map).fillna("(no module)")

# Expected from CSV modules only (not constant/rept, not skipped categories)
expected_svs = active_svs[
    active_svs["Input_Category"].isin(inv_category_map.keys())
].copy()
expected_keys = set(zip(expected_svs["Part_B"], expected_svs["Part_C"]))

# Actually modified (B,C) across all modules (includes constant_rept auto-fill)
actual_keys = all_modified_keys

# ── Inventory diagnostic sets ────────────────────────────────────────────────
inv_expected_modified = expected_keys & actual_keys
inv_expected_missing  = expected_keys - actual_keys
inv_unexpected        = actual_keys - expected_keys - const_rept_all_keys  # don't flag const/rept as unexpected

print(f"  Total inventory SVs:                  {len(inventory_df):,}")
print(f"  Skipped (Missing=T):                  {len(skipped_missing_keys):,}")
print(f"  Skipped (Used_in_DCR=F):              {len(skipped_not_dcr_keys):,}")
print(f"  Constant/Rept (auto-filled):          {len(const_rept_filled):,}")
print(f"  Expected from CSV modules:            {len(expected_keys):,}")
print(f"  Actually modified (Part B,C):          {len(actual_keys):,}")
print(f"  Expected & modified:                   {len(inv_expected_modified):,}")
print(f"  Expected but MISSING:                  {len(inv_expected_missing):,}")
print(f"  Modified but NOT in inventory:         {len(inv_unexpected):,}")
print()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 6 — Write all diagnostic CSVs
# ══════════════════════════════════════════════════════════════════════════════
print("Step 6: Writing diagnostics ...")

# ── paths_modified.csv ───────────────────────────────────────────────────────
modified_df = pd.DataFrame(modified_records)
if not modified_df.empty:
    modified_df = modified_df.sort_values(["Module", "Part B", "Part C"]).reset_index(drop=True)
fp = OUTPUT_DIR / "paths_modified.csv"
modified_df.to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(modified_df):>6,} records")

# ── paths_unchanged.csv ─────────────────────────────────────────────────────
unchanged_keys = baseline_keys - all_modified_keys
unchanged_records = []
for k in sorted(unchanged_keys):
    for p in baseline_bucket[k]:
        unchanged_records.append({"Part B": k[0], "Part C": k[1], "Pathname": p})
unchanged_df = pd.DataFrame(unchanged_records)
if not unchanged_df.empty:
    unchanged_df = unchanged_df.sort_values(["Part B", "Part C"]).reset_index(drop=True)
fp = OUTPUT_DIR / "paths_unchanged.csv"
unchanged_df.to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(unchanged_df):>6,} records")

# ── paths_not_in_baseline.csv ───────────────────────────────────────────────
nib_df = pd.DataFrame(module_not_in_base)
if not nib_df.empty:
    nib_df = nib_df.sort_values(["Module", "Part B", "Part C"]).reset_index(drop=True)
fp = OUTPUT_DIR / "paths_not_in_baseline.csv"
nib_df.to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(nib_df):>6,} records")

# ── inventory_expected_modified.csv ──────────────────────────────────────────
inv_mod_rows = expected_svs[
    expected_svs.apply(lambda r: (r["Part_B"], r["Part_C"]) in inv_expected_modified, axis=1)
].copy()
inv_mod_rows = inv_mod_rows.sort_values(["Module", "Part_B", "Part_C"]).reset_index(drop=True)
fp = OUTPUT_DIR / "inventory_expected_modified.csv"
inv_mod_rows.to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(inv_mod_rows):>6,} records")

# ── inventory_expected_missing.csv ───────────────────────────────────────────
inv_miss_rows = expected_svs[
    expected_svs.apply(lambda r: (r["Part_B"], r["Part_C"]) in inv_expected_missing, axis=1)
].copy()
inv_miss_rows = inv_miss_rows.sort_values(["Module", "Part_B", "Part_C"]).reset_index(drop=True)
fp = OUTPUT_DIR / "inventory_expected_missing.csv"
inv_miss_rows.to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(inv_miss_rows):>6,} records")

# ── inventory_constant_rept.csv ──────────────────────────────────────────────
inv_cr_rows = inv_const_rept_all.copy()
inv_cr_rows["Auto_Filled"] = inv_cr_rows.apply(
    lambda r: (r["Part_B"], r["Part_C"]) in const_rept_filled, axis=1
)
inv_cr_rows["From_CSV"] = inv_cr_rows.apply(
    lambda r: (r["Part_B"], r["Part_C"]) in already_from_csv, axis=1
)
inv_cr_rows = inv_cr_rows.sort_values(["Input_Category", "Part_B", "Part_C"]).reset_index(drop=True)
fp = OUTPUT_DIR / "inventory_constant_rept.csv"
inv_cr_rows.to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(inv_cr_rows):>6,} records")

# ── inventory_skipped_missing.csv ────────────────────────────────────────────
fp = OUTPUT_DIR / "inventory_skipped_missing.csv"
inv_skipped_missing.sort_values(["Input_Category", "Part_B", "Part_C"]).reset_index(drop=True).to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(inv_skipped_missing):>6,} records")

# ── inventory_skipped_not_in_dcr.csv ─────────────────────────────────────────
fp = OUTPUT_DIR / "inventory_skipped_not_in_dcr.csv"
inv_skipped_not_dcr.sort_values(["Input_Category", "Part_B", "Part_C"]).reset_index(drop=True).to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(inv_skipped_not_dcr):>6,} records")

# ── inventory_unexpected.csv ─────────────────────────────────────────────────
inv_unexp_rows = []
for mk in sorted(inv_unexpected):
    mods = [r["Module"] for r in modified_records
            if (r["Part B"], r["Part C"]) == mk]
    inv_unexp_rows.append({
        "Part_B": mk[0], "Part_C": mk[1],
        "Module": mods[0] if mods else "unknown",
    })
inv_unexp_df = pd.DataFrame(inv_unexp_rows)
fp = OUTPUT_DIR / "inventory_unexpected.csv"
inv_unexp_df.to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(inv_unexp_df):>6,} records")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 7 — Compilation summary
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
n_mod_unique  = len(all_modified_keys)
n_unch_unique = len(unchanged_keys)
n_nib_unique  = len(set((r["Part B"], r["Part C"]) for r in module_not_in_base)) if module_not_in_base else 0

lines = [
    "=" * 65,
    "  Product A Historical Validation — Compilation Summary",
    "=" * 65,
    "",
    f"  Baseline DSS:  {BASELINE_DSS.name}",
    f"  Output DSS:    {OUTPUT_DSS.name}",
    f"  Inventory:     {INVENTORY_XLSX.name}",
    "",
    f"  Baseline monthly paths (total):        {len(baseline_paths):>6,}",
    f"  Baseline unique (Part B, Part C):      {len(baseline_keys):>6,}",
    "",
    "─── Module Contributions ───",
    "",
]

for label in MODULE_CONFIG:
    if label in module_modified:
        n = len(module_modified[label])
        lines.append(f"    {label:40s}  {n:>5} (B,C)")
    else:
        lines.append(f"    {label:40s}  (not found)")

lines += [
    "",
    "─── Path Categories ───",
    "",
    f"    Modified (Part B,C) combinations:    {n_mod_unique:>6,}",
    f"    Modified DSS paths (total):           {len(modified_records):>6,}",
    "",
    f"    Unchanged (Part B,C) combinations:   {n_unch_unique:>6,}",
    f"    Unchanged DSS paths (total):          {len(unchanged_records):>6,}",
    "",
    f"    Not in baseline (Part B,C):          {n_nib_unique:>6,}",
    "",
    "─── Inventory Flags ───",
    "",
    f"    Total inventory SVs:                 {len(inventory_df):>6,}",
    f"    Skipped — Missing=T:                 {len(skipped_missing_keys):>6,}",
    f"    Skipped — Used_in_DCR=F:             {len(skipped_not_dcr_keys):>6,}",
    f"    Constant/Rept (total):               {len(const_rept_all_keys):>6,}",
    f"      Auto-filled by script:             {len(const_rept_filled):>6,}",
    f"      Already in module CSVs:            {len(const_rept_keys & already_from_csv):>6,}",
    "",
    "─── Inventory Cross-Reference (active SVs) ───",
    "",
    f"    Expected from CSV modules:           {len(expected_keys):>6,}",
    f"    Expected & successfully modified:     {len(inv_expected_modified):>6,}",
    f"    Expected but MISSING:                {len(inv_expected_missing):>6,}",
    f"    Modified but NOT in inventory:        {len(inv_unexpected):>6,}",
    "",
    "─── Categories Without Validation Modules ───",
    "",
]

for cat in sorted(CATEGORIES_WITHOUT_VALIDATION):
    cnt_active = len(inventory_df[
        (inventory_df["Input_Category"] == cat)
        & (~inventory_df["Missing"])
        & (inventory_df["Used_in_DCR"])
        & (~inventory_df["Constant_Rept"])
    ])
    cnt_cr = len(inventory_df[
        (inventory_df["Input_Category"] == cat)
        & (inventory_df["Constant_Rept"] == True)
        & (~inventory_df["Missing"])
        & (inventory_df["Used_in_DCR"])
    ])
    cnt_total = len(inventory_df[inventory_df["Input_Category"] == cat])
    lines.append(f"    {cat:40s}  {cnt_total:>5} total  ({cnt_cr} const/rept, {cnt_active} other)")

lines += [
    "",
    "─── Source CSV Modification Dates ───",
    "",
]
for label, mtimes in module_csv_mtimes.items():
    lines.append(f"    {label}")
    for csv_name, mtime_str in sorted(mtimes.items()):
        lines.append(f"      {csv_name:50s}  {mtime_str}")
    lines.append("")

lines += [
    "─── Modules Skipped (no data CSVs) ───",
    "",
]
if missing_modules:
    for m in missing_modules:
        lines.append(f"    {m}")
else:
    lines.append("    (none)")

lines += ["", "=" * 65]

summary_path = OUTPUT_DIR / "compilation_summary.txt"
summary_path.write_text("\n".join(lines), encoding="utf-8")

print(f"\n  {summary_path.name}")
print()
for line in lines:
    print(line)

print(f"\nDone.  Output: {OUTPUT_DSS}")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 8 — Compute modification statistics & generate report
# ══════════════════════════════════════════════════════════════════════════════
print()
print("Step 8: Computing modification statistics ...")
t_stats = time.time()
n_stats, _cached = compute_modification_statistics(all_modified_keys, baseline_bucket)
stats_csv = OUTPUT_DIR / "modification_statistics.csv"
print(f"  {stats_csv.name:45s}  {n_stats:>6,} paths  ({time.time()-t_stats:.1f}s)")

print("  Generating summary report & figures ...")
generate_stats_report(stats_csv, cached_series=_cached)

print(f"\nAll steps complete.  Outputs in: {OUTPUT_DIR}")
