"""
Compile Product B (1000-year stochastic) SV CSVs and DSS
========================================================
Consolidates every module's per-chunk ``_product_b_final/*.csv`` into ten
unified ``ProductB_SV_n{01..10}.csv`` files and writes the corresponding
``ProductB_SV_n{01..10}.dss`` files by overlaying chunk values onto the
CalSim baseline DSS template. Auto-fills inventory Constant/Rept SVs from
the baseline 12-month pattern and runs Product A and CalSim baseline
diagnostic comparisons.

Inputs
------
- Each module's GENERATED ``_product_b_final/*.csv`` (per the MODULE_CONFIG_B
  scan order below: calsimhydro, calsimhydro_ee, evaporation, rim_inflow,
  delta_channel_depletion, small_watersheds, storage_curves, instream_flows,
  tulare_gw_terms, climate, miscellaneous, upper_watershed, closure_terms,
  day_volume_fractions).
- ``BASE/CalSim3/__calsim_sv_default__.dss`` (baseline template, Constant/Rept
  12-month repeat source, units extractor).
- ``GENERATED/postprocessing/sv_compile/product_a_validation/
  ProductA_Historical_Validation_SV.dss`` (optional Product A comparison
  input; written by ``product_a_historical_validation.py``).
- ``inventory/_MASTER_INVENTORY_FOR_STOCHASTIC_INPUT_GENERATION_.xlsx``.

Outputs (all written to ``product_b_compilation/``)
---------------------------------------------------
- ``compiled_input_files/<module>/*.csv`` -- local copies of source CSVs
- ``_product_b_compiled_sv/ProductB_SV_n{01..10}.csv`` (+ matching .dss)
- ``inventory_*`` cross-reference CSVs (expected/missing/constant_rept/
  skipped_missing/skipped_not_in_dcr/unexpected)
- ``sv_coverage_by_chunk.csv``
- ``product_b_vs_a_comparison.csv``,
  ``product_b_vs_calsim_base_comparison.csv``
- ``compilation_summary.txt``
- ``figures/vs_product_a/`` and ``figures/vs_calsim_base/`` PNGs
  (weighted/unweighted annual %diff per category, exceedance rank shift,
  monthly climatology, per-category chunk spread)

Dependencies
------------
- ``utils.dss_io`` (open_dss, create_junction, remove_junction)
- ``utils.paths``
- pydsstools, numpy, pandas, matplotlib

Usage
-----
    python postprocessing/sv_compile/product_b_compilation.py
    python postprocessing/sv_compile/product_b_compilation.py --chunks 1 2 3
    python postprocessing/sv_compile/product_b_compilation.py --skip-comparison
    python postprocessing/sv_compile/product_b_compilation.py --skip-dss
    python postprocessing/sv_compile/product_b_compilation.py --summary-figures

CLI flags
---------
- ``--skip-comparison`` Skip the Product A vs B comparison step.
- ``--skip-dss``        Skip DSS file generation (CSV only).
- ``--chunks N [N ...]`` Process only specific chunks (default: all 10).
- ``--summary-figures`` Regenerate figures from a previous comparison CSV.

Note: this script runs procedurally at module load (no ``main()`` wrapper).
Wrapping the ~2700-line body in a function is deferred to avoid a
high-risk scoping reshuffle on a script with no tests; the file is never
imported elsewhere (only run as ``python <path>``).
"""

import sys
import time
import shutil
import argparse
import atexit
import warnings
import functools
import numpy as np
import pandas as pd
from pathlib import Path
from collections import OrderedDict

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from utils import dss_io
from utils.paths import get_base_dir, get_generated_dir, get_module_generated_dir, get_inventory_dir

warnings.filterwarnings("ignore", message="Mean of empty slice")
warnings.filterwarnings("ignore", message="All-NaN slice encountered")


# ======================================================================
# Configuration
# ======================================================================
_base = get_base_dir()
_gen  = get_module_generated_dir("postprocessing/sv_compile")

BASELINE_DSS   = _base / "CalSim3" / "__calsim_sv_default__.dss"
OUTPUT_DIR     = _gen / "product_b_compilation"
COMPILED_DIR   = OUTPUT_DIR / "_product_b_compiled_sv"
COMPILED_CSV   = OUTPUT_DIR / "compiled_input_files"

# Product A compiled DSS (for comparison)
PRODUCT_A_DSS  = _gen / "product_a_validation" / "ProductA_Historical_Validation_SV.dss"

INVENTORY_XLSX = get_inventory_dir() / "_MASTER_INVENTORY_FOR_STOCHASTIC_INPUT_GENERATION_.xlsx"

DSS_PATTERN = "/*/*/*/*/1MON/*"

# Product B canonical period: 100 WY per chunk  (Oct 1921 - Sep 2021)
N_CHUNKS     = 10
CHUNK_TAGS   = [f"n{i:02d}" for i in range(1, N_CHUNKS + 1)]
PB_START_YM  = (1921, 10)   # first data month
PB_END_YM    = (2021,  9)   # last data month

# Product A comparison window (WY 1972-2018)
PA_START = pd.Timestamp(1971, 10, 31)
PA_END   = pd.Timestamp(2018,  9, 30)

# CalSim baseline comparison window (full period, WY 1922-2021)
CB_START = pd.Timestamp(1921, 10, 31)
CB_END   = pd.Timestamp(2021,  9, 30)


# Junction lifecycle (mklink /J + Windows long-path handling) lives in
# utils.dss_io; this script uses dss_io.open_dss / create_junction /
# remove_junction.


# -- CLI arguments ---------------------------------------------------------
_parser = argparse.ArgumentParser(
    description="Product B (1000-year stochastic) -- CSV Compilation",
)
_parser.add_argument(
    "--skip-comparison", action="store_true", default=False,
    help="Skip the Product A vs B comparison step.",
)
_parser.add_argument(
    "--skip-dss", action="store_true", default=False,
    help="Skip DSS file generation (CSV only).",
)
_parser.add_argument(
    "--chunks", nargs="+", type=int,
    default=list(range(1, N_CHUNKS + 1)),
    help="Chunk numbers to compile, 1-10 (default: all).",
)
_parser.add_argument(
    "--summary-figures", action="store_true", default=False,
    help="Regenerate only the summary figures from a previous comparison CSV. "
         "Skips Steps 1-6, 8, and monthly climatology plots.",
)
CLI_ARGS = _parser.parse_args()

# Validate chunk numbers
for _c in CLI_ARGS.chunks:
    if _c < 1 or _c > N_CHUNKS:
        sys.exit(f"ERROR: Chunk number {_c} out of range (1-{N_CHUNKS})")

ACTIVE_CHUNKS = sorted(CLI_ARGS.chunks)
ACTIVE_TAGS   = [f"n{c:02d}" for c in ACTIVE_CHUNKS]


# ======================================================================
# Module definitions
# ======================================================================
_gen_dir = get_generated_dir()

# label -> (path to _product_b_final dir, inventory Input_Category, reader_type)
# reader_type: "standard" = Part B, Part C, Year, Month, Value
#              "rim_inflow" = CalSim, Matched_inflow, Year, Month, ..., qmap_postAdj
MODULE_CONFIG_B = OrderedDict([
    ("calsimhydro", (
        _gen_dir / "mod_hydrology/calsimhydro/output/_4_postprocess_product_b/_product_b_final",
        "CalSimHydro", "standard")),
    ("calsimhydro_ee", (
        _gen_dir / "mod_hydrology/calsimhydro_ee/output/_3_postprocess_product_b/_product_b_final",
        "CalSimHydroEE", "standard")),
    ("evaporation", (
        _gen_dir / "mod_reservoir/evaporation/output/_2_run_reservoir_evap/_product_b_final",
        "Reservoir Evaporation", "standard")),
    ("rim_inflow", (
        _gen_dir / "mod_hydrology/rim_inflow/output/_3_qmap_product_b",
        "Rim Inflow", "rim_inflow")),
    ("delta_channel_depletion", (
        _gen_dir / "mod_hydrology/delta_channel_depletion/output/_3_postprocess_product_b/_product_b_final",
        "Delta Channel Depletion", "standard")),
    ("small_watersheds", (
        _gen_dir / "mod_hydrology/small_watersheds/output/_3_postprocess_product_b/_product_b_final",
        "Small Watersheds", "standard")),
    ("storage_curves", (
        _gen_dir / "mod_reservoir/storage_curves/output/_product_b_final",
        "Reservoir Storage Curves", "standard")),
    ("instream_flows", (
        _gen_dir / "mod_other/instream_flows/output/_product_b_final",
        "Instream Flows", "standard")),
    ("tulare_gw_terms", (
        _gen_dir / "mod_hydrology/tulare_gw_terms/output/_1_wyt_monthlyavg/_product_b_final",
        "Tulare Groundwater Terms", "standard")),
    ("climate", (
        _gen_dir / "mod_forcing/climate/output/_product_b_final",
        "Climate", "standard")),
    ("miscellaneous", (
        _gen_dir / "mod_other/miscellaneous/output/_product_b_final",
        "Other", "standard")),
    ("upper_watershed", (
        _gen_dir / "mod_other/upper_watershed/output/_product_b_final",
        "Upper Watershed Modules", "standard")),
    ("closure_terms", (
        _gen_dir / "mod_other/closure_terms/output/_product_b_final",
        "Closure Terms", "standard")),
    ("day_volume_fractions", (
        _gen_dir / "mod_other/day_volume_fractions/output/_product_b_final",
        "Day-Volume Fraction", "standard")),
])

# Inventory categories that do NOT have Product B modules
CATEGORIES_WITHOUT_PRODUCT_B = {"Salinity"}


# ======================================================================
# Helpers
# ======================================================================
def excel_to_part(name: str) -> str:
    return str(name).upper().replace(" ", "_")


def safe_write_ts(dss_out, pathname: str, ts_obj):
    """Write a TimeSeriesContainer, tolerating API differences."""
    ts_obj.pathname = pathname
    if hasattr(dss_out, "put_ts"):
        dss_out.put_ts(ts_obj)
    elif hasattr(dss_out, "write_ts"):
        dss_out.write_ts(ts_obj)


def ym_to_eom(year: int, month: int) -> pd.Timestamp:
    return pd.Timestamp(year=int(year), month=int(month), day=1).to_period("M").to_timestamp("M")


def dss_eom(ts_pytimes) -> pd.DatetimeIndex:
    return (pd.to_datetime(ts_pytimes).to_period("M") - 1).to_timestamp("M")


def path_key(pathname: str) -> tuple:
    parts = pathname.strip("/").split("/")
    return (parts[1].upper(), parts[2].upper())


@functools.lru_cache(maxsize=1)
def read_master_inventory() -> pd.DataFrame:
    """Read the MASTER sheet from the inventory workbook."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(str(INVENTORY_XLSX), read_only=True, data_only=True)
    except PermissionError:
        sys.exit(
            f"ERROR: Cannot read the master inventory -- the file may be open in Excel.\n"
            f"  Close the file and re-run the script:\n"
            f"  {INVENTORY_XLSX}"
        )
    ws = wb["MASTER"]
    rows = list(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True))
    wb.close()
    records = []
    for r in rows:
        b, c, cat = r[2], r[3], r[8]
        if b and c and cat:
            records.append({
                "Part_B":          str(b).upper().replace(" ", "_"),
                "Part_C":          str(c).upper().replace(" ", "_"),
                "Input_Category":  str(cat).strip(),
                "Missing":         str(r[10]).strip().upper() == "T",
                "Constant_Rept":   str(r[11]).strip().upper() == "T",
                "Used_in_DCR":     str(r[12]).strip().upper() != "F",
            })
    return pd.DataFrame(records)


@functools.lru_cache(maxsize=1)
def _inventory_partc_map() -> dict:
    """Build Part_B -> Part_C lookup from the inventory (for rim_inflow normalization)."""
    inv = read_master_inventory()
    m = {}
    for _, r in inv.iterrows():
        m[r["Part_B"]] = r["Part_C"]
    return m


# ======================================================================
# Step 0: Detect chunk tag from filename
# ======================================================================
def detect_chunk_tag(filename: str) -> str:
    """Extract chunk tag (n01-n10) from a CSV filename.

    Tries several patterns used across modules.
    Returns empty string if no chunk tag found.
    """
    import re
    name = filename.lower()
    # Match _n01, _n02 .. _n10
    m = re.search(r'_n(\d{2})(?:\.|_|$)', name)
    if m:
        num = int(m.group(1))
        if 1 <= num <= N_CHUNKS:
            return f"n{num:02d}"
    return ""


# ======================================================================
# Step 1: Read & normalize module CSVs
# ======================================================================
def read_standard_csv(csv_path: Path) -> pd.DataFrame:
    """Read a standard SV CSV (Part B, Part C, Year, Month, Value)."""
    _p = str(csv_path)
    if sys.platform == "win32" and len(_p) >= 260 and not _p.startswith("\\\\?\\"):
        _p = "\\\\?\\" + _p
    df = pd.read_csv(_p)
    df.columns = [c.strip() for c in df.columns]
    required = {"Part B", "Part C", "Year", "Month", "Value"}
    if not required.issubset(set(df.columns)):
        missing = required - set(df.columns)
        print(f"    WARNING: {csv_path.name} missing columns {sorted(missing)} -- skipping")
        return pd.DataFrame(columns=["Part B", "Part C", "Year", "Month", "Value"])
    df["Part B"] = df["Part B"].apply(excel_to_part)
    df["Part C"] = df["Part C"].apply(excel_to_part)
    df["Year"]   = df["Year"].astype(int)
    df["Month"]  = df["Month"].astype(int)
    df["Value"]  = pd.to_numeric(df["Value"], errors="coerce")
    return df[["Part B", "Part C", "Year", "Month", "Value"]].dropna(subset=["Value"])


def read_rim_inflow_csv(csv_path: Path) -> pd.DataFrame:
    """Read a rim_inflow non-standard CSV and normalise to standard format.

    Input columns:  CalSim, Matched_inflow, Year, Month, vic_val, qmap_preAdj, qmap_postAdj
    Output columns: Part B, Part C, Year, Month, Value
    """
    _p = str(csv_path)
    if sys.platform == "win32" and len(_p) >= 260 and not _p.startswith("\\\\?\\"):
        _p = "\\\\?\\" + _p
    df = pd.read_csv(_p)
    df.columns = [c.strip() for c in df.columns]
    if "CalSim" not in df.columns or "qmap_postAdj" not in df.columns:
        # Try case-insensitive lookup
        col_map = {c.lower(): c for c in df.columns}
        if "calsim" not in col_map or "qmap_postadj" not in col_map:
            print(f"    WARNING: {csv_path.name} missing CalSim/qmap_postAdj columns -- skipping")
            return pd.DataFrame(columns=["Part B", "Part C", "Year", "Month", "Value"])
        calsim_col = col_map["calsim"]
        value_col  = col_map["qmap_postadj"]
    else:
        calsim_col = "CalSim"
        value_col  = "qmap_postAdj"

    partc_map = _inventory_partc_map()

    out = pd.DataFrame()
    out["Part B"] = df[calsim_col].apply(excel_to_part)
    out["Part C"] = out["Part B"].map(partc_map).fillna("FLOW-INFLOW")
    out["Year"]   = df["Year"].astype(int)
    out["Month"]  = df["Month"].astype(int)
    out["Value"]  = pd.to_numeric(df[value_col], errors="coerce")
    return out[["Part B", "Part C", "Year", "Month", "Value"]].dropna(subset=["Value"])


def scan_module_csvs(label, src_dir, reader_type):
    """Scan a module directory for CSVs, group by chunk tag.

    Returns:
        chunk_data: dict  chunk_tag -> list[pd.DataFrame]
        n_files: int
        n_svs: set of (Part B, Part C)
        source_paths: list[Path]  -- original CSV file paths (for copying)
    """
    chunk_data = {tag: [] for tag in CHUNK_TAGS}
    n_files = 0
    all_svs = set()
    source_paths = []

    if not src_dir.is_dir():
        return chunk_data, 0, set(), []

    csv_files = sorted(src_dir.glob("*.csv"))
    reader_fn = read_rim_inflow_csv if reader_type == "rim_inflow" else read_standard_csv

    for csv_path in csv_files:
        tag = detect_chunk_tag(csv_path.name)
        if not tag:
            continue  # not a chunk CSV (could be a summary or report file)
        if tag not in chunk_data:
            continue

        df = reader_fn(csv_path)
        if df.empty:
            continue

        chunk_data[tag].append(df)
        n_files += 1
        source_paths.append(csv_path)
        for b, c in zip(df["Part B"], df["Part C"]):
            all_svs.add((b, c))

    return chunk_data, n_files, all_svs, source_paths


# ======================================================================
# Step 2: Constant/Rept auto-fill from baseline
# ======================================================================
def _extract_wy_pattern(ser_clean):
    """Extract a 12-month repeating pattern from non-missing baseline data."""
    WY_MONTHS = [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]
    wy = ser_clean.index.year + (ser_clean.index.month >= 10).astype(int)
    for water_year in sorted(wy.unique(), reverse=True):
        wy_slice = ser_clean[wy == water_year]
        months_present = set(wy_slice.index.month)
        if months_present == set(WY_MONTHS):
            pattern = {}
            for dt, val in zip(wy_slice.index, wy_slice.values):
                pattern[dt.month] = val
            return pattern, True
    pattern = {}
    for m in WY_MONTHS:
        month_vals = ser_clean[ser_clean.index.month == m]
        if not month_vals.empty:
            pattern[m] = month_vals.iloc[-1]
    return pattern, len(pattern) == 12


def build_constant_rept_rows(dss_in, baseline_bucket, part_key):
    """Build 100 WY of constant/rept values from the baseline pattern.

    Returns a DataFrame with Part B, Part C, Year, Month, Value
    covering Oct 1921 - Sep 2021 (1200 months).
    """
    if part_key not in baseline_bucket:
        return pd.DataFrame()

    # Merge all D-part blocks into one unified series
    merged = {}
    for pathname in baseline_bucket[part_key]:
        ts = dss_in.read_ts(pathname, trim_missing=False)
        eom = dss_eom(ts.pytimes)
        eom_pd = pd.DatetimeIndex(eom)
        ts_vals = np.array(ts.values, dtype=float)
        for i, dt in enumerate(eom_pd):
            if ts_vals[i] > -900:
                merged[dt] = ts_vals[i]

    if not merged:
        return pd.DataFrame()

    merged_dates = sorted(merged.keys())
    merged_ser = pd.Series(
        [merged[d] for d in merged_dates],
        index=pd.DatetimeIndex(merged_dates),
    )
    pattern, complete = _extract_wy_pattern(merged_ser)

    if not complete:
        missing_months = sorted(set(range(1, 13)) - set(pattern.keys()))
        month_names = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
                       7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
        missing_str = ", ".join(month_names[m] for m in missing_months)
        print(f"  WARNING: Constant/Rept pattern for {part_key} is missing "
              f"months: {missing_str}")

    # Tile pattern across the Product B window
    rows = []
    partb, partc = part_key
    start = pd.Timestamp(PB_START_YM[0], PB_START_YM[1], 1)
    end = pd.Timestamp(PB_END_YM[0], PB_END_YM[1], 1)
    months = pd.date_range(start, end, freq="MS")
    for dt in months:
        m = dt.month
        if m in pattern:
            rows.append({
                "Part B": partb,
                "Part C": partc,
                "Year": dt.year,
                "Month": m,
                "Value": pattern[m],
            })

    return pd.DataFrame(rows)


# ======================================================================
# Step 3: Read Product A DSS for comparison
# ======================================================================
def read_product_a_monthly_means(product_a_dss: Path, baseline_bucket: dict,
                                 keys_to_read: set) -> dict:
    """Read Product A compiled DSS and compute monthly means per (Part B, Part C).

    Returns dict: (Part_B, Part_C) -> {month: mean_value}
    """
    result = {}
    with dss_io.open_dss(product_a_dss, version=6, catalog_flag=True) as dss:
        pa_paths = dss.getPathnameList(DSS_PATTERN)
        pa_bucket = {}
        for p in pa_paths:
            k = path_key(p)
            pa_bucket.setdefault(k, []).append(p)

        for pk in sorted(keys_to_read):
            if pk not in pa_bucket:
                continue

            merged = {}
            for pathname in pa_bucket[pk]:
                try:
                    ts = dss.read_ts(pathname, trim_missing=False)
                except Exception:
                    continue
                eom = dss_eom(ts.pytimes)
                vals = np.array(ts.values, dtype=float)
                for i, dt in enumerate(eom):
                    if vals[i] > -900 and PA_START <= dt <= PA_END:
                        merged[dt] = vals[i]

            if not merged:
                continue

            ser = pd.Series(merged)
            monthly_means = {}
            for m in range(1, 13):
                mm = ser[ser.index.month == m]
                if not mm.empty:
                    monthly_means[m] = mm.mean()
            result[pk] = monthly_means

    return result


def read_product_a_monthly_series(product_a_dss: Path,
                                  keys_to_read: set) -> dict:
    """Read Product A compiled DSS monthly series per (Part B, Part C)."""
    result = {}
    with dss_io.open_dss(product_a_dss, version=6, catalog_flag=True) as dss:
        pa_paths = dss.getPathnameList(DSS_PATTERN)
        pa_bucket = {}
        for p in pa_paths:
            k = path_key(p)
            pa_bucket.setdefault(k, []).append(p)

        for pk in sorted(keys_to_read):
            if pk not in pa_bucket:
                continue
            merged = {}
            for pathname in pa_bucket[pk]:
                try:
                    ts = dss.read_ts(pathname, trim_missing=False)
                except Exception:
                    continue
                eom = dss_eom(ts.pytimes)
                vals = np.array(ts.values, dtype=float)
                for i, dt in enumerate(eom):
                    if vals[i] > -900 and PA_START <= dt <= PA_END:
                        merged[pd.Timestamp(dt)] = vals[i]
            if merged:
                result[pk] = pd.Series(merged).sort_index()

    return result


def read_calsim_base_monthly_means(baseline_bucket: dict,
                                   keys_to_read: set,
                                   start: pd.Timestamp,
                                   end: pd.Timestamp) -> dict:
    """Read CalSim baseline DSS monthly means per (Part B, Part C).

    Returns dict: (Part_B, Part_C) -> {month: mean_value}
    """
    result = {}
    with dss_io.open_dss(BASELINE_DSS, version=6, catalog_flag=False) as dss:
        for pk in sorted(keys_to_read):
            if pk not in baseline_bucket:
                continue
            merged = {}
            for pathname in baseline_bucket[pk]:
                try:
                    ts = dss.read_ts(pathname, trim_missing=False)
                except Exception:
                    continue
                eom = dss_eom(ts.pytimes)
                vals = np.array(ts.values, dtype=float)
                for i, dt in enumerate(eom):
                    if vals[i] > -900 and start <= dt <= end:
                        merged[dt] = vals[i]
            if not merged:
                continue
            ser = pd.Series(merged)
            monthly_means = {}
            for m in range(1, 13):
                mm = ser[ser.index.month == m]
                if not mm.empty:
                    monthly_means[m] = mm.mean()
            result[pk] = monthly_means
    return result


def read_calsim_base_monthly_series(baseline_bucket: dict,
                                    keys_to_read: set,
                                    start: pd.Timestamp,
                                    end: pd.Timestamp) -> dict:
    """Read CalSim baseline DSS monthly series per (Part B, Part C)."""
    result = {}
    with dss_io.open_dss(BASELINE_DSS, version=6, catalog_flag=False) as dss:
        for pk in sorted(keys_to_read):
            if pk not in baseline_bucket:
                continue
            merged = {}
            for pathname in baseline_bucket[pk]:
                try:
                    ts = dss.read_ts(pathname, trim_missing=False)
                except Exception:
                    continue
                eom = dss_eom(ts.pytimes)
                vals = np.array(ts.values, dtype=float)
                for i, dt in enumerate(eom):
                    if vals[i] > -900 and start <= dt <= end:
                        merged[pd.Timestamp(dt)] = vals[i]
            if merged:
                result[pk] = pd.Series(merged).sort_index()
    return result


def _load_compiled_chunks_from_csv(active_tags):
    """Reload compiled Product B chunk CSVs for --summary-figures mode."""
    compiled_chunks = {}
    required = {"Part B", "Part C", "Year", "Month", "Value"}
    for tag in active_tags:
        csv_path = COMPILED_DIR / f"ProductB_SV_{tag}.csv"
        if not csv_path.exists():
            continue
        df = pd.read_csv(csv_path)
        df.columns = [c.strip() for c in df.columns]
        if not required.issubset(df.columns):
            continue
        df["Part B"] = df["Part B"].apply(excel_to_part)
        df["Part C"] = df["Part C"].apply(excel_to_part)
        df["Year"] = pd.to_numeric(df["Year"], errors="coerce").astype("Int64")
        df["Month"] = pd.to_numeric(df["Month"], errors="coerce").astype("Int64")
        df["Value"] = pd.to_numeric(df["Value"], errors="coerce")
        df = df.dropna(subset=["Year", "Month", "Value"]).copy()
        df["Year"] = df["Year"].astype(int)
        df["Month"] = df["Month"].astype(int)
        compiled_chunks[tag] = df[["Part B", "Part C", "Year", "Month", "Value"]]
    return compiled_chunks


def _compute_pb_chunk_means(compiled_chunks, active_tags, start_ym=None, end_ym=None):
    """Compute Product B monthly means per chunk, with optional year-month filter.

    Parameters
    ----------
    start_ym, end_ym : int or None
        Year-month as YYYYMM (e.g. 197110 for Oct 1971).  None = no filter.

    Returns dict: tag -> {(Part_B, Part_C): {month: mean_value}}
    """
    pb_chunk_means = {}
    for tag in active_tags:
        if tag not in compiled_chunks:
            pb_chunk_means[tag] = {}
            continue
        df_chunk = compiled_chunks[tag]
        if start_ym is not None and end_ym is not None:
            ym = df_chunk["Year"] * 100 + df_chunk["Month"]
            df_chunk = df_chunk[(ym >= start_ym) & (ym <= end_ym)]
        tag_means = {}
        for (b, c), grp in df_chunk.groupby(["Part B", "Part C"]):
            mm = {}
            for m_val, mg in grp.groupby("Month"):
                mm[int(m_val)] = mg["Value"].mean()
            tag_means[(b, c)] = mm
        pb_chunk_means[tag] = tag_means
    return pb_chunk_means


def _build_comparison_df(ref_means, pb_chunk_means, all_compiled_svs, active_tags):
    """Build comparison DataFrame: reference monthly means vs PB chunk means.

    Returns DataFrame with columns: Part_B, Part_C, Month, Ref_mean,
    {tag}_mean, chunk_median, pct_diff, Input_Category
    """
    cmp_rows = []
    for pk in sorted(all_compiled_svs):
        if pk not in ref_means:
            continue
        ref_mm = ref_means[pk]
        for m in range(1, 13):
            row = {
                "Part_B": pk[0],
                "Part_C": pk[1],
                "Month": m,
                "Ref_mean": ref_mm.get(m, np.nan),
            }
            chunk_means_for_month = []
            for tag in active_tags:
                val = pb_chunk_means.get(tag, {}).get(pk, {}).get(m, np.nan)
                row[f"{tag}_mean"] = val
                if np.isfinite(val):
                    chunk_means_for_month.append(val)
            if chunk_means_for_month:
                row["chunk_median"] = np.median(chunk_means_for_month)
                ref_val = ref_mm.get(m, np.nan)
                if np.isfinite(ref_val) and abs(ref_val) > 1e-6:
                    row["pct_diff"] = (
                        (row["chunk_median"] - ref_val) / abs(ref_val)
                    ) * 100.0
                else:
                    row["pct_diff"] = np.nan
            else:
                row["chunk_median"] = np.nan
                row["pct_diff"] = np.nan
            cmp_rows.append(row)

    cmp_df = pd.DataFrame(cmp_rows)
    inv = read_master_inventory()
    cat_map = {}
    for _, r in inv.iterrows():
        cat_map[(r["Part_B"], r["Part_C"])] = r["Input_Category"]
    cmp_df["Input_Category"] = cmp_df.apply(
        lambda r: cat_map.get((r["Part_B"], r["Part_C"]), "Unknown"), axis=1
    )
    return cmp_df


def _generate_comparison_figures(cmp_df, fig_dir, ref_label, active_tags,
                                units_map, skip_climatology=False,
                                ref_series_by_pk=None,
                                compiled_chunks=None):
    """Generate summary, chunk_spread, and monthly_climatology figures."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.gridspec import GridSpec

    _FS = 7
    plt.rcParams.update({
        "font.size": _FS, "axes.titlesize": _FS,
        "axes.labelsize": _FS, "xtick.labelsize": _FS,
        "ytick.labelsize": _FS, "legend.fontsize": _FS,
        "figure.dpi": 300, "savefig.dpi": 300,
        "figure.facecolor": "white",
    })

    _MONTH_LABELS = ["Oct", "Nov", "Dec", "Jan", "Feb", "Mar",
                     "Apr", "May", "Jun", "Jul", "Aug", "Sep"]
    _SUM_UNITS = {"TAF", "IN", "INCHES"}
    _CATS_GROUPED_BY_MODULE = [
        "CalSimHydro", "CalSimHydroEE", "Rim Inflow",
        "Delta Channel Depletion", "Small Watersheds",
        "Tulare Groundwater Terms",
        "Reservoir Evaporation", "Reservoir Storage Curves",
        "Climate",
        "Closure Terms", "Day-Volume Fraction", "Salinity",
        "Instream Flows", "Other", "Upper Watershed Modules",
    ]
    _CAT_TO_MODULE_GROUP = {
        "CalSimHydro":              "mod_hydrology",
        "CalSimHydroEE":            "mod_hydrology",
        "Rim Inflow":               "mod_hydrology",
        "Delta Channel Depletion":  "mod_hydrology",
        "Small Watersheds":         "mod_hydrology",
        "Tulare Groundwater Terms": "mod_hydrology",
        "Reservoir Evaporation":    "mod_reservoir",
        "Reservoir Storage Curves": "mod_reservoir",
        "Climate":                  "mod_forcing",
        "Closure Terms":            "mod_other",
        "Day-Volume Fraction":      "mod_other",
        "Salinity":                 "mod_other",
        "Instream Flows":           "mod_other",
        "Other":                    "mod_other",
        "Upper Watershed Modules":  "mod_other",
    }
    _LABEL_BREAKS = {
        "Delta Channel Depletion":  "Delta Channel\nDepletion",
        "Tulare Groundwater Terms": "Tulare Groundwater\nTerms",
        "Reservoir Evaporation":    "Reservoir\nEvaporation",
        "Reservoir Storage Curves": "Reservoir\nStorage Curves",
        "Upper Watershed Modules":  "Upper Watershed\nModules",
        "Instream Flows":           "Instream\nFlows",
        "Small Watersheds":         "Small\nWatersheds",
    }

    def _ordered_categories(categories):
        present = set(categories)
        ordered = [c for c in _CATS_GROUPED_BY_MODULE if c in present]
        ordered += sorted(present - set(ordered))
        return ordered

    def _add_module_group_brackets(ax, fig, categories):
        groups = []
        prev_group = None
        for i, cat in enumerate(categories):
            group_name = _CAT_TO_MODULE_GROUP.get(cat, "")
            if group_name == prev_group and groups:
                groups[-1]["end"] = i
            else:
                groups.append({"name": group_name, "start": i, "end": i})
                prev_group = group_name

        fig.canvas.draw()
        trans = ax.get_xaxis_transform()
        bracket_y = -0.48
        tick_y = -0.44
        label_y = -0.54

        for group in groups:
            if not group["name"]:
                continue
            x_start = group["start"] + 1
            x_end = group["end"] + 1
            x_mid = (x_start + x_end) / 2.0
            ax.plot([x_start - 0.3, x_end + 0.3], [bracket_y, bracket_y],
                    transform=trans, color="0.3", lw=0.8, clip_on=False)
            ax.plot([x_start - 0.3, x_start - 0.3], [tick_y, bracket_y],
                    transform=trans, color="0.3", lw=0.8, clip_on=False)
            ax.plot([x_end + 0.3, x_end + 0.3], [tick_y, bracket_y],
                    transform=trans, color="0.3", lw=0.8, clip_on=False)
            ax.text(x_mid, label_y, group["name"],
                    transform=trans, ha="center", va="top",
                    fontsize=6, fontstyle="italic", color="0.3",
                    clip_on=False)

    def _weighted_quantile(values, weights, quantiles):
        values = np.asarray(values, dtype=float)
        weights = np.asarray(weights, dtype=float)
        quantiles = np.asarray(quantiles, dtype=float)
        mask = np.isfinite(values) & np.isfinite(weights) & (weights > 0)
        values = values[mask]
        weights = weights[mask]
        if len(values) == 0:
            return np.full(len(quantiles), np.nan)
        sorter = np.argsort(values)
        values = values[sorter]
        weights = weights[sorter]
        cumulative = np.cumsum(weights) - 0.5 * weights
        cumulative /= np.sum(weights)
        return np.interp(quantiles, cumulative, values,
                         left=values[0], right=values[-1])

    def _weighted_box_stats(values, weights, label):
        values = np.asarray(values, dtype=float)
        weights = np.asarray(weights, dtype=float)
        mask = np.isfinite(values) & np.isfinite(weights) & (weights > 0)
        values = values[mask]
        weights = weights[mask]
        if len(values) == 0:
            return None
        q1, med, q3 = _weighted_quantile(values, weights, [0.25, 0.50, 0.75])
        iqr = q3 - q1
        if np.isfinite(iqr) and iqr > 0:
            lower_fence = q1 - 1.5 * iqr
            upper_fence = q3 + 1.5 * iqr
            inside = values[(values >= lower_fence) & (values <= upper_fence)]
            outside = values[(values < lower_fence) | (values > upper_fence)]
            if len(inside) > 0:
                whislo = float(np.nanmin(inside))
                whishi = float(np.nanmax(inside))
            else:
                whislo = float(q1)
                whishi = float(q3)
        else:
            outside = np.array([])
            whislo = float(q1)
            whishi = float(q3)
        return {
            "label": label,
            "whislo": whislo,
            "q1": float(q1),
            "med": float(med),
            "q3": float(q3),
            "whishi": whishi,
            "fliers": outside.tolist(),
        }

    fig_dir.mkdir(parents=True, exist_ok=True)
    cats_present = sorted(cmp_df["Input_Category"].dropna().unique())

    inv = read_master_inventory()
    const_rept_map = {
        (r["Part_B"], r["Part_C"]): bool(r["Constant_Rept"])
        for _, r in inv.iterrows()
    }

    # -- Aggregate to mean-annual values --
    annual_rows = []
    for (b, c), grp in cmp_df.groupby(["Part_B", "Part_C"]):
        unit = units_map.get((b, c), "UNKNOWN")
        cat = grp["Input_Category"].iloc[0]
        use_sum = unit.upper() in _SUM_UNITS
        ref_vals = grp["Ref_mean"].values
        ref_ann = float(np.nansum(ref_vals) if use_sum
                        else np.nanmean(ref_vals))
        row = {"Part_B": b, "Part_C": c, "Units": unit,
               "Input_Category": cat, "Ref_annual": ref_ann,
               "Constant_Rept": const_rept_map.get((b, c), False)}
        for tag in active_tags:
            col = f"{tag}_mean"
            if col not in grp.columns:
                continue
            ch_vals = grp[col].values
            ch_ann = float(np.nansum(ch_vals) if use_sum
                           else np.nanmean(ch_vals))
            row[f"{tag}_annual"] = ch_ann
            if (np.isfinite(ref_ann) and abs(ref_ann) > 1e-6
                    and np.isfinite(ch_ann)):
                row[f"{tag}_pct_diff"] = (
                    (ch_ann - ref_ann) / abs(ref_ann) * 100.0
                )
                row[f"{tag}_abs_diff"] = ch_ann - ref_ann
            else:
                row[f"{tag}_pct_diff"] = np.nan
                row[f"{tag}_abs_diff"] = np.nan
        annual_rows.append(row)

    annual_df = pd.DataFrame(annual_rows)

    # Exclude CalSimHydro terms that are identical in historical and
    # stochastic (constant repeating or water-demand/wastewater policy
    # inputs that are unchanged by design).
    _CALSIMHYDRO_PARTC_EXCL = {"URBAN-DEMAND", "WASTEWATER"}
    _excl_mask = (
        (annual_df["Input_Category"] == "CalSimHydro")
        & (annual_df["Part_C"].isin(_CALSIMHYDRO_PARTC_EXCL))
    )
    annual_df = annual_df[~_excl_mask].copy()

    # -- Weighted all-category summary ---------------------------------
    summary_df = annual_df[~annual_df["Constant_Rept"]].copy()
    summary_rows = []
    if not summary_df.empty:
        for cat, cat_df in summary_df.groupby("Input_Category"):
            ref_abs = cat_df["Ref_annual"].abs().replace([np.inf, -np.inf], np.nan)
            valid_ref = ref_abs.notna() & (ref_abs > 1e-9)
            if valid_ref.any():
                base_weights = ref_abs.where(valid_ref, np.nan)
            else:
                base_weights = pd.Series(1.0, index=cat_df.index)

            for idx, row in cat_df.iterrows():
                pct_cols = [f"{tag}_pct_diff" for tag in active_tags
                            if f"{tag}_pct_diff" in cat_df.columns
                            and np.isfinite(row.get(f"{tag}_pct_diff", np.nan))]
                if not pct_cols:
                    continue
                term_weight = base_weights.loc[idx]
                if not np.isfinite(term_weight) or term_weight <= 0:
                    term_weight = 0.0
                obs_weight = term_weight / len(pct_cols) if pct_cols else np.nan
                for tag in active_tags:
                    pct_col = f"{tag}_pct_diff"
                    ann_col = f"{tag}_annual"
                    abs_col = f"{tag}_abs_diff"
                    pct_val = row.get(pct_col, np.nan)
                    if not np.isfinite(pct_val):
                        continue
                    summary_rows.append({
                        "Input_Category": cat,
                        "Part_B": row["Part_B"],
                        "Part_C": row["Part_C"],
                        "Units": row["Units"],
                        "Chunk": tag,
                        "Ref_annual": row["Ref_annual"],
                        "Chunk_annual": row.get(ann_col, np.nan),
                        "Pct_Diff": pct_val,
                        "Abs_Diff": row.get(abs_col, np.nan),
                        "Weight_Base": term_weight,
                        "Observation_Weight": obs_weight,
                    })

    weighted_summary_df = pd.DataFrame(summary_rows)
    if not weighted_summary_df.empty:
        totals = weighted_summary_df.groupby("Input_Category")["Observation_Weight"].transform("sum")
        weighted_summary_df["Weight_Share"] = np.where(
            totals > 0,
            weighted_summary_df["Observation_Weight"] / totals,
            np.nan,
        )
        weighted_summary_df.to_csv(
            fig_dir / "weighted_annual_pctdiff_by_category.csv", index=False
        )

        plot_categories = _ordered_categories(weighted_summary_df["Input_Category"].unique())
        term_counts = weighted_summary_df.groupby("Input_Category")[["Part_B", "Part_C"]].apply(
            lambda x: len(set(zip(x["Part_B"], x["Part_C"])))
        )
        stats = []
        box_labels = []
        plot_categories_used = []
        for cat in plot_categories:
            sub = weighted_summary_df[weighted_summary_df["Input_Category"] == cat]
            stat = _weighted_box_stats(
                sub["Pct_Diff"].values,
                sub["Observation_Weight"].values,
                cat,
            )
            if stat is None:
                continue
            stats.append(stat)
            box_labels.append(
                f"{_LABEL_BREAKS.get(cat, cat)}\n(n={int(term_counts.get(cat, 0))})"
            )
            plot_categories_used.append(cat)

        def _save_category_boxplot(bx_stats, bx_labels, bx_cats, fname, title,
                                   ylim=None, showfliers=False, symlog=False,
                                   linthresh=5):
            fig, ax = plt.subplots(
                figsize=(min(7, 0.65 * len(bx_cats) + 2), 5.0)
            )
            bp = ax.bxp(bx_stats, patch_artist=True, widths=0.6,
                        showfliers=showfliers)
            for patch in bp["boxes"]:
                patch.set_facecolor("#5B9BD5")
                patch.set_alpha(0.7)
            for cap in bp["caps"]:
                cap.set_visible(False)
            if showfliers:
                for flier in bp["fliers"]:
                    flier.set(marker="o", markersize=1,
                              markerfacecolor="k", markeredgecolor="none")
            ax.set_xticklabels(bx_labels, rotation=90, ha="center")
            ax.axhline(0.0, color="red", ls="--", lw=0.6, alpha=0.5)
            if symlog:
                ax.set_yscale("symlog", linthresh=linthresh)
                ax.yaxis.set_major_formatter(
                    plt.FuncFormatter(lambda v, _: f"{v:g}")
                )
                ax.yaxis.set_minor_locator(
                    plt.matplotlib.ticker.AutoMinorLocator()
                )
                ax.grid(which="minor", axis="y", color="#cccccc",
                        linestyle="--", linewidth=0.4, alpha=0.7)
            if ylim is not None:
                ax.set_ylim(ylim)
            ax.set_ylabel(f"Average Annual % Diff (vs {ref_label})")
            ax.set_title(title)
            _add_module_group_brackets(ax, fig, bx_cats)
            fig.subplots_adjust(bottom=0.40)
            fig.savefig(fig_dir / fname, bbox_inches="tight")
            plt.close(fig)
            print(f"    Figure: {fig_dir.name}/{fname}")

        if stats:
            _base_title = (
                f"Product B Chunk Average Annual % Difference by Input Category\n"
                f"weighted by abs({ref_label} annual average), excl. Constant/Rept"
            )
            _save_category_boxplot(
                stats, box_labels, plot_categories_used,
                "weighted_annual_pctdiff_by_category.png",
                _base_title,
            )
            _save_category_boxplot(
                stats, box_labels, plot_categories_used,
                "weighted_annual_pctdiff_by_category_clipped.png",
                _base_title + "  [y: \u221240 to +40%]",
                ylim=(-40, 40), showfliers=False,
            )

        # -- Unweighted version (uniform weights) ----------------------
        unweighted_stats = []
        unweighted_labels = []
        unweighted_cats_used = []
        for cat in plot_categories:
            sub = weighted_summary_df[weighted_summary_df["Input_Category"] == cat]
            uw_stat = _weighted_box_stats(
                sub["Pct_Diff"].values,
                np.ones(len(sub)),
                cat,
            )
            if uw_stat is None:
                continue
            unweighted_stats.append(uw_stat)
            unweighted_labels.append(
                f"{_LABEL_BREAKS.get(cat, cat)}\n(n={int(term_counts.get(cat, 0))})"
            )
            unweighted_cats_used.append(cat)

        if unweighted_stats:
            _uw_title = (
                "Product B Chunk Average Annual % Difference by Input Category\n"
                "unweighted, excl. Constant/Rept"
            )
            _save_category_boxplot(
                unweighted_stats, unweighted_labels, unweighted_cats_used,
                "unweighted_annual_pctdiff_by_category.png",
                _uw_title,
            )
            _save_category_boxplot(
                unweighted_stats, unweighted_labels, unweighted_cats_used,
                "unweighted_annual_pctdiff_by_category_clipped.png",
                _uw_title + "  [y: \u221240 to +40%]",
                ylim=(-40, 40), showfliers=False,
            )
            _save_category_boxplot(
                unweighted_stats, unweighted_labels, unweighted_cats_used,
                "unweighted_annual_pctdiff_by_category_symlog.png",
                _uw_title + "  [symlog scale]",
                showfliers=False, symlog=True,
            )

    # -- Per-category WY exceedance distribution diagnostics -----------
    def _safe_fig_name(name):
        safe = "".join(ch.lower() if ch.isalnum() else "_" for ch in str(name))
        while "__" in safe:
            safe = safe.replace("__", "_")
        return safe.strip("_") or "category"

    def _annual_values_from_monthly_df(df_monthly, unit):
        if df_monthly.empty:
            return np.array([], dtype=float)
        work = df_monthly[["Year", "Month", "Value"]].copy()
        work = work[np.isfinite(work["Value"])]
        if work.empty:
            return np.array([], dtype=float)
        work["WY"] = work["Year"] + (work["Month"] >= 10).astype(int)
        grouped = work.groupby("WY")["Value"]
        counts = grouped.count()
        if str(unit).upper() in _SUM_UNITS:
            annual = grouped.sum()
        else:
            annual = grouped.mean()
        annual = annual[counts == 12]
        return annual.replace([np.inf, -np.inf], np.nan).dropna().values

    def _annual_values_from_series(ser, unit):
        if ser is None or len(ser) == 0:
            return np.array([], dtype=float)
        idx = pd.DatetimeIndex(ser.index)
        df_monthly = pd.DataFrame({
            "Year": idx.year,
            "Month": idx.month,
            "Value": pd.to_numeric(ser.values, errors="coerce"),
        })
        return _annual_values_from_monthly_df(df_monthly, unit)

    def _build_pb_annual_lookup():
        """Returns {pk: np.array} -- pooled 1000-year annual values (all chunks)."""
        lookup = {}
        if not compiled_chunks:
            return lookup
        for tag in active_tags:
            df_chunk = compiled_chunks.get(tag)
            if df_chunk is None or df_chunk.empty:
                continue
            for pk, grp in df_chunk.groupby(["Part B", "Part C"]):
                unit = units_map.get(pk, "UNKNOWN")
                vals = _annual_values_from_monthly_df(grp, unit)
                if len(vals) == 0:
                    continue
                lookup.setdefault(pk, []).append(vals)
        return {
            pk: np.concatenate(parts)
            for pk, parts in lookup.items()
            if parts
        }

    def _save_rank_shift_concept_figure():
        rng = np.random.default_rng(42)
        hist_vals = rng.gamma(shape=4.2, scale=15.0, size=5000)
        pb_vals = rng.gamma(shape=4.2, scale=17.0, size=5000) + 4.0
        ex_grid = np.linspace(99, 1, 250)
        q_grid = 1.0 - ex_grid / 100.0
        hist_curve = np.quantile(hist_vals, q_grid)
        pb_curve = np.quantile(pb_vals, q_grid)
        hist_scale = float(np.nanmedian(np.abs(hist_vals)))
        pct_curve = []
        for q_prob, hq, pbq in zip(q_grid, hist_curve, pb_curve):
            denom = abs(hq) if abs(hq) > 1e-6 else hist_scale
            pct_curve.append((pbq - hq) / denom * 100.0)
        pct_curve = np.asarray(pct_curve)

        target_ex = 10.0
        target_q = 1.0 - target_ex / 100.0
        pb_target = float(np.quantile(pb_vals, target_q))
        hist_same_rank = float(np.quantile(hist_vals, target_q))
        pct_diff = (pb_target - hist_same_rank) / abs(hist_same_rank) * 100.0
        target_idx = int(np.argmin(np.abs(ex_grid - target_ex)))

        fig, axes = plt.subplots(1, 2, figsize=(8.6, 3.8))
        ax0, ax1 = axes
        ax0.plot(ex_grid, hist_curve, color="#4d4d4d", lw=1.6,
                 label="Historical reference")
        ax0.plot(ex_grid, pb_curve, color="#1f77b4", lw=1.8,
                 label="Product B")
        ax0.scatter([target_ex], [pb_target], color="#1f77b4", s=24, zorder=5)
        ax0.scatter([target_ex], [hist_same_rank], color="#4d4d4d", s=24, zorder=5)
        ax0.plot([target_ex, target_ex], [hist_same_rank, pb_target],
                 color="red", lw=1.2, ls="--")
        ax0.annotate(
            f"pct diff = {pct_diff:+.1f}%",
            xy=(target_ex, (hist_same_rank + pb_target) / 2.0),
            xytext=(-60, 0), textcoords="offset points",
            fontsize=7, color="red", ha="right", va="center",
            arrowprops={"arrowstyle": "-", "color": "red", "lw": 0.8},
        )
        ax0.set_xlim(100, 0)
        ax0.set_xlabel("Exceedance probability (%)")
        ax0.set_ylabel("WY annual value")
        ax0.set_title("Values at the same exceedance probability")
        ax0.grid(color="#dddddd", linestyle="--", linewidth=0.45)
        ax0.legend(frameon=False, loc="upper left")

        ax1.plot(ex_grid, pct_curve, color="#1f4e79", lw=1.8)
        ax1.scatter([ex_grid[target_idx]], [pct_curve[target_idx]],
                    color="red", s=24, zorder=5)
        ax1.axhline(0.0, color="red", ls="--", lw=0.7, alpha=0.55)
        ax1.set_xlim(100, 0)
        ax1.set_xlabel("Exceedance probability (%)")
        ax1.set_ylabel("% difference (Product B vs reference)")
        ax1.set_title("Percent difference across the distribution")
        ax1.grid(color="#dddddd", linestyle="--", linewidth=0.45)
        ax1.text(
            0.02, 0.03,
            "Positive means Product B value exceeds historical at the same exceedance probability.",
            transform=ax1.transAxes, fontsize=6, color="0.25",
            va="bottom", ha="left",
        )
        fig.suptitle(
            f"WY exceedance distribution concept (vs {ref_label})",
            y=1.02, fontsize=8,
        )
        fig.tight_layout()
        out_path = fig_dir / "wy_exceedance_rank_shift_concept.png"
        fig.savefig(out_path, bbox_inches="tight")
        plt.close(fig)
        print(f"    Figure: {fig_dir.name}/{out_path.name}")

    _EXCEEDANCE_PROBS = np.array(
        [99, 95, 90, 80, 70, 60, 50, 40, 30, 20, 10, 5, 1],
        dtype=float,
    )
    _QUANTILE_PROBS = 1.0 - (_EXCEEDANCE_PROBS / 100.0)
    _NEAR_COPY_RTOL = 0.005
    _NEAR_COPY_ATOL_FRAC = 0.01

    def _generate_summary_exceedance_figures():
        """Per-section exceedance figures using summary-level aggregation.

        Mirrors generate_summary_tables() grouping (Part-C groups, UNIMP flows,
        Rim Total, individual Part-B terms).  For each summary term plots 10
        individual chunk curves (n01-n10) plus a median; one figure per section
        per metric (rank-shift and pct-diff).
        """
        import math

        if not ref_series_by_pk or not compiled_chunks:
            return

        _inv = read_master_inventory()
        _const_keys = set(zip(
            _inv.loc[_inv["Constant_Rept"] == True, "Part_B"],
            _inv.loc[_inv["Constant_Rept"] == True, "Part_C"],
        ))
        _cat_lk = {(r["Part_B"], r["Part_C"]): r["Input_Category"]
                   for _, r in _inv.iterrows()}

        _MEAN_CAT = {"Climate", "Reservoir Evaporation"}
        _PARTC_SECTS = [
            ("CalSimHydro",              "CalSimHydro"),
            ("CalSimHydroEE",            "CalSimHydroEE"),
            ("Reservoir Evaporation",    "ResEvap"),
            ("Delta Channel Depletion",  "DCD"),
            ("Small Watersheds",         "SWS"),
            ("Tulare Groundwater Terms", "TulareGW"),
            ("Climate",                  "Climate"),
        ]
        _UNIMP_TERMS = [
            "UNIMP_FOLS", "UNIMP_ME", "UNIMP_OROV",
            "UNIMP_SJ", "UNIMP_SRBB", "UNIMP_ST", "UNIMP_TRIN",
            "UNIMP_TU", "UNIMP_YUBA",
        ]
        _UNIMP_SET = set(_UNIMP_TERMS)
        _RIM_EXCL = (
            _UNIMP_SET
            | {"UNIMP_SHAS", "UNIMP_WH", "FOLSM_INFLOW", "FOLSOM_INFLOW",
               "I_MELON_FCST", "I_MLRTN_IMP"}
        )
        _PARTB_SECTS = [
            ("Reservoir Storage Curves",  "ResCurves"),
            ("Instream Flows",            "InstreamFlows"),
            ("Other",                     "Miscellaneous"),
            ("Upper Watershed Modules",   "UpperWatershed"),
        ]

        out_dir = fig_dir / "summary_term_exceedance"
        out_dir.mkdir(exist_ok=True)

        _n_ch = len(active_tags)
        _clrs = [plt.cm.tab10(i % 10) for i in range(_n_ch)]

        # -- Monthly aggregation helpers --------------------------------------
        def _agg_ref(pk_set, agg):
            from collections import defaultdict
            monthly = defaultdict(list)
            for pk in pk_set:
                ser = ref_series_by_pk.get(pk)
                if ser is None:
                    continue
                for dt, val in zip(ser.index, ser.values):
                    if np.isfinite(val):
                        monthly[dt].append(val)
            if not monthly:
                return None
            fn = sum if agg == "sum" else (lambda v: float(np.mean(v)))
            return pd.Series({dt: fn(vals)
                               for dt, vals in monthly.items()}).sort_index()

        def _agg_chunk(tag, pk_set, agg):
            df_ch = compiled_chunks.get(tag)
            if df_ch is None or df_ch.empty:
                return None
            pk_df = pd.DataFrame(list(pk_set), columns=["Part B", "Part C"])
            sub = df_ch.merge(pk_df, on=["Part B", "Part C"])
            if sub.empty:
                return None
            grp = sub.groupby(["Year", "Month"])["Value"]
            agg_vals = grp.sum() if agg == "sum" else grp.mean()
            idx = [
                pd.Timestamp(int(y), int(m), 1).to_period("M").to_timestamp("M")
                for y, m in agg_vals.index
            ]
            return pd.Series(agg_vals.values,
                             index=pd.DatetimeIndex(idx)).sort_index()

        def _to_annual(ser, unit):
            if ser is None or ser.empty:
                return np.array([], dtype=float)
            df_m = pd.DataFrame({
                "Year": ser.index.year,
                "Month": ser.index.month,
                "Value": pd.to_numeric(ser.values, errors="coerce"),
            })
            return _annual_values_from_monthly_df(df_m, unit)

        # -- Exceedance metric computation ------------------------------------
        def _compute_curves(ref_ann, chunk_ann_by_tag):
            if len(ref_ann) < 5:
                return {}, {}
            s_ref = np.sort(ref_ann)
            scale = float(np.nanmedian(np.abs(ref_ann)))
            if not np.isfinite(scale) or scale <= 1e-6:
                scale = max(abs(float(np.nanmean(ref_ann))), 1.0)
            rank_d, pct_d = {}, {}
            for tag, pb_ann in chunk_ann_by_tag.items():
                if len(pb_ann) < 5:
                    continue
                rc, pc = [], []
                for ex_prob, q_prob in zip(_EXCEEDANCE_PROBS, _QUANTILE_PROBS):
                    rq = float(np.nanquantile(ref_ann, q_prob))
                    pq = float(np.nanquantile(pb_ann, q_prob))
                    cdf = np.searchsorted(s_ref, pq, side="right") / len(s_ref)
                    rc.append(float((cdf - q_prob) * 100.0))
                    denom = abs(rq) if abs(rq) > 1e-6 else scale
                    pc.append(float((pq - rq) / denom * 100.0))
                rank_d[tag] = np.array(rc)
                pct_d[tag] = np.array(pc)
            return rank_d, pct_d

        def _run_section(term_pk_sets, agg_by_term, unit_by_term):
            term_rank, term_pct, valid = {}, {}, []
            for term, pk_set in term_pk_sets.items():
                agg = agg_by_term.get(term, "sum")
                unit = unit_by_term.get(term, "UNKNOWN")
                ref_ann = _to_annual(_agg_ref(pk_set, agg), unit)
                ch_ann = {}
                for tag in active_tags:
                    ann = _to_annual(_agg_chunk(tag, pk_set, agg), unit)
                    if len(ann) >= 5:
                        ch_ann[tag] = ann
                rank_d, pct_d = _compute_curves(ref_ann, ch_ann)
                if rank_d:
                    term_rank[term] = rank_d
                    term_pct[term] = pct_d
                    valid.append(term)
            return term_rank, term_pct, valid

        # -- Figure generation ------------------------------------------------
        def _plot_section(s_safe, s_label, term_rank, term_pct, term_names):
            if not term_names:
                return
            n = len(term_names)
            nc = min(3, n)
            nr = math.ceil(n / nc)
            fw = max(4.5, 3.6 * nc)
            fh = max(3.2, 3.0 * nr)

            for curves_d, ylabel, fsuffix in [
                (term_rank, "Rank shift (ppt)",       "rank_shift"),
                (term_pct,  f"% diff vs {ref_label}", "pct_diff"),
            ]:
                if not any(curves_d.get(t) for t in term_names):
                    continue
                fig, axes = plt.subplots(nr, nc, figsize=(fw, fh),
                                         squeeze=False)
                afl = axes.flatten()
                for i, term in enumerate(term_names):
                    ax = afl[i]
                    tc = curves_d.get(term, {})
                    for j, tag in enumerate(active_tags):
                        v = tc.get(tag)
                        if v is None:
                            continue
                        ax.plot(_EXCEEDANCE_PROBS, v,
                                color=_clrs[j], lw=0.9, alpha=0.65, zorder=2)
                    all_c = [tc[t] for t in active_tags if t in tc]
                    if all_c:
                        med = np.nanmedian(np.stack(all_c), axis=0)
                        ax.plot(_EXCEEDANCE_PROBS, med,
                                color="black", lw=2.0, zorder=4)
                    ax.axhline(0.0, color="red", ls="--", lw=0.7, alpha=0.55)
                    ax.set_xlim(100, 0)
                    ax.set_xticks([99, 90, 70, 50, 30, 10, 1])
                    ax.tick_params(labelsize=_FS - 1)
                    ax.set_title(term, fontsize=_FS)
                    ax.set_ylabel(ylabel, fontsize=_FS - 1)
                    ax.set_xlabel("Exceedance prob. (%)", fontsize=_FS - 1)
                    ax.grid(color="#dddddd", linestyle="--",
                            linewidth=0.4, alpha=0.7)
                    avals = (np.concatenate(list(tc.values()))
                             if tc else np.array([0.0]))
                    max_abs = float(np.nanmax(np.abs(avals)))
                    lim = max(10.0, np.ceil(max_abs / 10) * 10)
                    ax.set_ylim(-lim, lim)
                for k in range(n, len(afl)):
                    afl[k].set_visible(False)
                leg_h = [
                    plt.Line2D([0], [0], color=_clrs[j], lw=0.9, label=tag)
                    for j, tag in enumerate(active_tags)
                ] + [plt.Line2D([0], [0], color="black", lw=2.0, label="Median")]
                leg_ncol = min(6, _n_ch + 1)
                fig.suptitle(
                    f"{s_label}: Product B vs {ref_label}  |  {ylabel}",
                    fontsize=_FS + 1,
                )
                fig.tight_layout()
                fig.legend(handles=leg_h, loc="upper center",
                           ncol=leg_ncol, fontsize=_FS - 1,
                           frameon=False,
                           bbox_to_anchor=(0.5, 0.0),
                           bbox_transform=fig.transFigure)
                op = out_dir / f"{s_safe}_{fsuffix}.png"
                fig.savefig(op, bbox_inches="tight")
                plt.close(fig)
                print(f"    Figure: {out_dir.name}/{op.name}")

        # -- Part-C-grouped sections ------------------------------------------
        # Part C terms to exclude from CalSimHydro (unchanged by design).
        _CALSIMHYDRO_PC_EXCL = {"URBAN-DEMAND", "WASTEWATER"}

        for cat_name, s_safe in _PARTC_SECTS:
            agg = "mean" if cat_name in _MEAN_CAT else "sum"
            pks_by_pc = {}
            for (pb, pc), cat in _cat_lk.items():
                if cat == cat_name:
                    if cat_name == "CalSimHydro" and pc in _CALSIMHYDRO_PC_EXCL:
                        continue
                    pks_by_pc.setdefault(pc, set()).add((pb, pc))
            if not pks_by_pc:
                continue
            unit_by_term = {
                pc: units_map.get(next(iter(pks)), "UNKNOWN")
                for pc, pks in pks_by_pc.items()
            }
            tr, tp, valid = _run_section(
                pks_by_pc,
                {pc: agg for pc in pks_by_pc},
                unit_by_term,
            )
            _plot_section(s_safe, cat_name, tr, tp,
                          [t for t in sorted(pks_by_pc) if t in tr])

        # -- Rim Inflow -- Unimpaired ------------------------------------------
        rim_pks = {(pb, pc) for (pb, pc), cat in _cat_lk.items()
                   if cat == "Rim Inflow"}
        u_pks = {u: {pk for pk in rim_pks if pk[0] == u}
                 for u in _UNIMP_TERMS}
        u_pks = {u: s for u, s in u_pks.items() if s}
        if u_pks:
            unit_u = {u: units_map.get(next(iter(s)), "TAF")
                      for u, s in u_pks.items()}
            tr, tp, valid = _run_section(u_pks,
                                         {u: "sum" for u in u_pks},
                                         unit_u)
            _plot_section("RimUNIMP", "Rim Inflow - Unimpaired",
                          tr, tp, valid)

        # -- Rim Inflow -- Total -----------------------------------------------
        rim_tot = {pk for pk in rim_pks
                   if pk[0] not in _RIM_EXCL and "_UHH" not in pk[0]}
        if rim_tot:
            unit_tot = units_map.get(next(iter(rim_tot)), "TAF")
            tr, tp, valid = _run_section(
                {"Total": rim_tot}, {"Total": "sum"}, {"Total": unit_tot}
            )
            _plot_section("RimTotal", "Rim Inflow - Total", tr, tp, valid)

        # -- Part-B sections --------------------------------------------------
        for cat_name, s_safe in _PARTB_SECTS:
            pks_by_pb = {}
            for (pb, pc), cat in _cat_lk.items():
                if cat == cat_name and (pb, pc) not in _const_keys:
                    pks_by_pb.setdefault(pb, set()).add((pb, pc))
            if not pks_by_pb:
                continue
            unit_by_term = {
                pb: units_map.get(next(iter(pks)), "UNKNOWN")
                for pb, pks in pks_by_pb.items()
            }
            tr, tp, valid = _run_section(
                pks_by_pb,
                {pb: "sum" for pb in pks_by_pb},
                unit_by_term,
            )
            _plot_section(s_safe, cat_name, tr, tp,
                          [t for t in sorted(pks_by_pb) if t in tr])

        print(f"    Figures: {out_dir.name}/")

    if ref_series_by_pk and compiled_chunks:
        dist_df = annual_df[~annual_df["Constant_Rept"]].copy()
        pb_annual_lookup = _build_pb_annual_lookup()
        ref_annual_lookup = {}
        for pk, ser in ref_series_by_pk.items():
            unit = units_map.get(pk, "UNKNOWN")
            vals = _annual_values_from_series(ser, unit)
            if len(vals) > 0:
                ref_annual_lookup[pk] = vals

        metric_rows = []
        for _, row in dist_df.iterrows():
            pk = (row["Part_B"], row["Part_C"])
            ref_vals = ref_annual_lookup.get(pk)
            pb_vals = pb_annual_lookup.get(pk)
            if ref_vals is None or pb_vals is None:
                continue
            ref_vals = np.asarray(ref_vals, dtype=float)
            pb_vals = np.asarray(pb_vals, dtype=float)
            ref_vals = ref_vals[np.isfinite(ref_vals)]
            pb_vals = pb_vals[np.isfinite(pb_vals)]
            if len(ref_vals) < 10 or len(pb_vals) < 10:
                continue

            sorted_ref = np.sort(ref_vals)
            ref_mean_annual = float(np.nanmean(ref_vals))
            pb_mean_annual = float(np.nanmean(pb_vals))
            ref_scale = float(np.nanmedian(np.abs(ref_vals)))
            if not np.isfinite(ref_scale) or ref_scale <= 1e-6:
                ref_scale = abs(ref_mean_annual)
            if not np.isfinite(ref_scale) or ref_scale <= 1e-6:
                ref_scale = 1.0

            rank_values = []
            pct_values = []
            ref_quantiles = []
            pb_quantiles = []
            term_metric_rows = []
            for ex_prob, q_prob in zip(_EXCEEDANCE_PROBS, _QUANTILE_PROBS):
                ref_q = float(np.nanquantile(ref_vals, q_prob))
                pb_q = float(np.nanquantile(pb_vals, q_prob))
                hist_cdf = np.searchsorted(sorted_ref, pb_q, side="right") / len(sorted_ref)
                rank_shift = float((hist_cdf - q_prob) * 100.0)
                denom = abs(ref_q) if abs(ref_q) > 1e-6 else ref_scale
                pct_diff = float((pb_q - ref_q) / denom * 100.0)
                rank_values.append(rank_shift)
                pct_values.append(pct_diff)
                ref_quantiles.append(ref_q)
                pb_quantiles.append(pb_q)
                term_metric_rows.append({
                    "Input_Category": row["Input_Category"],
                    "Part_B": row["Part_B"],
                    "Part_C": row["Part_C"],
                    "Units": row["Units"],
                    "Exceedance_Probability": ex_prob,
                    "Quantile_Probability": q_prob,
                    "Ref_Quantile": ref_q,
                    "ProductB_Quantile": pb_q,
                    "Rank_Shift_Ppt": rank_shift,
                    "Pct_Diff": pct_diff,
                    "Pct_Diff_Denominator": denom,
                    "Ref_WY_Count": len(ref_vals),
                    "ProductB_WY_Count": len(pb_vals),
                    "Ref_Mean_Annual": ref_mean_annual,
                    "ProductB_Mean_Annual": pb_mean_annual,
                })

            ref_quantiles = np.asarray(ref_quantiles, dtype=float)
            pb_quantiles = np.asarray(pb_quantiles, dtype=float)
            quantile_abs_diff = np.abs(pb_quantiles - ref_quantiles)
            max_abs_quantile_diff = float(np.nanmax(quantile_abs_diff))
            max_abs_pct_diff = float(np.nanmax(np.abs(pct_values)))
            near_copy_atol = max(1e-8, _NEAR_COPY_ATOL_FRAC * ref_scale)
            near_reference_copy = bool(np.allclose(
                pb_quantiles, ref_quantiles,
                rtol=_NEAR_COPY_RTOL, atol=near_copy_atol,
                equal_nan=False,
            ))

            for metric_row in term_metric_rows:
                metric_row["Near_Reference_Copy"] = near_reference_copy
                metric_row["Near_Copy_Rtol"] = _NEAR_COPY_RTOL
                metric_row["Near_Copy_Atol"] = near_copy_atol
                metric_row["Max_Abs_Quantile_Diff"] = max_abs_quantile_diff
                metric_row["Max_Abs_Pct_Diff_Across_Quantiles"] = max_abs_pct_diff
            metric_rows.extend(term_metric_rows)

        if metric_rows:
            metrics_df = pd.DataFrame(metric_rows)
            metrics_path = fig_dir / "wy_exceedance_distribution_metrics.csv"
            metrics_df.to_csv(metrics_path, index=False)

            summary = metrics_df.groupby(
                ["Input_Category", "Part_B", "Part_C", "Units"],
                as_index=False,
            ).agg(
                Mean_Rank_Shift_Ppt=("Rank_Shift_Ppt", "mean"),
                Mean_Abs_Rank_Shift_Ppt=("Rank_Shift_Ppt", lambda x: np.nanmean(np.abs(x))),
                Median_Pct_Diff=("Pct_Diff", "median"),
                Max_Abs_Pct_Diff=("Pct_Diff", lambda x: np.nanmax(np.abs(x))),
                Ref_WY_Count=("Ref_WY_Count", "first"),
                ProductB_WY_Count=("ProductB_WY_Count", "first"),
                Ref_Mean_Annual=("Ref_Mean_Annual", "first"),
                ProductB_Mean_Annual=("ProductB_Mean_Annual", "first"),
                Near_Reference_Copy=("Near_Reference_Copy", "first"),
                Max_Abs_Quantile_Diff=("Max_Abs_Quantile_Diff", "first"),
                Max_Abs_Pct_Diff_Across_Quantiles=("Max_Abs_Pct_Diff_Across_Quantiles", "first"),
            )
            summary_path = fig_dir / "wy_exceedance_distribution_summary.csv"
            summary.to_csv(summary_path, index=False)
            print(f"    CSV: {metrics_path.name}")
            print(f"    CSV: {summary_path.name}")
            n_near_copy = int(summary["Near_Reference_Copy"].sum())
            if n_near_copy:
                print(f"    Near-reference-copy terms: {n_near_copy}")

            _save_rank_shift_concept_figure()

        _generate_summary_exceedance_figures()
    elif ref_series_by_pk is not None or compiled_chunks is not None:
        print("    WARNING: WY exceedance plots skipped; missing reference or Product B series.")

    # -- Per-category chunk spread (one figure per category) --
    spread_dir = fig_dir / "chunk_spread_by_category"
    spread_dir.mkdir(exist_ok=True)

    for cat in cats_present:
        cat_df = annual_df[annual_df["Input_Category"] == cat]
        if cat_df.empty:
            continue
        unique_units = sorted(cat_df["Units"].dropna().unique())
        if not unique_units:
            unique_units = ["UNKNOWN"]
        n_unit_panels = len(unique_units)
        n_sv = len(cat_df)

        n_cols_bot = max(1, n_unit_panels)
        gs = GridSpec(2, n_cols_bot, height_ratios=[1, 1])
        fig = plt.figure(figsize=(6.5, 5.0))

        ax_top = fig.add_subplot(gs[0, :])
        pct_data = []
        for tag in active_tags:
            col = f"{tag}_pct_diff"
            if col in cat_df.columns:
                pct_data.append(cat_df[col].dropna().values)
            else:
                pct_data.append(np.array([]))
        if any(len(d) > 0 for d in pct_data):
            bp = ax_top.boxplot(
                pct_data, vert=True, patch_artist=True, widths=0.6,
                labels=list(active_tags),
            )
            for patch in bp["boxes"]:
                patch.set_facecolor("#5B9BD5")
                patch.set_alpha(0.7)
            for cap in bp["caps"]:
                cap.set_visible(False)
            for flier in bp["fliers"]:
                flier.set(marker="o", markersize=1,
                          markerfacecolor="k", markeredgecolor="none")
        ax_top.axhline(0.0, color="red", ls="--", lw=0.6, alpha=0.5)
        ax_top.set_ylabel(f"Mean Annual % Diff (vs {ref_label})")
        ax_top.set_title(f"{cat}  (n={n_sv} SVs)")
        ax_top.tick_params(axis="x", rotation=45)

        for u_idx, unit in enumerate(unique_units):
            ax = fig.add_subplot(gs[1, u_idx])
            unit_df = cat_df[cat_df["Units"] == unit]
            n_unit_sv = len(unit_df)
            abs_data = []
            for tag in active_tags:
                col = f"{tag}_abs_diff"
                if col in unit_df.columns:
                    abs_data.append(unit_df[col].dropna().values)
                else:
                    abs_data.append(np.array([]))
            if any(len(d) > 0 for d in abs_data):
                bp = ax.boxplot(
                    abs_data, vert=True, patch_artist=True, widths=0.6,
                    labels=list(active_tags),
                )
                for patch in bp["boxes"]:
                    patch.set_facecolor("#E8A54B")
                    patch.set_alpha(0.7)
                for cap in bp["caps"]:
                    cap.set_visible(False)
                for flier in bp["fliers"]:
                    flier.set(marker="o", markersize=1,
                              markerfacecolor="k", markeredgecolor="none")
            ax.axhline(0.0, color="red", ls="--", lw=0.6, alpha=0.5)
            ax.set_ylabel(f"Mean Annual Abs Diff ({unit})")
            if n_unit_panels > 1:
                ax.set_title(f"{unit}  (n={n_unit_sv})")
            ax.tick_params(axis="x", rotation=45)

        fig.tight_layout()
        cat_safe = cat.replace(" ", "_").replace("/", "_")
        fig.savefig(
            spread_dir / f"{cat_safe}.png", bbox_inches="tight"
        )
        plt.close(fig)

    print(f"    Figures: {fig_dir.name}/chunk_spread_by_category/ "
          f"({len(cats_present)} categories)")

    # -- Monthly climatology (per SV, per category) --
    if not skip_climatology:
        scatter_dir = fig_dir / "monthly_climatology"
        scatter_dir.mkdir(exist_ok=True)

        for cat in cats_present:
            cat_cmp = cmp_df[cmp_df["Input_Category"] == cat]
            sv_list = cat_cmp.groupby(["Part_B", "Part_C"]).ngroups
            if sv_list == 0:
                continue
            n_plotted = 0
            for (partb, partc), sv_cmp in cat_cmp.groupby(["Part_B", "Part_C"]):
                wy_month_nums = [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]
                month_pos = {month: i + 1 for i, month in enumerate(wy_month_nums)}
                sv_cmp = sv_cmp.copy()
                sv_cmp["Month_Position"] = sv_cmp["Month"].map(month_pos)
                sv_cmp = sv_cmp.sort_values("Month_Position")
                months_data = sv_cmp["Month_Position"].values

                fig, ax = plt.subplots(figsize=(5.0, 2.5))
                chunk_cols = [f"{t}_mean" for t in active_tags
                              if f"{t}_mean" in sv_cmp.columns]
                for i, col in enumerate(chunk_cols):
                    vals = sv_cmp[col].values
                    ax.plot(months_data, vals, color="tab:orange", alpha=0.3,
                            lw=0.8, label="Product B" if i == 0 else None)
                ref_vals = sv_cmp["Ref_mean"].values
                ax.plot(months_data, ref_vals, color="tab:blue", lw=1.2,
                        label=ref_label)
                ax.set_xticks(range(1, 13))
                ax.set_xticklabels(_MONTH_LABELS)
                ax.set_xlabel("Month")
                ax.set_ylabel("Monthly Mean")
                ax.set_title(f"{partb}/{partc}")
                ax.legend(loc="best", framealpha=0.7)
                fig.tight_layout()

                cat_safe = cat.replace(" ", "_").replace("/", "_")
                cat_plot_dir = scatter_dir / cat_safe
                cat_plot_dir.mkdir(parents=True, exist_ok=True)
                fname = f"{partb}__{partc}.png".replace("/", "_")
                fig.savefig(cat_plot_dir / fname, bbox_inches="tight")
                plt.close(fig)
                n_plotted += 1
            print(f"      {cat}: {n_plotted} climatology plots")


# ======================================================================
# MAIN EXECUTION
# ======================================================================
print("=" * 72)
print("  Product B (1000-year stochastic) -- CSV Compilation")
print("=" * 72)

if not BASELINE_DSS.exists():
    sys.exit(f"ERROR: Baseline DSS not found:\n  {BASELINE_DSS}")
if not INVENTORY_XLSX.exists():
    sys.exit(f"ERROR: Master inventory not found:\n  {INVENTORY_XLSX}")

# Prompt to close Excel if locked
while True:
    try:
        with open(INVENTORY_XLSX, "r+b"):
            pass
        break
    except PermissionError:
        print(
            f"\nWARNING: The master inventory appears to be open in Excel.\n"
            f"  {INVENTORY_XLSX}\n"
            f"  Close the file, then press Enter to continue (or Ctrl+C to abort)..."
        )
        try:
            input()
        except KeyboardInterrupt:
            sys.exit("\nAborted.")

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
if not CLI_ARGS.skip_dss:
    COMPILED_DIR.mkdir(parents=True, exist_ok=True)

print(f"  Active chunks: {', '.join(ACTIVE_TAGS)}")
print(f"  Output: {OUTPUT_DIR}")
print()

# ==================================================================
# FAST PATH: --summary-figures (regenerate figures from previous run)
# ==================================================================
if CLI_ARGS.summary_figures:
    print("--- Running in --summary-figures mode (figures only) ---")
    print()

    # Catalog baseline DSS for unit extraction
    with dss_io.open_dss(BASELINE_DSS, version=6, catalog_flag=True) as _dss_b:
        _bp = _dss_b.getPathnameList(DSS_PATTERN)
    baseline_bucket = {}
    for p in _bp:
        k = path_key(p)
        baseline_bucket.setdefault(k, []).append(p)

    # Collect all SVs across available comparison CSVs
    all_compiled_svs = set()
    _comparisons_to_plot = []  # (csv_path, ref_col, ref_label, fig_subdir)

    _cmp_a_csv = OUTPUT_DIR / "product_b_vs_a_comparison.csv"
    if _cmp_a_csv.exists():
        _comparisons_to_plot.append(
            (_cmp_a_csv, "Product_A_mean", "Product A", "vs_product_a"))

    _cmp_b_csv = OUTPUT_DIR / "product_b_vs_calsim_base_comparison.csv"
    if _cmp_b_csv.exists():
        _comparisons_to_plot.append(
            (_cmp_b_csv, "CalSim_Base_mean", "CalSim Base", "vs_calsim_base"))

    if not _comparisons_to_plot:
        sys.exit(f"ERROR: No comparison CSVs found from a previous run in:\n  {OUTPUT_DIR}")

    # Extract units (once, for all SVs)
    for _csv_path, _, _, _ in _comparisons_to_plot:
        _df_tmp = pd.read_csv(_csv_path)
        all_compiled_svs |= set(zip(_df_tmp["Part_B"], _df_tmp["Part_C"]))

    units_map = {}
    with dss_io.open_dss(BASELINE_DSS, version=6, catalog_flag=False) as _dss_u:
        for _pk in sorted(all_compiled_svs):
            if _pk not in baseline_bucket:
                continue
            try:
                _ts_u = _dss_u.read_ts(
                    baseline_bucket[_pk][0], trim_missing=False
                )
                units_map[_pk] = _ts_u.units
            except Exception:
                units_map[_pk] = "UNKNOWN"
    print(f"  Extracted units for {len(units_map):,} SVs")

    fig_root = OUTPUT_DIR / "figures"
    fig_root.mkdir(exist_ok=True)

    print("  Loading compiled Product B CSVs for WY exceedance figures ...")
    _summary_compiled_chunks = _load_compiled_chunks_from_csv(ACTIVE_TAGS)
    if _summary_compiled_chunks:
        print(f"  Loaded {len(_summary_compiled_chunks):,} compiled chunk CSVs")
    else:
        print("  WARNING: No compiled chunk CSVs found; WY exceedance figures will be skipped.")

    for _csv_path, _ref_col, _ref_label, _fig_subdir in _comparisons_to_plot:
        cmp_df = pd.read_csv(_csv_path)
        print(f"  Loaded {_csv_path.name} ({len(cmp_df):,} rows)")

        # Rename reference column to Ref_mean for the figure function
        if _ref_col in cmp_df.columns:
            cmp_df = cmp_df.rename(columns={_ref_col: "Ref_mean"})

        if _ref_label == "Product A" and PRODUCT_A_DSS.exists():
            ref_series = read_product_a_monthly_series(
                PRODUCT_A_DSS, all_compiled_svs
            )
        elif _ref_label == "CalSim Base":
            ref_series = read_calsim_base_monthly_series(
                baseline_bucket, all_compiled_svs, CB_START, CB_END
            )
        else:
            ref_series = {}

        print(f"  Generating {_ref_label} figures ...")
        try:
            _generate_comparison_figures(
                cmp_df, fig_root / _fig_subdir, _ref_label,
                ACTIVE_TAGS, units_map, skip_climatology=True,
                ref_series_by_pk=ref_series,
                compiled_chunks=_summary_compiled_chunks,
            )
        except ImportError:
            print("  WARNING: matplotlib not available, skipping figures.")

    print(f"\nDone (--summary-figures).  Output in: {OUTPUT_DIR}")
    sys.exit(0)


# ==================================================================
# STEP 1 -- Scan module directories & read CSVs
# ==================================================================
print("Step 1: Scanning module directories for Product B CSVs ...")
print("-" * 72)

t0_scan = time.time()

# chunk_tag -> module_label -> list[DataFrame]
all_chunk_data = {tag: {} for tag in CHUNK_TAGS}
module_svs     = {}          # label -> set of (Part B, Part C)
module_files   = {}          # label -> int (file count)
module_src_paths = {}        # label -> list[Path] (for copying)
missing_modules = []

for label, (src_dir, _inv_cat, reader_type) in MODULE_CONFIG_B.items():
    chunk_data, n_files, svs, src_paths = scan_module_csvs(label, src_dir, reader_type)
    if n_files > 0:
        module_files[label] = n_files
        module_svs[label]   = svs
        module_src_paths[label] = src_paths
        for tag in CHUNK_TAGS:
            if chunk_data[tag]:
                all_chunk_data[tag][label] = chunk_data[tag]
        print(f"  [OK]   {label:40s}  {n_files:>4} files  |  "
              f"{len(svs):>4} unique (B,C)")
    else:
        missing_modules.append(label)
        print(f"  [SKIP] {label:40s}  (no chunk CSVs found in {src_dir})")

print("-" * 72)
print(f"  Modules found: {len(module_files)} / {len(MODULE_CONFIG_B)}")
if missing_modules:
    print(f"  Modules skipped: {', '.join(missing_modules)}")
print(f"  Scan completed in {time.time()-t0_scan:.1f}s")
print()

if not module_files:
    sys.exit("ERROR: No module CSVs found. Nothing to compile.")

# Collect all (Part B, Part C) provided by modules across all chunks
all_module_keys = set()
for svs in module_svs.values():
    all_module_keys |= svs
print(f"  Total unique (Part B, Part C) from modules: {len(all_module_keys):,}")


# -- Copy source CSVs into compiled_input_files/<module>/ --
print()
print("  Copying source CSVs into compiled_input_files/ ...")
COMPILED_CSV.mkdir(parents=True, exist_ok=True)
for label, src_paths in module_src_paths.items():
    dest_dir = COMPILED_CSV / label
    dest_dir.mkdir(parents=True, exist_ok=True)
    for csv_path in src_paths:
        shutil.copy2(csv_path, dest_dir / csv_path.name)
    print(f"    {label:40s}  -> {len(src_paths)} file(s) copied")


# ==================================================================
# STEP 2 -- Inventory cross-reference
# ==================================================================
print()
print("Step 2: Cross-referencing against master inventory ...")

inventory_df = read_master_inventory()
print(f"  Inventory rows loaded: {len(inventory_df):,}")

# Category -> module label mapping
inv_category_map = {}
for label, (_dir, inv_cat, _reader) in MODULE_CONFIG_B.items():
    inv_category_map[inv_cat] = label

# Skipped: Missing=T
inv_skipped_missing = inventory_df[inventory_df["Missing"] == True].copy()
skipped_missing_keys = set(zip(inv_skipped_missing["Part_B"], inv_skipped_missing["Part_C"]))

# Skipped: Used_in_DCR=F (and not already Missing)
inv_skipped_not_dcr = inventory_df[
    (~inventory_df["Missing"]) & (~inventory_df["Used_in_DCR"])
].copy()
skipped_not_dcr_keys = set(zip(inv_skipped_not_dcr["Part_B"], inv_skipped_not_dcr["Part_C"]))

# Constant/Rept SVs
inv_const_rept_all = inventory_df[
    (inventory_df["Constant_Rept"] == True)
    & (~inventory_df["Missing"])
    & (inventory_df["Used_in_DCR"])
].copy()
const_rept_all_keys = set(zip(inv_const_rept_all["Part_B"], inv_const_rept_all["Part_C"]))

# Active SVs: not missing, used in DCR, not constant/rept,
# in a category with a module (or a known no-module category)
expected_svs = inventory_df[
    (~inventory_df["Missing"])
    & (inventory_df["Used_in_DCR"])
    & (~inventory_df["Constant_Rept"])
    & (inventory_df["Input_Category"].isin(inv_category_map.keys()))
].copy()
expected_svs["Module"] = expected_svs["Input_Category"].map(inv_category_map).fillna("(no module)")
expected_keys = set(zip(expected_svs["Part_B"], expected_svs["Part_C"]))

# Which expected SVs are actually provided?
inv_expected_modified = expected_keys & all_module_keys
inv_expected_missing  = expected_keys - all_module_keys
# Which module SVs are unexpected (not in inventory)?
inv_unexpected = all_module_keys - expected_keys - const_rept_all_keys

print(f"  Skipped (Missing=T):           {len(skipped_missing_keys):>6,}")
print(f"  Skipped (Used_in_DCR=F):       {len(skipped_not_dcr_keys):>6,}")
print(f"  Constant/Rept (total):         {len(const_rept_all_keys):>6,}")
print(f"  Expected from modules:         {len(expected_keys):>6,}")
print(f"  Expected & provided:           {len(inv_expected_modified):>6,}")
print(f"  Expected but MISSING:          {len(inv_expected_missing):>6,}")
print(f"  Provided but NOT in inventory: {len(inv_unexpected):>6,}")
print()


# ==================================================================
# STEP 3 -- Auto-fill Constant/Rept SVs from baseline
# ==================================================================
print("Step 3: Auto-filling Constant/Rept SVs from baseline DSS ...")

# Constant/Rept keys that are NOT already provided by module CSVs
const_rept_to_fill = const_rept_all_keys - all_module_keys

# Also include Constant/Rept from categories WITHOUT modules
const_rept_other = inventory_df[
    (inventory_df["Constant_Rept"] == True)
    & (~inventory_df["Missing"])
    & (inventory_df["Used_in_DCR"])
    & (~inventory_df["Input_Category"].isin(inv_category_map.keys()))
].copy()
const_rept_to_fill |= set(zip(const_rept_other["Part_B"], const_rept_other["Part_C"]))
const_rept_to_fill -= all_module_keys

print(f"  Constant/Rept SVs to auto-fill: {len(const_rept_to_fill):,}")

# Catalog baseline DSS
with dss_io.open_dss(BASELINE_DSS, version=6, catalog_flag=True) as _dss_b:
    _bp = _dss_b.getPathnameList(DSS_PATTERN)
baseline_bucket = {}
for p in _bp:
    k = path_key(p)
    baseline_bucket.setdefault(k, []).append(p)
print(f"  Baseline unique (Part B, Part C): {len(baseline_bucket):,}")

# Build constant/rept DataFrames (one per Part B/C, same for all chunks)
const_rept_frames = {}   # (Part B, Part C) -> DataFrame
const_rept_filled = set()

t0_cr = time.time()
with dss_io.open_dss(BASELINE_DSS, version=6, catalog_flag=True) as dss_in:
    for pk in sorted(const_rept_to_fill):
        df_cr = build_constant_rept_rows(dss_in, baseline_bucket, pk)
        if not df_cr.empty:
            const_rept_frames[pk] = df_cr
            const_rept_filled.add(pk)

print(f"  Successfully auto-filled: {len(const_rept_filled):,} (B,C)  ({time.time()-t0_cr:.1f}s)")
print(f"  Already provided by module CSVs: {len(const_rept_all_keys & all_module_keys):,}")
print()


# ==================================================================
# STEP 4 -- Merge per-chunk data in memory
# ==================================================================
print("Step 4: Merging per-chunk data ...")
print("-" * 72)

t0_compile = time.time()
chunk_stats = {}           # tag -> {n_rows, n_svs, modules}
chunk_sv_sets = {}         # tag -> set of (Part B, Part C)
chunk_sv_month_counts = {} # tag -> Series indexed by (Part B, Part C)
compiled_chunks = {}       # tag -> DataFrame  (kept in memory for DSS / comparison)

for chunk_idx in ACTIVE_CHUNKS:
    tag = f"n{chunk_idx:02d}"
    t_chunk = time.time()

    frames = []

    # Module CSVs for this chunk
    modules_in_chunk = []
    if tag in all_chunk_data:
        for label, dfs in all_chunk_data[tag].items():
            for df in dfs:
                frames.append(df)
            modules_in_chunk.append(label)

    # Constant/Rept rows (same for every chunk)
    for pk, df_cr in const_rept_frames.items():
        frames.append(df_cr)

    if not frames:
        print(f"  {tag}: WARNING -- no data, skipping")
        continue

    compiled = pd.concat(frames, ignore_index=True)

    # Trim to canonical Product B window (Oct 1921 - Sep 2021)
    ym = compiled["Year"] * 100 + compiled["Month"]
    compiled = compiled[(ym >= 192110) & (ym <= 202109)].copy()

    # De-duplicate: keep last occurrence for each (Part B, Part C, Year, Month)
    compiled = compiled.drop_duplicates(
        subset=["Part B", "Part C", "Year", "Month"], keep="last"
    )

    # Sort
    compiled = compiled.sort_values(
        ["Part B", "Part C", "Year", "Month"]
    ).reset_index(drop=True)

    # Write per-chunk compiled CSV
    COMPILED_DIR.mkdir(parents=True, exist_ok=True)
    chunk_csv_path = COMPILED_DIR / f"ProductB_SV_{tag}.csv"
    compiled.to_csv(chunk_csv_path, index=False)

    compiled_chunks[tag] = compiled

    sv_keys = set(zip(compiled["Part B"], compiled["Part C"]))
    chunk_sv_sets[tag] = sv_keys

    # Per-SV month counts for this chunk
    sv_counts = compiled.groupby(["Part B", "Part C"]).size()
    chunk_sv_month_counts[tag] = sv_counts
    wrong_months = sv_counts[sv_counts != 1200]
    if not wrong_months.empty:
        print(f"  {tag}: WARNING -- {len(wrong_months)} SVs with != 1200 months:")
        for (b, c), cnt in wrong_months.items():
            print(f"    {b} / {c}: {cnt} months")

    chunk_stats[tag] = {
        "n_rows": len(compiled),
        "n_svs": len(sv_keys),
        "modules": modules_in_chunk,
    }

    print(f"  {tag}:  {len(compiled):>10,} rows  |  {len(sv_keys):>5,} (B,C)  |  "
          f"{len(modules_in_chunk)} modules  ({time.time()-t_chunk:.1f}s)")

print("-" * 72)
print(f"  Merge completed in {time.time()-t0_compile:.1f}s")
print()


# ==================================================================
# STEP 4b -- Write per-chunk DSS files
# ==================================================================
if not CLI_ARGS.skip_dss:
    print("Step 4b: Writing per-chunk DSS files ...")
    print("-" * 72)

    from pydsstools.core import TimeSeriesContainer

    t0_dss = time.time()
    dss_chunk_paths = {}  # tag -> Path

    # -- 4b-i: Pre-build chunk lookups ---------------------------------
    print("  Building per-chunk value lookups ...")
    chunk_lookups = {}   # tag -> {(B,C): {(year,month): value}}
    all_compiled_pks = set()

    for chunk_idx in ACTIVE_CHUNKS:
        tag = f"n{chunk_idx:02d}"
        if tag not in compiled_chunks:
            continue
        df_chunk = compiled_chunks[tag]
        lookup = {}
        for (b, c), grp in df_chunk.groupby(["Part B", "Part C"]):
            ym_dict = dict(zip(zip(grp["Year"].astype(int), grp["Month"].astype(int)),
                               grp["Value"]))
            lookup[(b, c)] = ym_dict
            all_compiled_pks.add((b, c))
        chunk_lookups[tag] = lookup

    # -- 4b-ii: Cache baseline TS data (read each pathname once) -------
    print("  Caching baseline time series data ...")
    t0_cache = time.time()
    baseline_ts_cache = {}  # pathname -> dict

    with dss_io.open_dss(BASELINE_DSS, version=6, catalog_flag=True) as dss_in:
        for pk in sorted(all_compiled_pks):
            if pk not in baseline_bucket:
                continue
            for pathname in baseline_bucket[pk]:
                if pathname in baseline_ts_cache:
                    continue
                try:
                    ts = dss_in.read_ts(pathname, trim_missing=False)
                except Exception:
                    continue
                eom = dss_eom(ts.pytimes)
                eom_pd = pd.DatetimeIndex(eom)
                baseline_ts_cache[pathname] = {
                    'start_dt':   ts.pytimes[0].strftime("%d%b%Y %H:%M"),
                    'values':     np.array(ts.values, dtype=float),
                    'units':      ts.units,
                    'type':       ts.type,
                    'interval':   ts.interval,
                    'eom_years':  eom_pd.year.values.copy(),
                    'eom_months': eom_pd.month.values.copy(),
                }
    print(f"  Cached {len(baseline_ts_cache):,} baseline pathnames ({time.time()-t0_cache:.1f}s)")

    # -- 4b-iii: Prepare output files (copy baseline sequentially) -----
    # The chunk-write loop sets up the junction once for all 10 writes (rather
    # than per-iteration) to avoid the mklink/rmdir overhead; that's why we
    # call dss_io.create_junction / remove_junction explicitly here instead
    # of using dss_io.open_dss for each write inside the loop.
    use_junction = dss_io.needs_junction(COMPILED_DIR / "ProductB_SV_n01.dss")
    if use_junction:
        dss_io.create_junction(COMPILED_DIR)
        atexit.register(dss_io.remove_junction)

    def _get_dss_str(path):
        if use_junction:
            return str(dss_io._DSS_LINK / path.name)
        return str(path)

    for tag in sorted(chunk_lookups):
        chunk_dss = COMPILED_DIR / f"ProductB_SV_{tag}.dss"
        for f in [chunk_dss] + [chunk_dss.with_suffix(e) for e in [".dsd", ".dsk", ".dsc"]]:
            if f.exists():
                f.unlink()
        shutil.copy2(BASELINE_DSS, chunk_dss)
        for ext in [".dsd", ".dsk", ".dsc"]:
            src = BASELINE_DSS.with_suffix(ext)
            if src.exists():
                shutil.copy2(src, chunk_dss.with_suffix(ext))

    # -- 4b-iv: Worker function ----------------------------------------
    def _write_chunk(tag):
        """Overlay cached baseline with chunk values and write to DSS."""
        chunk_dss = COMPILED_DIR / f"ProductB_SV_{tag}.dss"
        dss_str = _get_dss_str(chunk_dss)
        lookup = chunk_lookups[tag]

        n_paths_written = 0
        n_svs_written = set()

        with dss_io.open_dss(dss_str, version=6, catalog_flag=False,
                             use_junction=False) as dss_out:
            for pk, ym_vals in lookup.items():
                if pk not in baseline_bucket:
                    continue
                for pathname in baseline_bucket[pk]:
                    if pathname not in baseline_ts_cache:
                        continue
                    cached = baseline_ts_cache[pathname]
                    ts_vals = cached['values'].copy()

                    eom_years  = cached['eom_years']
                    eom_months = cached['eom_months']
                    modified = False
                    for i in range(len(ts_vals)):
                        v = ym_vals.get((eom_years[i], eom_months[i]))
                        if v is not None:
                            ts_vals[i] = v
                            modified = True

                    if modified:
                        tsc = TimeSeriesContainer()
                        tsc.pathname      = pathname
                        tsc.startDateTime = cached['start_dt']
                        tsc.numberValues  = len(ts_vals)
                        tsc.units         = cached['units']
                        tsc.type          = cached['type']
                        tsc.interval      = cached['interval']
                        tsc.values        = ts_vals
                        safe_write_ts(dss_out, pathname, tsc)
                        n_paths_written += 1

                n_svs_written.add(pk)

        return tag, n_paths_written, len(n_svs_written)

    # -- 4b-v: Execute (sequential) ------------------------------------
    # NOTE: DSS writing must be sequential. The HEC-DSS C library uses
    # global state that is not thread-safe; concurrent writes to separate
    # files can corrupt DSS internal pointers.
    tags_to_write = sorted(chunk_lookups)
    print(f"  Writing {len(tags_to_write)} chunks (sequential) ...")
    for tag in tags_to_write:
        t_one = time.time()
        tag, n_paths, n_svs = _write_chunk(tag)
        dss_chunk_paths[tag] = COMPILED_DIR / f"ProductB_SV_{tag}.dss"
        print(f"  {tag}:  {n_paths:>6,} DSS paths written  |  "
              f"{n_svs:>5,} (B,C)  ({time.time()-t_one:.1f}s)")

    if use_junction:
        dss_io.remove_junction()

    print("-" * 72)
    print(f"  DSS compilation completed in {time.time()-t0_dss:.1f}s")
    print()
else:
    print("Step 4b: Skipped (--skip-dss)")
    print()


# ==================================================================
# STEP 5 -- SV coverage by chunk
# ==================================================================
print("Step 5: Writing SV coverage diagnostics ...")

# Build a DataFrame showing which SVs appear in which chunks
all_compiled_svs = set()
for sv_set in chunk_sv_sets.values():
    all_compiled_svs |= sv_set

coverage_rows = []
for pk in sorted(all_compiled_svs):
    row = {"Part_B": pk[0], "Part_C": pk[1]}
    for tag in ACTIVE_TAGS:
        row[tag] = pk in chunk_sv_sets.get(tag, set())
    # Source module
    source_mod = "constant_rept" if pk in const_rept_filled else "unknown"
    for label, svs in module_svs.items():
        if pk in svs:
            source_mod = label
            break
    row["Source_Module"] = source_mod
    coverage_rows.append(row)

coverage_df = pd.DataFrame(coverage_rows)
# Flag any SVs that are not in ALL active chunks
bool_cols = [c for c in ACTIVE_TAGS if c in coverage_df.columns]
if bool_cols:
    coverage_df["All_Chunks"] = coverage_df[bool_cols].all(axis=1)

# Add per-chunk month counts and a summary flag for unexpected counts
for tag in ACTIVE_TAGS:
    sv_counts = chunk_sv_month_counts.get(tag, pd.Series(dtype=int))
    coverage_df[f"months_{tag}"] = coverage_df.apply(
        lambda r: sv_counts.get((r["Part_B"], r["Part_C"]), 0), axis=1
    )
month_cols = [f"months_{t}" for t in ACTIVE_TAGS if f"months_{t}" in coverage_df.columns]
if month_cols:
    coverage_df["Min_Months"] = coverage_df[month_cols].replace(0, pd.NA).min(axis=1)
    coverage_df["Max_Months"] = coverage_df[month_cols].replace(0, pd.NA).max(axis=1)
    coverage_df["Months_Uniform"] = coverage_df["Min_Months"] == coverage_df["Max_Months"]

fp = OUTPUT_DIR / "sv_coverage_by_chunk.csv"
coverage_df.to_csv(fp, index=False)
incomplete = coverage_df[~coverage_df.get("All_Chunks", True)].shape[0]
non_uniform = int((~coverage_df.get("Months_Uniform", pd.Series([True]*len(coverage_df)))).sum())
not_1200 = int((coverage_df.get("Min_Months", pd.Series([1200]*len(coverage_df))) != 1200).sum())
print(f"  {fp.name:45s}  {len(coverage_df):>6,} SVs")
if incomplete > 0:
    print(f"  WARNING: {incomplete} SVs are missing from at least one chunk!")
if not_1200 > 0:
    print(f"  WARNING: {not_1200} SVs have != 1200 months in at least one chunk!")
if non_uniform > 0:
    print(f"  WARNING: {non_uniform} SVs have different month counts across chunks!")
print()


# ==================================================================
# STEP 6 -- Inventory diagnostic CSVs
# ==================================================================
print("Step 6: Writing inventory diagnostics ...")

# inventory_expected_modified.csv
inv_mod_rows = expected_svs[
    expected_svs.apply(lambda r: (r["Part_B"], r["Part_C"]) in inv_expected_modified, axis=1)
].copy()
inv_mod_rows = inv_mod_rows.sort_values(["Module", "Part_B", "Part_C"]).reset_index(drop=True)
fp = OUTPUT_DIR / "inventory_expected_modified.csv"
inv_mod_rows.to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(inv_mod_rows):>6,} records")

# inventory_expected_missing.csv
inv_miss_rows = expected_svs[
    expected_svs.apply(lambda r: (r["Part_B"], r["Part_C"]) in inv_expected_missing, axis=1)
].copy()
inv_miss_rows = inv_miss_rows.sort_values(["Module", "Part_B", "Part_C"]).reset_index(drop=True)
fp = OUTPUT_DIR / "inventory_expected_missing.csv"
inv_miss_rows.to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(inv_miss_rows):>6,} records")

# inventory_constant_rept.csv
inv_cr_rows = inv_const_rept_all.copy()
inv_cr_rows["Auto_Filled"] = inv_cr_rows.apply(
    lambda r: (r["Part_B"], r["Part_C"]) in const_rept_filled, axis=1
)
inv_cr_rows["From_CSV"] = inv_cr_rows.apply(
    lambda r: (r["Part_B"], r["Part_C"]) in all_module_keys, axis=1
)
inv_cr_rows = inv_cr_rows.sort_values(["Input_Category", "Part_B", "Part_C"]).reset_index(drop=True)
fp = OUTPUT_DIR / "inventory_constant_rept.csv"
inv_cr_rows.to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(inv_cr_rows):>6,} records")

# inventory_skipped_missing.csv
fp = OUTPUT_DIR / "inventory_skipped_missing.csv"
inv_skipped_missing.sort_values(["Input_Category", "Part_B", "Part_C"]).reset_index(
    drop=True
).to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(inv_skipped_missing):>6,} records")

# inventory_skipped_not_in_dcr.csv
fp = OUTPUT_DIR / "inventory_skipped_not_in_dcr.csv"
inv_skipped_not_dcr.sort_values(["Input_Category", "Part_B", "Part_C"]).reset_index(
    drop=True
).to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(inv_skipped_not_dcr):>6,} records")

# inventory_unexpected.csv
inv_unexp_rows = []
for mk in sorted(inv_unexpected):
    source = "unknown"
    for label, svs in module_svs.items():
        if mk in svs:
            source = label
            break
    inv_unexp_rows.append({"Part_B": mk[0], "Part_C": mk[1], "Module": source})
inv_unexp_df = pd.DataFrame(inv_unexp_rows)
fp = OUTPUT_DIR / "inventory_unexpected.csv"
inv_unexp_df.to_csv(fp, index=False)
print(f"  {fp.name:45s}  {len(inv_unexp_df):>6,} records")
print()


# ==================================================================
# STEP 7 -- Product B comparisons (vs Product A & vs CalSim Base)
# ==================================================================
if not CLI_ARGS.skip_comparison:
    print("Step 7: Comparing Product B chunks against references ...")
    t0_cmp = time.time()

    # -- Extract SV units from baseline DSS (shared by both comparisons) --
    try:
        _bt_cache = baseline_ts_cache
    except NameError:
        _bt_cache = {}
    units_map = {}
    for _pn, _cached in _bt_cache.items():
        _pk = path_key(_pn)
        if _pk not in units_map:
            units_map[_pk] = _cached["units"]
    _missing_unit_keys = all_compiled_svs - set(units_map.keys())
    if _missing_unit_keys:
        with dss_io.open_dss(BASELINE_DSS, version=6, catalog_flag=False) as _dss_u:
            for _pk in sorted(_missing_unit_keys):
                if _pk not in baseline_bucket:
                    continue
                try:
                    _ts_u = _dss_u.read_ts(
                        baseline_bucket[_pk][0], trim_missing=False
                    )
                    units_map[_pk] = _ts_u.units
                except Exception:
                    units_map[_pk] = "UNKNOWN"
    print(f"  Extracted units for {len(units_map):,} SVs")

    fig_root = OUTPUT_DIR / "figures"
    fig_root.mkdir(exist_ok=True)

    # -- 7a: Product A comparison (WY 1972-2021) --
    if not PRODUCT_A_DSS.exists():
        print("  WARNING: Product A DSS not found, skipping Product A comparison.")
        print(f"    Expected: {PRODUCT_A_DSS}")
    else:
        print()
        print("  7a: Comparing against Product A (WY 1972-2021) ...")
        pa_means = read_product_a_monthly_means(
            PRODUCT_A_DSS, baseline_bucket, all_compiled_svs
        )
        print(f"    Product A: {len(pa_means):,} (B,C) with data in comparison window")
        pa_series = read_product_a_monthly_series(
            PRODUCT_A_DSS, all_compiled_svs
        )

        pb_means_pa = _compute_pb_chunk_means(
            compiled_chunks, ACTIVE_TAGS,
            start_ym=197110, end_ym=202109,
        )

        cmp_a = _build_comparison_df(
            pa_means, pb_means_pa, all_compiled_svs, ACTIVE_TAGS
        )

        # Write CSV (rename Ref_mean -> Product_A_mean for backward compat)
        cmp_a_out = cmp_a.rename(columns={"Ref_mean": "Product_A_mean"})
        fp = OUTPUT_DIR / "product_b_vs_a_comparison.csv"
        cmp_a_out.to_csv(fp, index=False)
        n_svs_compared = cmp_a_out.groupby(["Part_B", "Part_C"]).ngroups
        print(f"    {fp.name:45s}  {n_svs_compared:>6,} SVs compared")

        print("    Generating Product A comparison figures ...")
        try:
            _generate_comparison_figures(
                cmp_a, fig_root / "vs_product_a", "Product A",
                ACTIVE_TAGS, units_map, skip_climatology=False,
                ref_series_by_pk=pa_series,
                compiled_chunks=compiled_chunks,
            )
        except ImportError:
            print("    WARNING: matplotlib not available, skipping figures.")

    # -- 7b: CalSim baseline comparison (full 1921-2021) --
    print()
    print("  7b: Comparing against CalSim baseline (WY 1922-2021) ...")
    cb_means = read_calsim_base_monthly_means(
        baseline_bucket, all_compiled_svs, CB_START, CB_END
    )
    print(f"    CalSim Base: {len(cb_means):,} (B,C) with data in comparison window")
    cb_series = read_calsim_base_monthly_series(
        baseline_bucket, all_compiled_svs, CB_START, CB_END
    )

    pb_means_cb = _compute_pb_chunk_means(
        compiled_chunks, ACTIVE_TAGS,
        start_ym=192110, end_ym=202109,
    )

    cmp_b = _build_comparison_df(
        cb_means, pb_means_cb, all_compiled_svs, ACTIVE_TAGS
    )

    # Write CSV
    cmp_b_out = cmp_b.rename(columns={"Ref_mean": "CalSim_Base_mean"})
    fp = OUTPUT_DIR / "product_b_vs_calsim_base_comparison.csv"
    cmp_b_out.to_csv(fp, index=False)
    n_svs_compared = cmp_b_out.groupby(["Part_B", "Part_C"]).ngroups
    print(f"    {fp.name:45s}  {n_svs_compared:>6,} SVs compared")

    print("    Generating CalSim Base comparison figures ...")
    try:
        _generate_comparison_figures(
            cmp_b, fig_root / "vs_calsim_base", "CalSim Base",
            ACTIVE_TAGS, units_map, skip_climatology=False,
            ref_series_by_pk=cb_series,
            compiled_chunks=compiled_chunks,
        )
    except ImportError:
        print("    WARNING: matplotlib not available, skipping figures.")

    print(f"\n  Comparisons completed in {time.time()-t0_cmp:.1f}s")
    print()
else:
    print("Step 7: Skipped (--skip-comparison)")
    print()


# ==================================================================
# STEP 8 -- Compilation summary
# ==================================================================
print("Step 8: Writing compilation summary ...")

n_total_compiled = sum(cs["n_rows"] for cs in chunk_stats.values())
n_total_svs = len(all_compiled_svs)

lines = [
    "=" * 65,
    "  Product B (1000-year) -- Compilation Summary",
    "=" * 65,
    "",
    f"  Baseline DSS:  {BASELINE_DSS.name}",
    f"  Inventory:     {INVENTORY_XLSX.name}",
    f"  Source CSVs:   {COMPILED_CSV}",
    f"  CSV output:    {COMPILED_DIR}",
    f"  DSS output:    {COMPILED_DIR}",
    "",
    f"  Chunks compiled: {len(chunk_stats)} / {N_CHUNKS}",
    f"  Total unique (Part B, Part C): {n_total_svs:>6,}",
    "",
    "--- Module Contributions ---",
    "",
]

for label in MODULE_CONFIG_B:
    if label in module_svs:
        n = len(module_svs[label])
        nf = module_files.get(label, 0)
        lines.append(f"    {label:40s}  {n:>5} (B,C)  |  {nf:>4} files")
    else:
        lines.append(f"    {label:40s}  (not found)")

lines += [
    "",
    "--- Constant/Rept ---",
    "",
    f"    Total Constant/Rept in inventory:    {len(const_rept_all_keys):>6,}",
    f"    Auto-filled by script:               {len(const_rept_filled):>6,}",
    f"    Already in module CSVs:              {len(const_rept_all_keys & all_module_keys):>6,}",
    "",
    "--- Per-Chunk Statistics ---",
    "",
]

for tag in ACTIVE_TAGS:
    if tag in chunk_stats:
        cs = chunk_stats[tag]
        lines.append(f"    {tag}:  {cs['n_rows']:>10,} rows  |  "
                      f"{cs['n_svs']:>5,} (B,C)  |  "
                      f"modules: {', '.join(cs['modules'][:5])}"
                      f"{'...' if len(cs['modules']) > 5 else ''}")
    else:
        lines.append(f"    {tag}:  (not compiled)")

lines += [
    "",
    "--- Inventory Flags ---",
    "",
    f"    Total inventory SVs:                 {len(inventory_df):>6,}",
    f"    Skipped -- Missing=T:                {len(skipped_missing_keys):>6,}",
    f"    Skipped -- Used_in_DCR=F:            {len(skipped_not_dcr_keys):>6,}",
    "",
    "--- Inventory Cross-Reference ---",
    "",
    f"    Expected from modules:               {len(expected_keys):>6,}",
    f"    Expected & provided:                 {len(inv_expected_modified):>6,}",
    f"    Expected but MISSING:                {len(inv_expected_missing):>6,}",
    f"    Provided but NOT in inventory:       {len(inv_unexpected):>6,}",
    "",
    "--- Categories Without Product B Modules ---",
    "",
]

for cat in sorted(CATEGORIES_WITHOUT_PRODUCT_B):
    cnt_active = len(inventory_df[
        (inventory_df["Input_Category"] == cat)
        & (~inventory_df["Missing"])
        & (inventory_df["Used_in_DCR"])
        & (~inventory_df["Constant_Rept"])
    ])
    cnt_cr = len(inventory_df[
        (inventory_df["Input_Category"] == cat)
        & (inventory_df["Constant_Rept"] == True)
        & (~inventory_df["Missing"])
        & (inventory_df["Used_in_DCR"])
    ])
    cnt_total = len(inventory_df[inventory_df["Input_Category"] == cat])
    lines.append(f"    {cat:40s}  {cnt_total:>5} total  "
                 f"({cnt_cr} const/rept, {cnt_active} other)")

lines += [
    "",
    "--- Modules Skipped (no data) ---",
    "",
]
if missing_modules:
    for m in missing_modules:
        lines.append(f"    {m}")
else:
    lines.append("    (none)")

lines += [
    "",
    "--- Source CSV Modification Dates ---",
    "",
]
import datetime as _dt
for label in MODULE_CONFIG_B:
    if label not in module_src_paths:
        continue
    lines.append(f"    {label}")
    for csv_path in sorted(module_src_paths[label], key=lambda p: p.name):
        mtime = csv_path.stat().st_mtime
        mtime_str = _dt.datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
        lines.append(f"      {csv_path.name:60s}  {mtime_str}")
    lines.append("")

lines += ["=" * 65]

summary_path = OUTPUT_DIR / "compilation_summary.txt"
summary_path.write_text("\n".join(lines), encoding="utf-8")

print(f"\n  {summary_path.name}")
print()
for line in lines:
    print(line)

print(f"\nDone.  Output in: {OUTPUT_DIR}")
